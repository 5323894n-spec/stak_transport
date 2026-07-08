# -*- coding: utf-8 -*-
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

DAY_TYPE = "weekday"
CRITICAL = "\u043a\u0440\u0438\u0442\u0438\u0447\u043d\u043e"
HEADER_PREFIX = "\u0420\u0430\u0441\u043f\u0438\u0441\u0430\u043d\u0438\u0435 \u043c\u0430\u0440\u0448\u0440\u0443\u0442\u0430"
EXPORT_HEADERS = (
    "\u041c\u0430\u0440\u0448\u0440\u0443\u0442",
    "\u0422\u0438\u043f \u0434\u043d\u044f",
    "\u0412\u044b\u0445\u043e\u0434",
    "\u0421\u043c\u0435\u043d\u0430",
    "\u0420\u0435\u0439\u0441",
    "\u041d\u0430\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u0438\u0435",
)


def make_client(tmp_path, monkeypatch):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    db.DB_PATH = str(tmp_path / "atp-test.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
    finally:
        con.close()

    client = TestClient(app)
    token = client.post("/api/login", json={"username": "admin", "password": "admin"}).json()["token"]
    client.headers.update({"Authorization": "Bearer " + token})
    return client


def create_route(client):
    payload = {
        "number": "12",
        "name": "Center - Station",
        "comm_type": "city",
        "start_point": "Center",
        "end_point": "Station",
        "length_km": 11.5,
        "length_back_km": 10.8,
        "trip_time_min": 35,
        "trip_time_back_min": 38,
        "interval_min": 12,
        "outputs_count": 3,
        "bus_types": "large",
        "work_days": "daily",
        "version": 1,
        "active": 1,
    }
    response = client.post("/api/refs/routes", json=payload)
    assert response.status_code == 200, response.text
    return response.json()["id"]


def add_trip(client, route_id, **overrides):
    payload = {
        "route_id": route_id,
        "day_type": DAY_TYPE,
        "output_number": 1,
        "shift_number": 1,
        "trip_number": 1,
        "direction": "forward",
        "dep_time": "06:00",
        "arr_time": "06:35",
        "distance_km": 11.5,
        "break_after_min": 6,
        "break_type": "",
    }
    payload.update(overrides)
    response = client.post("/api/trips", json=payload)
    assert response.status_code == 200, response.text
    return payload


def test_generate_schedule_returns_summary(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)

    response = client.post("/api/trips/generate", json={
        "route_id": route_id,
        "day_type": DAY_TYPE,
        "outputs": 3,
        "first_dep": "05:30",
        "last_dep": "08:30",
        "trip_time": 35,
        "trip_time_back": 38,
        "interval": 12,
        "rest_min": 6,
        "lunch_min": 40,
        "distance": 11.5,
        "distance_back": 10.8,
        "mode": "interval",
    })
    assert response.status_code == 200, response.text
    assert response.json()["trips"] > 0

    summary = client.get(f"/api/routes/{route_id}/schedule-summary?day_type={DAY_TYPE}")
    assert summary.status_code == 200, summary.text
    data = summary.json()
    assert data["trips_count"] > 0
    assert data["outputs_count"] == 3
    assert data["bus_need"] == 3
    assert data["driver_need"] >= 3
    assert data["distance_km"] > 0
    assert data["first_dep"] == "05:30"


def test_route_check_reports_overlap_and_recommendation(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)
    add_trip(client, route_id, trip_number=1, dep_time="06:00", arr_time="06:40")
    add_trip(client, route_id, trip_number=2, dep_time="06:30", arr_time="07:10")

    response = client.get(f"/api/routes/{route_id}/check?day_type={DAY_TYPE}")
    assert response.status_code == 200, response.text
    problems = response.json()["problems"]
    assert any(p["severity"] == CRITICAL and p["kind"] == "overlap" for p in problems)
    assert all("recommendation" in p for p in problems)


def test_route_check_reports_duplicate_trip_number(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)
    add_trip(client, route_id, trip_number=1, dep_time="06:00", arr_time="06:35")
    add_trip(client, route_id, trip_number=1, dep_time="06:45", arr_time="07:20")

    response = client.get(f"/api/routes/{route_id}/check?day_type={DAY_TYPE}")
    assert response.status_code == 200, response.text
    assert any(p["kind"] == "duplicate_trip_number" for p in response.json()["problems"])


def test_bulk_shift_moves_trip_times(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)
    add_trip(client, route_id, dep_time="06:00", arr_time="06:35")

    response = client.post("/api/trips/bulk-shift", json={
        "route_id": route_id,
        "day_type": DAY_TYPE,
        "minutes": 15,
    })
    assert response.status_code == 200, response.text
    assert response.json()["updated"] == 1

    trips = client.get(f"/api/trips?route_id={route_id}&day_type={DAY_TYPE}").json()["items"]
    assert trips[0]["dep_time"] == "06:15"
    assert trips[0]["arr_time"] == "06:50"


def test_renumber_orders_trips_inside_each_output(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)
    add_trip(client, route_id, output_number=1, trip_number=9, dep_time="07:00", arr_time="07:35")
    add_trip(client, route_id, output_number=1, trip_number=8, dep_time="06:00", arr_time="06:35")

    response = client.post("/api/trips/renumber", json={"route_id": route_id, "day_type": DAY_TYPE})
    assert response.status_code == 200, response.text
    assert response.json()["updated"] == 2

    trips = client.get(f"/api/trips?route_id={route_id}&day_type={DAY_TYPE}").json()["items"]
    assert [t["trip_number"] for t in trips] == [1, 2]
    assert [t["dep_time"] for t in trips] == ["06:00", "07:00"]


def test_schedule_export_xlsx(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)
    add_trip(client, route_id, dep_time="06:00", arr_time="06:35")

    response = client.get(f"/api/routes/{route_id}/schedule-export.xlsx?day_type={DAY_TYPE}")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(response.content), read_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    assert values[0][0].startswith(HEADER_PREFIX)
    assert values[1][:6] == EXPORT_HEADERS
    assert values[2][4] == 1



def test_schedule_export_accepts_cyrillic_day_type(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)
    day_type = "\u0431\u0443\u0434\u043d\u0438"
    add_trip(client, route_id, day_type=day_type, dep_time="06:00", arr_time="06:35")

    response = client.get(f"/api/routes/{route_id}/schedule-export.xlsx?day_type={day_type}")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def test_trip_break_after_save_shifts_following_trip_chain(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)
    add_trip(client, route_id, trip_number=1, dep_time="09:15", arr_time="10:00", break_after_min=0)
    add_trip(client, route_id, trip_number=2, dep_time="10:12", arr_time="10:50", break_after_min=6)
    add_trip(client, route_id, trip_number=3, dep_time="11:00", arr_time="11:40", break_after_min=6)

    trips = client.get(f"/api/trips?route_id={route_id}&day_type={DAY_TYPE}").json()["items"]
    first = next(t for t in trips if t["trip_number"] == 1)
    first["break_after_min"] = 40
    first["break_type"] = "обед"

    response = client.post("/api/trips", json=first)
    assert response.status_code == 200, response.text
    assert response.json()["shifted"] == 2

    trips = client.get(f"/api/trips?route_id={route_id}&day_type={DAY_TYPE}").json()["items"]
    by_number = {t["trip_number"]: t for t in trips}
    assert by_number[1]["arr_time"] == "10:00"
    assert by_number[2]["dep_time"] == "10:40"
    assert by_number[2]["arr_time"] == "11:18"
    assert by_number[3]["dep_time"] == "11:28"
    assert by_number[3]["arr_time"] == "12:08"


def test_route_check_reports_break_gap_violation(tmp_path, monkeypatch):
    client = make_client(tmp_path, monkeypatch)
    route_id = create_route(client)

    import app.db as db

    con = db.connect()
    try:
        con.execute(
            """
            INSERT INTO route_trips(route_id,day_type,output_number,shift_number,trip_number,
              direction,dep_time,arr_time,distance_km,break_after_min,break_type)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """,
            (route_id, DAY_TYPE, 1, 1, 1, "forward", "09:15", "10:00", 11.5, 40, "обед"),
        )
        con.execute(
            """
            INSERT INTO route_trips(route_id,day_type,output_number,shift_number,trip_number,
              direction,dep_time,arr_time,distance_km,break_after_min,break_type)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """,
            (route_id, DAY_TYPE, 1, 1, 2, "forward", "10:12", "10:50", 11.5, 6, ""),
        )
        con.commit()
    finally:
        con.close()

    response = client.get(f"/api/routes/{route_id}/check?day_type={DAY_TYPE}")
    assert response.status_code == 200, response.text
    problems = response.json()["problems"]
    assert any(p["kind"] == "break_gap" and p["trip_number"] == 2 for p in problems)
