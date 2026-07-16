# -*- coding: utf-8 -*-
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


@pytest.fixture
def exported_schedule(tmp_path):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    db.DB_PATH = str(tmp_path / "timetable-exports.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
        route_id = con.execute(
            "INSERT INTO routes(number,name,trip_time_min,trip_time_back_min,outputs_count) "
            "VALUES('E1','Export route',10,10,2)"
        ).lastrowid
        stop_ids = []
        for code, name in (("EA", "Export terminal A"), ("EB", "Export terminal B")):
            stop_ids.append(con.execute(
                "INSERT INTO stops(external_code,name,active) VALUES(?,?,1)",
                (code, name),
            ).lastrowid)
        for direction, ordered in (("forward", stop_ids), ("backward", stop_ids[::-1])):
            for sequence, stop_id in enumerate(ordered, 1):
                con.execute(
                    "INSERT INTO route_stops(route_id,direction,sequence,stop_id,"
                    "run_time_sec,dwell_time_sec,is_timing_point) VALUES(?,?,?,?,?,?,1)",
                    (route_id, direction, sequence, stop_id,
                     0 if sequence == 1 else 600, 0),
                )
        con.execute(
            "INSERT INTO day_periods(route_id,day_type,name,start_min,end_min,interval_min,"
            "travel_time_factor,transition_mode,transition_window_min,color,priority,active,"
            "created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (route_id, "будни", "Утро", 360, 381, 10, 1.0, "abrupt", 0,
             "#3b82f6", 0, 1, "2026-07-16", "2026-07-16"),
        )
        con.commit()
    finally:
        con.close()
    client = TestClient(app)
    token = client.post(
        "/api/login", json={"username": "admin", "password": "admin"}
    ).json()["token"]
    client.headers.update({"Authorization": "Bearer " + token})
    preview = client.post(
        f"/api/routes/{route_id}/schedule-generation/preview",
        json={"day_type": "будни", "outputs": 3, "terminal_layover_min": 5},
    )
    assert preview.status_code == 200, preview.text
    applied = client.post(
        f"/api/routes/{route_id}/schedule-generation/apply",
        json={"day_type": "будни", "preview_token": preview.json()["preview_token"]},
    )
    assert applied.status_code == 200, applied.text
    matrix = client.get(
        f"/api/routes/{route_id}/stop-times?day_type=будни"
    ).json()
    return client, route_id, stop_ids[0], matrix["trips"][0]["trip_id"]


def test_route_matrix_and_trip_exports_have_stop_headers(exported_schedule):
    client, route_id, _, trip_id = exported_schedule
    response = client.get(
        f"/api/routes/{route_id}/stop-times/export.xlsx?day_type=будни"
    )
    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    assert sheet["A1"].value.startswith("Поостановочное расписание")
    assert "Остановка" in [cell.value for cell in sheet[3]]
    assert sheet.freeze_panes == "B4"
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True

    response = client.get(f"/api/trips/{trip_id}/stop-times/export.xlsx")
    assert response.status_code == 200, response.text
    trip_sheet = load_workbook(io.BytesIO(response.content)).active
    assert trip_sheet["A1"].value.startswith("Лист рейса")
    assert [cell.value for cell in trip_sheet[3]][:4] == [
        "№", "Остановка", "Прибытие", "Отправление"
    ]


def test_stop_pavilion_export_lists_all_departures(exported_schedule):
    client, _, stop_id, _ = exported_schedule
    response = client.get(
        f"/api/stops/{stop_id}/timetable.xlsx?day_type=будни"
    )
    assert response.status_code == 200, response.text
    sheet = load_workbook(io.BytesIO(response.content)).active
    assert sheet["A1"].value.startswith("Расписание остановки")
    assert sheet.max_row > 3
    assert "Маршрут" in [cell.value for cell in sheet[3]]


def test_trip_export_preserves_manual_adjustment_fill(exported_schedule):
    import app.db as db

    client, _, _, trip_id = exported_schedule
    con = db.connect()
    try:
        route_stop_id = con.execute(
            "SELECT route_stop_id FROM trip_stop_times WHERE trip_id=? "
            "ORDER BY sequence LIMIT 1",
            (trip_id,),
        ).fetchone()[0]
    finally:
        con.close()
    response = client.patch(
        f"/api/trips/{trip_id}/stop-times/{route_stop_id}",
        json={
            "departure_time": "06:01",
            "strategy": "shift_following",
            "reason": "Проверка отметки Excel",
        },
    )
    assert response.status_code == 200, response.text
    response = client.get(f"/api/trips/{trip_id}/stop-times/export.xlsx")
    assert response.status_code == 200, response.text
    sheet = load_workbook(io.BytesIO(response.content)).active
    color = sheet["A4"].fill.fgColor.rgb
    assert color in ("00FFF2CC", "FFF2CC")
