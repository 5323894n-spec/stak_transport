# -*- coding: utf-8 -*-
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook


DATE = "2026-07-06"
DAY_TYPE = "будни"


def make_client(tmp_path):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    db.DB_PATH = str(tmp_path / "atp-summary.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
    finally:
        con.close()

    client = TestClient(app)
    token = client.post("/api/login", json={"username": "admin", "password": "admin"}).json()["token"]
    client.headers.update({"Authorization": "Bearer " + token})
    return client


def seed_clean_summary_source():
    import app.db as db

    con = db.connect()
    try:
        driver_id = con.execute(
            "INSERT INTO drivers(tab_number,fio,status,default_schedule) VALUES(?,?,?,?)",
            ("7001", "Summary Driver", "работает", "2/2"),
        ).lastrowid
        bus_id = con.execute(
            "INSERT INTO buses(garage_number,plate,status,fuel_rate,assigned_driver_id) VALUES(?,?,?,?,?)",
            ("G-701", "А701АА69", "исправен", 35.0, driver_id),
        ).lastrowid
        con.execute("UPDATE drivers SET assigned_bus_id=? WHERE id=?", (bus_id, driver_id))
        route_id = con.execute(
            "INSERT INTO routes(number,name,start_point,end_point,length_km,length_back_km,active,version) "
            "VALUES(?,?,?,?,?,?,?,?)",
            ("77", "Вокзал - Центр", "Вокзал", "Центр", 12.0, 12.0, 1, 3),
        ).lastrowid
        con.executemany(
            "INSERT INTO route_trips(route_id,day_type,output_number,shift_number,trip_number,direction,dep_time,arr_time,distance_km) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            [
                (route_id, DAY_TYPE, 1, 1, 1, "прямое", "06:00", "06:45", 12.0),
                (route_id, DAY_TYPE, 1, 1, 2, "обратное", "07:00", "07:45", 12.0),
            ],
        )
        con.execute(
            "INSERT INTO roster_assignments(driver_id,date,route_id,day_type,output_number,shift_number,trip_from,trip_to,"
            "start_time,end_time,hours,distance_km,trips_count) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (driver_id, DATE, route_id, DAY_TYPE, 1, 1, 1, 2, "06:00", "07:45", 1.5, 24.0, 2),
        )
        con.commit()
        return {"driver_id": driver_id, "bus_id": bus_id, "route_id": route_id}
    finally:
        con.close()


def seed_missing_assignment_source():
    import app.db as db

    con = db.connect()
    try:
        route_id = con.execute(
            "INSERT INTO routes(number,name,active,version) VALUES(?,?,?,?)",
            ("88", "No assignment route", 1, 1),
        ).lastrowid
        con.execute(
            "INSERT INTO route_trips(route_id,day_type,output_number,shift_number,trip_number,direction,dep_time,arr_time,distance_km) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            (route_id, DAY_TYPE, 1, 1, 1, "прямое", "08:00", "08:30", 7.0),
        )
        con.commit()
        return route_id
    finally:
        con.close()


def test_summary_schema_and_history_endpoint_exist(tmp_path):
    client = make_client(tmp_path)

    import app.db as db

    con = db.connect()
    try:
        tables = {row["name"] for row in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    finally:
        con.close()

    assert {"summary_schedules", "summary_schedule_lines", "summary_schedule_errors"}.issubset(tables)
    response = client.get("/api/summary-schedules")
    assert response.status_code == 200, response.text
    assert response.json() == {"items": []}


def test_generate_summary_creates_snapshot_lines_and_views(tmp_path):
    client = make_client(tmp_path)
    ctx = seed_clean_summary_source()

    response = client.post("/api/summary-schedules/generate", json={
        "date_from": DATE,
        "date_to": DATE,
        "route_ids": [ctx["route_id"]],
        "include_inactive": False,
        "comment": "test generation",
    })

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["summary"]["trips_count"] == 2
    assert data["summary"]["routes_count"] == 1
    assert data["summary"]["runs_count"] == 1
    assert data["summary"]["vehicles_count"] == 1
    assert data["summary"]["drivers_count"] == 1
    assert len(data["lines"]) == 2
    assert data["lines"][0]["route_number"] == "77"
    assert data["lines"][0]["driver_name"] == "Summary Driver"
    assert data["lines"][0]["garage_number"] == "G-701"
    assert data["views"]["by_outputs"][0]["trips_count"] == 2
    assert data["views"]["by_time"][0]["event"] in {"выезд из парка", "начало рейса"}

    detail = client.get(f"/api/summary-schedules/{data['summary']['id']}")
    assert detail.status_code == 200, detail.text
    assert detail.json()["summary"]["id"] == data["summary"]["id"]


def test_summary_errors_detect_missing_driver_and_bus(tmp_path):
    client = make_client(tmp_path)
    seed_missing_assignment_source()

    response = client.post("/api/summary-schedules/generate", json={"date_from": DATE, "date_to": DATE})

    assert response.status_code == 200, response.text
    errors = response.json()["errors"]
    assert any(e["level"] == "Критическая ошибка" and "водитель" in e["message"].lower() for e in errors)
    assert any(e["level"] == "Критическая ошибка" and "автобус" in e["message"].lower() for e in errors)
    assert response.json()["summary"]["errors_count"] >= 2
    assert any(line["error_flag"] == 1 for line in response.json()["lines"])


def test_summary_recheck_returns_same_summary_id(tmp_path):
    client = make_client(tmp_path)
    seed_clean_summary_source()
    generated = client.post("/api/summary-schedules/generate", json={"date_from": DATE, "date_to": DATE})
    summary_id = generated.json()["summary"]["id"]

    recheck = client.post(f"/api/summary-schedules/{summary_id}/check")

    assert recheck.status_code == 200, recheck.text
    assert recheck.json()["summary"]["id"] == summary_id


def test_summary_export_xlsx_has_required_sheets_and_format(tmp_path):
    client = make_client(tmp_path)
    seed_clean_summary_source()
    generated = client.post("/api/summary-schedules/generate", json={"date_from": DATE, "date_to": DATE})
    summary_id = generated.json()["summary"]["id"]

    response = client.get(f"/api/summary-schedules/{summary_id}/export.xlsx")

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == [
        "Титульный лист", "Сводка", "По маршрутам", "По выходам", "По водителям",
        "По автобусам", "По времени", "Ошибки", "Исходные данные",
    ]
    assert wb["Титульный лист"]["A2"].value == "СВОДНОЕ РАСПИСАНИЕ"
    assert wb["По маршрутам"].freeze_panes == "A2"
    assert wb["По маршрутам"].auto_filter.ref is not None
    assert wb["По маршрутам"]["A1"].font.bold is True


def test_summary_order_creation_blocks_critical_errors(tmp_path):
    client = make_client(tmp_path)
    seed_missing_assignment_source()
    generated = client.post("/api/summary-schedules/generate", json={"date_from": DATE, "date_to": DATE})
    summary_id = generated.json()["summary"]["id"]

    response = client.post(f"/api/summary-schedules/{summary_id}/order", json={"regenerate": True})

    assert response.status_code == 409
    assert "Критические" in response.text


def test_summary_order_creation_creates_order_lines(tmp_path):
    client = make_client(tmp_path)
    seed_clean_summary_source()
    generated = client.post("/api/summary-schedules/generate", json={"date_from": DATE, "date_to": DATE})
    summary_id = generated.json()["summary"]["id"]

    response = client.post(f"/api/summary-schedules/{summary_id}/order", json={"regenerate": True})

    assert response.status_code == 200, response.text
    assert response.json()["lines"] == 1
    order = client.get(f"/api/orders?date={DATE}")
    assert order.status_code == 200, order.text
    lines = order.json()["lines"]
    assert len(lines) == 1
    assert lines[0]["route_number"] == "77"
    assert lines[0]["driver_id"] is not None
    assert lines[0]["bus_id"] is not None