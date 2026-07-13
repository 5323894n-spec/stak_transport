# -*- coding: utf-8 -*-
import io
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook


DATE = "2026-07-06"
MONTH = "2026-07"
DAY_TYPE = "будни"


def make_client(tmp_path):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    db.DB_PATH = str(tmp_path / "atp-roster-excel.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
        con.execute(
            "INSERT INTO settings(key,value) VALUES('org_name',?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            ("Тестовое АТП",),
        )
        con.commit()
    finally:
        con.close()

    client = TestClient(app)
    token = client.post("/api/login", json={"username": "admin", "password": "admin"}).json()["token"]
    client.headers.update({"Authorization": "Bearer " + token})
    return client


def create_driver(client):
    response = client.post("/api/refs/drivers", json={
        "tab_number": "9101",
        "fio": "Иванов Иван Иванович",
        "status": "работает",
        "default_schedule": "2/2",
    })
    assert response.status_code == 200, response.text
    return response.json()["id"]


def create_route_with_two_shifts(client):
    response = client.post("/api/refs/routes", json={
        "number": "22",
        "name": "Депо - Центр",
        "comm_type": "городское",
        "start_point": "Депо",
        "end_point": "Центр",
        "length_km": 35,
        "length_back_km": 40,
        "trip_time_min": 90,
        "trip_time_back_min": 90,
        "interval_min": 15,
        "outputs_count": 1,
        "bus_types": "большой",
        "work_days": "ежедневно",
        "version": 1,
        "active": 1,
    })
    assert response.status_code == 200, response.text
    route_id = response.json()["id"]
    trips = [
        (1, 1, 1, "06:00", "09:00", 15.0),
        (1, 1, 2, "09:15", "13:30", 20.0),
        (1, 2, 1, "14:00", "18:00", 20.0),
        (1, 2, 2, "18:15", "22:00", 20.0),
    ]
    for output, shift, trip_no, dep, arr, dist in trips:
        response = client.post("/api/trips", json={
            "route_id": route_id,
            "day_type": DAY_TYPE,
            "output_number": output,
            "shift_number": shift,
            "trip_number": trip_no,
            "direction": "прямое" if trip_no % 2 else "обратное",
            "dep_time": dep,
            "arr_time": arr,
            "distance_km": dist,
            "break_after_min": 0,
            "break_type": "",
        })
        assert response.status_code == 200, response.text
    return route_id


def post_assignment(client, driver_id, route_id, shift_number):
    response = client.post("/api/roster/assignment", json={
        "driver_id": driver_id,
        "date": DATE,
        "route_id": route_id,
        "output_number": 1,
        "shift_number": shift_number,
        "trip_from": 1,
        "trip_to": 2,
        "comment": "",
    })
    assert response.status_code == 200, response.text
    return response.json()


def all_text(ws):
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )


def test_roster_export_xlsx_is_editable_monthly_schedule(tmp_path):
    client = make_client(tmp_path)
    driver_id = create_driver(client)
    route_id = create_route_with_two_shifts(client)
    post_assignment(client, driver_id, route_id, 1)
    post_assignment(client, driver_id, route_id, 2)

    response = client.get(f"/api/roster/export.xlsx?month={MONTH}")

    assert response.status_code == 200, response.text
    assert "roster_2026-07.xlsx" in response.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(response.content), data_only=False)
    assert wb.sheetnames == ["График месяца", "Назначения", "Проверки 424", "Справочники"]

    month_sheet = wb["График месяца"]
    assert month_sheet["A1"].value == "Тестовое АТП"
    assert month_sheet["A2"].value == "ГРАФИК РАБОТЫ ВОДИТЕЛЕЙ"
    assert "Июль 2026" in month_sheet["A3"].value
    month_text = all_text(month_sheet)
    assert "Иванов Иван Иванович" in month_text
    assert "№22" in month_text
    assert "06:00-13:30" in month_text
    assert "14:00-22:00" in month_text
    assert "ПЗВ18м" in month_text
    assert "ПЗВ 18 мин" not in month_text
    driver_row = next(
        row for row in range(8, month_sheet.max_row + 1)
        if month_sheet.cell(row=row, column=2).value == "Иванов Иван Иванович"
    )
    assert month_sheet.row_dimensions[driver_row].height <= 70
    assert month_sheet.column_dimensions["C"].width <= 11.5

    assignments = wb["Назначения"]
    headers = [assignments.cell(row=4, column=col).value for col in range(1, 17)]
    assert headers == [
        "Дата", "Таб.№", "Водитель", "Статус дня", "Маршрут", "Выход", "Смена", "Рейсы",
        "Начало", "Окончание", "ПЗВ, мин", "Обед/перерыв, мин", "Линейное время, ч",
        "Рабочее время с ПЗВ, ч", "Ночные часы", "Комментарий",
    ]
    assert assignments.cell(row=5, column=1).value == DATE
    assert assignments.cell(row=5, column=9).value == "06:00"
    assert assignments.cell(row=5, column=10).value == "13:30"
    assert assignments.cell(row=5, column=11).value == 18
    assert assignments.cell(row=5, column=13).value.startswith("=")
    assert assignments.cell(row=5, column=14).value.startswith("=")
    assert assignments.cell(row=5, column=11).fill.fgColor.rgb in {"00FFF2CC", "FFFFF2CC"}
    assert assignments.protection.sheet is False

    warnings = wb["Проверки 424"]
    assert warnings["A1"].value == "Проверки 424"
    assert "Превышение продолжительности смены" in all_text(warnings)

    refs = wb["Справочники"]
    assert "ПЗВ по нормам, мин" in all_text(refs)


def test_roster_page_has_dedicated_roster_excel_button():
    app_js = Path(__file__).resolve().parents[1] / "static" / "app.js"
    text = app_js.read_text(encoding="utf-8")

    assert "График в Excel" in text
    assert "Табель в Excel" in text
    assert "/api/roster/export.xlsx?month=${month}" in text