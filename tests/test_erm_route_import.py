# -*- coding: utf-8 -*-
import io
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


ROUTE_TITLE = 'Маршрут № 1 "Железнодорожный вокзал - Ореховая улица"'


def make_client(tmp_path):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    db.DB_PATH = str(tmp_path / "atp-erm-test.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
    finally:
        con.close()

    client = TestClient(app)
    token = client.post(
        "/api/login", json={"username": "admin", "password": "admin"}
    ).json()["token"]
    client.headers.update({"Authorization": "Bearer " + token})
    return client


def _write_header(ws, row):
    headers = {
        2: "п.п.",
        3: "ID",
        6: "остановочный пункт",
        7: "наименования улиц",
        8: "широта",
        9: "долгота",
        10: "расст.      программн",
        11: "общ       программн",
        12: "время движения между ОП",
        13: "время движения нарастающ",
        14: "несоот",
    }
    for col, value in headers.items():
        ws.cell(row=row, column=col, value=value)


def _write_stop(
    ws, row, seq, stop_id, line_name, direction, stop, street,
    distance_m, cumulative_m, travel, cumulative,
):
    values = {
        2: seq,
        3: stop_id,
        4: line_name,
        5: direction,
        6: stop,
        7: street,
        8: 56.8 + seq / 1000,
        9: 35.9 + seq / 1000,
        10: distance_m,
        11: cumulative_m,
        12: travel,
        13: cumulative,
        14: 0,
        15: distance_m / 1000 if distance_m is not None else None,
        16: cumulative_m / 1000 if cumulative_m is not None else None,
    }
    for col, value in values.items():
        ws.cell(row=row, column=col, value=value)


def build_erm_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("параметры")
    ws.cell(row=2, column=6, value=ROUTE_TITLE)
    ws.cell(row=2, column=8, value="прямое направление")
    _write_header(ws, 3)
    _write_stop(ws, 4, 1, 100, "А001", "на Ореховую", "Железнодорожный вокзал", "Привокзальная площадь", 0, 0, "00:00:00", "00:00:00")
    _write_stop(ws, 5, 2, 101, "А001", "на Ореховую", "Автовокзал", "Улица Коминтерна", 500, 500, "00:03:00", "00:03:00")
    _write_stop(ws, 6, 3, 102, "А001", "на Ореховую", "Ореховая улица", "Ореховая улица", 1000, 1500, "00:07:00", "00:10:00")
    ws.cell(row=8, column=8, value="обратное направление")
    _write_header(ws, 9)
    _write_stop(ws, 10, 1, 200, "А001", "Ореховая-Вокзал", "Ореховая улица", "Ореховая улица", 0, 0, "00:00:00", "00:00:00")
    _write_stop(ws, 11, 2, 201, "А001", "Ореховая-Вокзал", "Железнодорожный вокзал", "Привокзальная площадь", 1300, 1300, "00:09:00", "00:09:00")

    ws = wb.create_sheet("из парка")
    ws.cell(row=2, column=6, value=ROUTE_TITLE)
    ws.cell(row=2, column=7, value="рейсы из парка")
    _write_header(ws, 3)
    _write_stop(ws, 4, 1, 300, "А001", "из парка", "Автопарк", "Гаражная улица", 0, 0, "00:00:00", "00:00:00")
    _write_stop(ws, 5, 2, 301, "А001", "из парка", "Железнодорожный вокзал", "Привокзальная площадь", 700, 700, "00:05:00", "00:05:00")

    ws = wb.create_sheet("в парк")
    ws.cell(row=2, column=6, value=ROUTE_TITLE)
    ws.cell(row=2, column=7, value="рейсы в парк")
    _write_header(ws, 3)
    _write_stop(ws, 4, 1, 400, "А001", "в парк", "Ореховая улица", "Ореховая улица", 0, 0, "00:00:00", "00:00:00")
    _write_stop(ws, 5, 2, 401, "А001", "в парк", "Автопарк", "Гаражная улица", 800, 800, "00:06:00", "00:06:00")

    data = io.BytesIO()
    wb.save(data)
    return data.getvalue()


def build_erm_workbook_with_depot_continuation():
    wb = load_workbook(io.BytesIO(build_erm_workbook()))
    ws = wb["из парка"]
    _write_header(ws, 8)
    _write_stop(
        ws, 9, 3, 302, "А001", "из парка", "Площадь", "Советская площадь",
        250, 950, "00:02:30", "00:07:30",
    )
    data = io.BytesIO()
    wb.save(data)
    return data.getvalue()


def build_erm_without_depot_in():
    wb = load_workbook(io.BytesIO(build_erm_workbook()))
    del wb["в парк"]
    data = io.BytesIO()
    wb.save(data)
    return data.getvalue()


def _depot_rows(route_id):
    import app.db as db

    con = db.connect()
    try:
        return db.rows(con.execute(
            "SELECT r.stop_id,r.direction,r.sequence,r.distance_from_prev_km,"
            "r.run_time_day_sec,r.run_time_night_sec,r.source,r.source_detail,"
            "s.external_code,s.name,s.address,s.latitude,s.longitude,"
            "s.source AS stop_source "
            "FROM route_depot_stops r JOIN stops s ON s.id=r.stop_id "
            "WHERE r.route_id=? ORDER BY r.direction,r.sequence",
            (route_id,),
        ))
    finally:
        con.close()


def test_parse_erm_route_workbook_extracts_route_and_sections():
    from app.erm_import import parse_erm_route_workbook

    parsed = parse_erm_route_workbook(build_erm_workbook())

    assert parsed["number"] == "1"
    assert parsed["name"] == "Железнодорожный вокзал - Ореховая улица"
    assert parsed["start_point"] == "Железнодорожный вокзал"
    assert parsed["end_point"] == "Ореховая улица"
    assert parsed["stops"] == "Железнодорожный вокзал, Автовокзал, Ореховая улица"
    assert parsed["stops_back"] == "Ореховая улица, Железнодорожный вокзал"
    assert parsed["length_km"] == 1.5
    assert parsed["length_back_km"] == 1.3
    assert parsed["trip_time_min"] == 10
    assert parsed["trip_time_back_min"] == 9
    assert parsed["details"]["summary"]["route_stops_forward"] == 3
    assert parsed["details"]["summary"]["route_stops_backward"] == 2
    assert parsed["details"]["summary"]["depot_sections"] == 2


def test_erm_route_import_creates_route_and_preserves_details(tmp_path):
    client = make_client(tmp_path)

    response = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["created"] is True
    assert data["updated"] is False
    assert data["route"]["number"] == "1"
    assert data["summary"]["depot_sections"] == 2

    routes = client.get("/api/refs/routes").json()["items"]
    route = next(r for r in routes if r["number"] == "1")
    assert route["length_km"] == 1.5
    assert route["trip_time_min"] == 10
    notes = json.loads(route["notes"])
    assert notes["source"] == "ЭРМ"
    assert notes["details"]["sheets"]["из парка"]["sections"][0]["stops"][1]["stop_name"] == "Железнодорожный вокзал"

    rows = _depot_rows(data["route_id"])
    assert [(row["direction"], row["external_code"]) for row in rows] == [
        ("depot_in", "200"),
        ("depot_in", "401"),
        ("depot_out", "300"),
        ("depot_out", "201"),
    ]
    assert [row["distance_from_prev_km"] for row in rows] == [0, 0.8, 0, 0.7]
    assert [row["run_time_day_sec"] for row in rows] == [0, 360, 0, 300]
    assert [row["run_time_night_sec"] for row in rows] == [0, 360, 0, 300]
    assert {row["source"] for row in rows} == {"erm_import"}
    assert json.loads(rows[2]["source_detail"])["sheet"] == "из парка"
    assert json.loads(rows[2]["source_detail"])["section"] == 1
    assert rows[2]["address"] == "Гаражная улица"
    assert rows[2]["latitude"] == pytest.approx(56.801)
    assert rows[2]["longitude"] == pytest.approx(35.901)
    assert rows[2]["stop_source"] == "erm_import"

    import app.db as db
    con = db.connect()
    try:
        states = db.rows(con.execute(
            "SELECT direction FROM route_depot_section_state "
            "WHERE route_id=? ORDER BY direction",
            (data["route_id"],),
        ))
    finally:
        con.close()
    assert [state["direction"] for state in states] == ["depot_in", "depot_out"]


def test_erm_route_import_updates_existing_route_by_number(tmp_path):
    client = make_client(tmp_path)
    create = client.post("/api/refs/routes", json={
        "number": "1",
        "name": "Старое имя",
        "version": 3,
        "notes": "ручная заметка",
    })
    assert create.status_code == 200, create.text

    response = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["created"] is False
    assert data["updated"] is True

    route = next(r for r in client.get("/api/refs/routes").json()["items"] if r["number"] == "1")
    assert route["name"] == "Железнодорожный вокзал - Ореховая улица"
    assert route["version"] == 4
    notes = json.loads(route["notes"])
    assert notes["previous_notes"] == "ручная заметка"
    network = client.get(f"/api/routes/{data['route_id']}/network").json()
    assert [row["stop"]["external_code"] for row in network["forward"]] == [
        "100", "101", "102"
    ]
    assert network["forward"][0]["stop"]["latitude"] == 56.801
    assert network["forward"][-1]["cumulative_km"] == 1.5
    assert network["forward"][1]["run_time_sec"] == 180
    assert network["backward"][-1]["cumulative_km"] == 1.3

    repeated = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert repeated.status_code == 200, repeated.text
    rows = _depot_rows(data["route_id"])
    assert len(rows) == 4
    assert [(row["direction"], row["sequence"]) for row in rows] == [
        ("depot_in", 1), ("depot_in", 2),
        ("depot_out", 1), ("depot_out", 2),
    ]


def test_erm_route_import_combines_repeated_depot_sections(tmp_path):
    client = make_client(tmp_path)

    response = client.post(
        "/api/import/routes/erm",
        files={
            "file": (
                "erm.xlsx",
                build_erm_workbook_with_depot_continuation(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    rows = [
        row for row in _depot_rows(response.json()["route_id"])
        if row["direction"] == "depot_out"
    ]
    assert [row["external_code"] for row in rows] == ["300", "201", "302"]
    assert [row["sequence"] for row in rows] == [1, 2, 3]
    assert rows[-1]["distance_from_prev_km"] == 0.25
    assert rows[-1]["run_time_day_sec"] == 150
    assert rows[-1]["run_time_night_sec"] == 150
    assert json.loads(rows[-1]["source_detail"])["section"] == 2


def test_erm_route_import_does_not_clear_absent_depot_direction(tmp_path):
    client = make_client(tmp_path)
    created = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert created.status_code == 200, created.text
    route_id = created.json()["route_id"]
    depot_in_before = [
        (row["external_code"], row["sequence"])
        for row in _depot_rows(route_id) if row["direction"] == "depot_in"
    ]

    updated = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_without_depot_in(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert updated.status_code == 200, updated.text
    rows = _depot_rows(route_id)
    assert [
        (row["external_code"], row["sequence"])
        for row in rows if row["direction"] == "depot_in"
    ] == depot_in_before
    assert any(row["direction"] == "depot_out" for row in rows)


def test_erm_route_import_rolls_back_route_and_depot_changes(tmp_path, monkeypatch):
    client = make_client(tmp_path)
    created = client.post("/api/refs/routes", json={
        "number": "1",
        "name": "До импорта",
        "version": 3,
        "notes": "старое",
    })
    assert created.status_code == 200, created.text
    route_id = created.json()["id"]

    import app.api_refs as api_refs
    original = api_refs.normalize_erm_depot_sections

    def fail_after_depot_write(con, imported_route_id, details):
        original(con, imported_route_id, details)
        raise RuntimeError("test rollback")

    monkeypatch.setattr(api_refs, "normalize_erm_depot_sections", fail_after_depot_write)
    with pytest.raises(RuntimeError, match="test rollback"):
        client.post(
            "/api/import/routes/erm",
            files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    route = next(
        row for row in client.get("/api/refs/routes").json()["items"]
        if row["id"] == route_id
    )
    assert route["name"] == "До импорта"
    assert route["version"] == 3
    assert route["notes"] == "старое"
    assert _depot_rows(route_id) == []


def test_erm_route_import_does_not_choose_ambiguous_same_name_stop(tmp_path):
    client = make_client(tmp_path)

    import app.db as db
    con = db.connect()
    try:
        con.execute(
            "INSERT INTO stops(external_code,name,address,latitude,longitude) "
            "VALUES(?,?,?,?,?)",
            ("X", "Автопарк", "Другая улица", 1.0, 1.0),
        )
        first_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.execute(
            "INSERT INTO stops(external_code,name,address,latitude,longitude) "
            "VALUES(?,?,?,?,?)",
            ("Y", "Автопарк", "Ещё одна улица", 2.0, 2.0),
        )
        second_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.commit()
    finally:
        con.close()

    response = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200, response.text
    first_depot_out = next(
        row for row in _depot_rows(response.json()["route_id"])
        if row["direction"] == "depot_out" and row["sequence"] == 1
    )
    assert first_depot_out["stop_id"] not in {first_id, second_id}
    assert first_depot_out["external_code"] == "300"


def test_erm_import_unique_name_with_conflicting_discriminators_creates_new_stop(tmp_path):
    client = make_client(tmp_path)

    import app.db as db
    con = db.connect()
    try:
        cursor = con.execute(
            "INSERT INTO stops(external_code,name,address,latitude,longitude) "
            "VALUES(?,?,?,?,?)",
            ("X", "Автопарк", "Другая улица", 1.0, 1.0),
        )
        conflicting_id = cursor.lastrowid
        con.commit()
    finally:
        con.close()

    response = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200, response.text
    first_depot_out = next(
        row for row in _depot_rows(response.json()["route_id"])
        if row["direction"] == "depot_out" and row["sequence"] == 1
    )
    assert first_depot_out["stop_id"] != conflicting_id
    assert first_depot_out["external_code"] == "300"
    assert first_depot_out["address"] == "Гаражная улица"
    assert first_depot_out["latitude"] == pytest.approx(56.801)
    assert first_depot_out["longitude"] == pytest.approx(35.901)


def test_empty_or_malformed_depot_sections_preserve_existing_rows(tmp_path):
    client = make_client(tmp_path)
    imported = client.post(
        "/api/import/routes/erm",
        files={"file": ("erm.xlsx", build_erm_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    route_id = imported.json()["route_id"]
    before = [
        (row["stop_id"], row["sequence"], row["source"])
        for row in _depot_rows(route_id) if row["direction"] == "depot_out"
    ]

    import app.db as db
    from app.route_depot import normalize_erm_depot_sections

    con = db.connect()
    try:
        normalize_erm_depot_sections(con, route_id, {
            "sheets": {"из парка": {"sections": [
                {"kind": "из парка", "stops": []},
                {"kind": "из парка", "stops": "bad"},
                {"kind": "из парка", "stops": [None, 7]},
            ]}}
        })
        con.commit()
        state_count = con.execute(
            "SELECT COUNT(*) FROM route_depot_section_state "
            "WHERE route_id=? AND direction='depot_out'",
            (route_id,),
        ).fetchone()[0]
    finally:
        con.close()

    after = [
        (row["stop_id"], row["sequence"], row["source"])
        for row in _depot_rows(route_id) if row["direction"] == "depot_out"
    ]
    assert after == before
    assert state_count == 1


def test_erm_depot_normalization_scans_stop_registry_once_and_reuses_new_stop(tmp_path):
    client = make_client(tmp_path)
    created = client.post("/api/refs/routes", json={
        "number": "77", "name": "Тестовый маршрут", "version": 1,
    })
    assert created.status_code == 200, created.text
    route_id = created.json()["id"]

    stops = [
        {
            "stop_id": 900 + index,
            "stop_name": f"Новая {index}",
            "street": f"Улица {index}",
            "lat": 55.0 + index / 1000,
            "lon": 37.0 + index / 1000,
            "distance_km": 0.1,
            "travel_time": "00:01:00",
        }
        for index in range(12)
    ]
    stops[0].update({
        "address": None,
        "latitude": None,
        "longitude": None,
    })
    stops.append(dict(stops[0]))
    details = {"sheets": {"из парка": {"sections": [{
        "kind": "из парка", "stops": stops,
    }]}}}

    import app.db as db
    from app.route_depot import normalize_erm_depot_sections

    statements = []
    con = db.connect()
    try:
        con.set_trace_callback(statements.append)
        normalize_erm_depot_sections(con, route_id, details)
        con.set_trace_callback(None)
        con.commit()
        rows = db.rows(con.execute(
            "SELECT r.stop_id,s.address,s.latitude,s.longitude "
            "FROM route_depot_stops r JOIN stops s ON s.id=r.stop_id "
            "WHERE r.route_id=? AND r.direction='depot_out' "
            "ORDER BY r.sequence",
            (route_id,),
        ))
        first_code_count = con.execute(
            "SELECT COUNT(*) FROM stops WHERE external_code='900'"
        ).fetchone()[0]
    finally:
        con.close()

    full_scans = [
        statement for statement in statements
        if statement.strip().upper() == "SELECT * FROM STOPS ORDER BY ID"
    ]
    assert len(full_scans) == 1
    assert len(rows) == 13
    assert rows[0]["stop_id"] == rows[-1]["stop_id"]
    assert first_code_count == 1
    assert rows[0]["address"] == "Улица 0"
    assert rows[0]["latitude"] == pytest.approx(55.0)
    assert rows[0]["longitude"] == pytest.approx(37.0)
