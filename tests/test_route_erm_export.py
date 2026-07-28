# -*- coding: utf-8 -*-
import datetime
import io
from types import SimpleNamespace
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


HEADERS = [
    "№", "ID", "Остановочный пункт", "Улица", "Широта", "Долгота",
    "День между ОП", "День нарастающим", "Ночь между ОП",
    "Ночь нарастающим", "Расстояние между ОП", "Расстояние нарастающим",
]


def _client(tmp_path):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    db.DB_PATH = str(tmp_path / "route-erm-export.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
    finally:
        con.close()
    client = TestClient(app)
    token = client.post(
        "/api/login", json={"username": "admin", "password": "admin"}
    ).json()["token"]
    client.headers.update({"Authorization": f"Bearer {token}"})
    return client


def _seed_route(*, number="44", name="Вокзал — Аэропорт"):
    import app.db as db

    con = db.connect()
    try:
        route_id = con.execute(
            "INSERT INTO routes(number,name,start_point,end_point,version) "
            "VALUES(?,?,?,?,?)",
            (number, name, "Вокзал", "Аэропорт", 7),
        ).lastrowid
        stops = [
            con.execute(
                "INSERT INTO stops(name,external_code,address,latitude,longitude) "
                "VALUES(?,?,?,?,?)",
                stop,
            ).lastrowid
            for stop in (
                ("Парк", "DEPOT", "Гаражная, 1", 55.70, 37.50),
                ("Вокзал", "START", "Вокзальная, 1", 55.71, 37.51),
                ("Площадь", None, "=HYPERLINK(\"bad\")", None, None),
                ("Аэропорт", "END", "Аэропорт, 1", 55.73, 37.53),
            )
        ]
        con.executemany(
            """INSERT INTO route_stops(
                route_id,direction,stop_id,sequence,distance_from_prev_km,
                run_time_day_sec,run_time_night_sec
            ) VALUES(?,?,?,?,?,?,?)""",
            [
                (route_id, "forward", stops[1], 1, 0, 0, 0),
                (route_id, "forward", stops[2], 2, 2.5, 180, 210),
                (route_id, "forward", stops[3], 3, 3.5, 240, 270),
                (route_id, "backward", stops[3], 1, 0, 0, 0),
                (route_id, "backward", stops[1], 2, 6.0, 420, 480),
            ],
        )
        con.executemany(
            """INSERT INTO route_depot_stops(
                route_id,direction,stop_id,sequence,distance_from_prev_km,
                run_time_day_sec,run_time_night_sec
            ) VALUES(?,?,?,?,?,?,?)""",
            [
                (route_id, "depot_out", stops[0], 1, 0, 0, 0),
                (route_id, "depot_out", stops[1], 2, 1.5, 90, 110),
                (route_id, "depot_in", stops[3], 1, 0, 0, 0),
                (route_id, "depot_in", stops[0], 2, 2.0, 120, 150),
            ],
        )
        con.commit()
        return route_id
    finally:
        con.close()


def _download(client, route_id, *, season="summer", effective_date="2026-06-01"):
    return client.get(
        f"/api/routes/{route_id}/erm-export.xlsx",
        params={"season": season, "effective_date": effective_date},
    )


def _workbook(response):
    assert response.status_code == 200, response.text
    return load_workbook(io.BytesIO(response.content), data_only=False)


def _values(sheet):
    return [cell.value for row in sheet.iter_rows() for cell in row]


def _header_rows(sheet):
    return [
        row
        for row in range(1, sheet.max_row + 1)
        if [sheet.cell(row, column).value for column in range(1, 13)] == HEADERS
    ]


def test_erm_endpoint_filename_sheets_headings_and_layout(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_route()
    response = _download(client, route_id)

    assert response.status_code == 200, response.text
    assert "ЭРМ_М044_20260601_ЛЕТО.xlsx" in unquote(
        response.headers["content-disposition"]
    )
    workbook = _workbook(response)
    assert workbook.sheetnames == ["Параметры", "Из парка", "В парк"]
    parameters = workbook["Параметры"]
    assert parameters["A1"].value == "ЭЛЕКТРОННАЯ МОДЕЛЬ МАРШРУТА"
    assert parameters["A4"].value == "ПАРАМЕТРЫ МАРШРУТА"
    assert [parameters.cell(5, column).value for column in range(1, 13)] == HEADERS
    assert parameters["A6"].value == "Прямое направление"
    assert _header_rows(parameters) == [5, 13]
    assert "Прямое направление" in _values(parameters)
    assert "Обратное направление" in _values(parameters)

    for sheet in workbook.worksheets:
        assert sheet.sheet_view.showGridLines is False
        assert str(sheet.page_setup.paperSize) == str(sheet.PAPERSIZE_A4)
        assert sheet.page_setup.fitToWidth == 1
        assert sheet.print_area and sheet.print_title_rows == "$1:$5"
        assert sheet.freeze_panes == "A6"
        _, _, max_col, _ = range_boundaries(
            str(sheet.print_area).split("!")[-1].replace("'", "").replace("$", "")
        )
        assert max_col == 12
        assert "Сформировано" in (sheet.oddFooter.left.text or "")
        assert "версия 7" in (sheet.oddFooter.left.text or "")


def test_erm_numeric_coordinates_formulas_reset_and_warnings(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_route()
    sheet = _workbook(_download(client, route_id))["Параметры"]

    assert sheet["E7"].value == pytest.approx(55.71)
    assert sheet["F7"].value == pytest.approx(37.51)
    assert isinstance(sheet["E7"].value, float)
    assert sheet["E8"].value is None and sheet["F8"].value is None
    assert sheet["B8"].value is None
    assert sheet["B8"].fill.fgColor.rgb.endswith("FFF3DD")
    assert sheet["E8"].fill.fgColor.rgb.endswith("FFF3DD")
    assert sheet["F8"].fill.fgColor.rgb.endswith("FFF3DD")
    assert "Примечание: отсутствуют технические данные" in _values(sheet)

    assert sheet["G8"].value == datetime.timedelta(seconds=180)
    assert sheet["G8"].number_format == "[h]:mm"
    assert sheet["H8"].value == "=SUM($G$7:G8)"
    assert sheet["H8"].number_format == "[h]:mm"
    assert sheet["J8"].value == "=SUM($I$7:I8)"
    assert sheet["L8"].value == "=SUM($K$7:K8)"
    assert sheet["H14"].value == "=SUM($G$14:G14)"
    assert sheet["J14"].value == "=SUM($I$14:I14)"
    assert sheet["L14"].value == "=SUM($K$14:K14)"
    assert "$G$7" not in sheet["H14"].value


def test_erm_empty_depot_sheets_remain_complete(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_route()
    import app.db as db

    con = db.connect()
    try:
        con.execute("DELETE FROM route_depot_stops WHERE route_id=?", (route_id,))
        con.commit()
    finally:
        con.close()

    workbook = _workbook(_download(client, route_id))
    for name in ("Из парка", "В парк"):
        sheet = workbook[name]
        assert sheet["A1"].value == "ЭЛЕКТРОННАЯ МОДЕЛЬ МАРШРУТА"
        assert [sheet.cell(5, column).value for column in range(1, 13)] == HEADERS
        assert _values(sheet).count("Нулевой рейс не заполнен") == 1
        assert sheet.print_area and sheet.freeze_panes == "A6"


def test_erm_auth_validation_unknown_route_and_filename_normalization(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_route()
    assert _download(TestClient(client.app), route_id).status_code in (401, 403)
    assert _download(client, route_id, season="autumn").status_code == 400
    assert _download(client, route_id, effective_date="01.06.2026").status_code == 400
    assert _download(client, route_id + 99999).status_code == 404

    import app.db as db

    con = db.connect()
    try:
        con.execute(
            "UPDATE routes SET number=?,name=? WHERE id=?",
            ("M1/A", "Тестовый маршрут", route_id),
        )
        con.commit()
    finally:
        con.close()
    visible_workbook = _workbook(_download(client, route_id))
    for sheet in visible_workbook.worksheets:
        assert sheet["A2"].value.startswith("Маршрут: M1/A —")
        assert sheet.oddHeader.left.text == "Маршрут M1/A"

    con = db.connect()
    try:
        con.execute(
            "UPDATE routes SET number=? WHERE id=?",
            ("A&B", route_id),
        )
        con.commit()
    finally:
        con.close()
    ampersand_response = _download(client, route_id)
    assert "ЭРМ_МA_B_20260601_ЛЕТО.xlsx" in unquote(
        ampersand_response.headers["content-disposition"]
    )
    ampersand_workbook = _workbook(ampersand_response)
    for sheet in ampersand_workbook.worksheets:
        assert sheet["A2"].value.startswith("Маршрут: A&B —")
        assert sheet.oddHeader.left.text == "Маршрут A&&B"

    from app.route_document_data import parse_document_options
    from app.route_document_xlsx import erm_filename

    options = parse_document_options("summer", "2026-06-01")
    assert erm_filename(SimpleNamespace(route_number="M1"), options).startswith("ЭРМ_М001_")
    assert erm_filename(SimpleNamespace(route_number="М044"), options).startswith("ЭРМ_М044_")
    assert erm_filename(SimpleNamespace(route_number="M44/A"), options).startswith("ЭРМ_М44_A_")


def test_erm_formula_like_text_long_values_and_no_error_literals(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_route(
        number="44/А очень длинный номер маршрута",
        name="Очень длинное имя маршрута " * 20,
    )
    import app.db as db

    long_stop_name = "+" + "Очень длинное имя остановочного пункта " * 12
    long_address = "-" + "Очень длинный адрес остановочного пункта " * 12
    con = db.connect()
    try:
        con.execute(
            "UPDATE stops SET name=?,external_code=?,address=? WHERE id=("
            "SELECT stop_id FROM route_stops WHERE route_id=? AND direction='forward' "
            "ORDER BY sequence LIMIT 1)",
            (long_stop_name, "@EXTERNAL", long_address, route_id),
        )
        con.commit()
    finally:
        con.close()

    response = _download(client, route_id)
    disposition = unquote(response.headers["content-disposition"])
    assert "ЭРМ_М44_А_очень_длинный_номер_маршрута_20260601_ЛЕТО.xlsx" in disposition
    assert "\r" not in disposition and "\n" not in disposition
    workbook = _workbook(response)
    parameters = workbook["Параметры"]
    assert parameters["A2"].value.startswith(
        "Маршрут: 44/А очень длинный номер маршрута —"
    )
    rendered = [
        cell
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value in (
            "'" + long_stop_name,
            "'@EXTERNAL",
            "'" + long_address,
        )
    ]
    assert len(rendered) >= 3 and all(cell.data_type == "s" for cell in rendered)
    assert 210 < parameters.row_dimensions[7].height <= 240
    formulas = [
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert formulas
    assert all(
        token not in formula
        for formula in formulas
        for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    )
    assert parameters.row_dimensions[2].height > 24
