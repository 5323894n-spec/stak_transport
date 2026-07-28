# -*- coding: utf-8 -*-
import io
from types import SimpleNamespace
from urllib.parse import unquote

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import range_boundaries


def _client(tmp_path):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app
    db.DB_PATH = str(tmp_path / "schedule-document.db")
    db.init_db()
    con = db.connect()
    try: ensure_admin(con)
    finally: con.close()
    client = TestClient(app)
    token = client.post("/api/login", json={"username": "admin", "password": "admin"}).json()["token"]
    client.headers.update({"Authorization": f"Bearer {token}"})
    return client


def _seed_route(*, number="44", name="Вокзал — Аэропорт", weekend=True):
    import app.db as db
    con = db.connect()
    try:
        route_id = con.execute("INSERT INTO routes(number,name,start_point,end_point,version) VALUES(?,?,?,?,?)", (number, name, "Вокзал", "Аэропорт", 7)).lastrowid
        stops = [con.execute("INSERT INTO stops(name,external_code) VALUES(?,?)", (stop_name, f"{external_code}-{route_id}")).lastrowid for stop_name, external_code in (("Парк", "DEPOT"), ("Вокзал", "START"), ("Аэропорт", "END"))]
        con.executemany("""INSERT INTO route_stops(route_id,direction,stop_id,sequence,distance_from_prev_km,run_time_day_sec,run_time_night_sec) VALUES(?,?,?,?,?,?,?)""", [(route_id,"forward",stops[1],1,0,0,0),(route_id,"forward",stops[2],2,12.5,4200,4500),(route_id,"backward",stops[2],1,0,0,0),(route_id,"backward",stops[1],2,12.5,4200,4500)])
        con.executemany("""INSERT INTO route_depot_stops(route_id,direction,stop_id,sequence,distance_from_prev_km,run_time_day_sec,run_time_night_sec) VALUES(?,?,?,?,?,?,?)""", [(route_id,"depot_out",stops[0],1,0,0,0),(route_id,"depot_out",stops[1],2,2,600,720),(route_id,"depot_in",stops[2],1,0,0,0),(route_id,"depot_in",stops[0],2,3,900,1080)])
        trips = [(route_id,"будни",1,1,1,"прямое","22:00","23:10",12.5,15,"отстой"),(route_id,"будни",1,1,2,"обратное","00:30","01:40",12.5,30,"обед")]
        if weekend: trips.append((route_id,"выходные",1,1,1,"прямое","09:00","10:10",12.5,10,"отстой"))
        con.executemany("""INSERT INTO route_trips(route_id,day_type,output_number,shift_number,trip_number,direction,dep_time,arr_time,distance_km,break_after_min,break_type) VALUES(?,?,?,?,?,?,?,?,?,?,?)""", trips)
        con.commit(); return route_id
    finally: con.close()


def _download(client, route_id, *, season="winter", effective_date="2025-12-01"):
    return client.get(f"/api/routes/{route_id}/schedule-document.xlsx", params={"season": season, "effective_date": effective_date})


def _workbook(response):
    assert response.status_code == 200, response.text
    return load_workbook(io.BytesIO(response.content), data_only=False)


def _built_workbook(route_id):
    import app.db as db
    from app.route_document_data import load_route_document_data, parse_document_options
    from app.route_document_xlsx import build_schedule_workbook
    con = db.connect()
    try: data = load_route_document_data(con, route_id)
    finally: con.close()
    return build_schedule_workbook(data, parse_document_options("winter", "2025-12-01"))


def test_schedule_document_endpoint_contract(tmp_path):
    client = _client(tmp_path); route_id = _seed_route(); response = _download(client, route_id)
    assert response.status_code == 200, response.text
    assert "Расписание_М044_20251201_ЗИМА.xlsx" in unquote(response.headers["content-disposition"])
    workbook = _workbook(response); assert workbook.sheetnames == ["Рабочие дни", "Выходные дни", "Хронометраж"]
    sheet = workbook["Рабочие дни"]
    assert sheet["A1"].value == "МАРШРУТНОЕ РАСПИСАНИЕ" and sheet.sheet_view.showGridLines is False
    assert sheet.page_setup.orientation == "landscape" and sheet.freeze_panes and sheet.print_area and sheet.print_title_rows == "$1:$8"
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert any(isinstance(value, str) and value.startswith("=") for value in values)
    kpis = ("Количество выходов","Количество рейсов","Общий пробег, км","Сумма перерывов","Продолжительность работы")
    assert all(label in values for label in kpis) and all(label in values for label in ("Из парка","В парк","Перерывы","Всего"))
    assert sheet["C8"].value == "Рейс 1\nотправление (Вокзал)"
    assert sheet["D8"].value == "Рейс 1\nприбытие (Аэропорт)"
    assert sheet["E8"].value == "Рейс 2\nотправление (Аэропорт)"
    assert sheet["F8"].value == "Рейс 2\nприбытие (Вокзал)"
    assert any(isinstance(value,str) and "Версия 7" in value for value in values)
    _,_,print_max_col,_ = range_boundaries(str(sheet.print_area).split("!")[-1].replace("'","").replace("$",""))
    assert all(next(cell.column for row in sheet.iter_rows() for cell in row if cell.value == label) <= print_max_col for label in kpis)
    source = _built_workbook(route_id)["Рабочие дни"]
    assert any(isinstance(cell.value,(int,float)) and cell.number_format == "[h]:mm" for row in source.iter_rows() for cell in row)
    assert source["B9"].value == (22*3600-600)/86400
    assert source["G9"].value == (1*3600+40*60+900)/86400
    assert source["R9"].value == 14700 and source["I9"].value == "=R9/86400"


def test_schedule_filename_normalizes_existing_route_prefixes():
    from app.route_document_data import parse_document_options
    from app.route_document_xlsx import schedule_filename
    options = parse_document_options("winter", "2025-12-01")
    assert schedule_filename(SimpleNamespace(route_number="M1"), options).startswith("Расписание_М001_")
    assert schedule_filename(SimpleNamespace(route_number="М044"), options).startswith("Расписание_М044_")
    assert schedule_filename(SimpleNamespace(route_number="M44/A"), options).startswith("Расписание_М44_A_")
    assert schedule_filename(SimpleNamespace(route_number="М44-А"), options).startswith("Расписание_М44_А_")


def test_schedule_document_auth_validation_and_unknown_route(tmp_path):
    client = _client(tmp_path); route_id = _seed_route()
    assert _download(TestClient(client.app), route_id).status_code in (401,403)
    bad_season = _download(client,route_id,season="autumn"); bad_date = _download(client,route_id,effective_date="01.12.2025"); missing = _download(client,route_id+99999)
    assert bad_season.status_code == 400 and "Сезон" in bad_season.json()["detail"]
    assert bad_date.status_code == 400 and "YYYY-MM-DD" in bad_date.json()["detail"]
    assert missing.status_code == 404 and "Маршрут не найден" in missing.json()["detail"]


def test_empty_weekend_has_metadata_kpis_and_explicit_message(tmp_path):
    client = _client(tmp_path); route_id = _seed_route(weekend=False); sheet = _workbook(_download(client,route_id))["Выходные дни"]
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert sheet["A1"].value == "МАРШРУТНОЕ РАСПИСАНИЕ" and "Выходные дни" in values and values.count("Расписание не заполнено") == 1
    assert all(label in values for label in ("Количество выходов","Количество рейсов","Общий пробег, км","Сумма перерывов","Продолжительность работы"))


def test_long_names_helpers_and_formulas_remain_safe(tmp_path):
    client = _client(tmp_path); route_id = _seed_route(number="44/А очень длинный номер маршрута", name="Очень длинное имя маршрута "*20)
    response = _download(client,route_id); workbook = _workbook(response); disposition = unquote(response.headers["content-disposition"])
    assert "\r" not in disposition and "\n" not in disposition
    assert "Расписание_М44_А_очень_длинный_номер_маршрута_20251201_ЗИМА.xlsx" in disposition
    for sheet in workbook.worksheets:
        formulas = [cell.value for row in sheet.iter_rows() for cell in row if isinstance(cell.value,str) and cell.value.startswith("=")]
        assert all(token not in formula for formula in formulas for token in ("#REF!","#DIV/0!","#VALUE!","#NAME?","#N/A"))
        _,_,max_col,_ = range_boundaries(str(sheet.print_area).split("!")[-1].replace("'","").replace("$",""))
        hidden = [column_index_from_string(column) for column,dimension in sheet.column_dimensions.items() if dimension.hidden]
        assert hidden and min(hidden) > max_col and sheet.row_dimensions[2].height > 24


def test_nested_weekend_variants_are_distinct_but_generic_is_simple(tmp_path):
    client = _client(tmp_path); route_id = _seed_route(weekend=False)
    import app.db as db
    con = db.connect()
    try:
        con.executemany("""INSERT INTO route_trips(route_id,day_type,output_number,shift_number,trip_number,direction,dep_time,arr_time,distance_km) VALUES(?,?,?,?,?,?,?,?,?)""", [(route_id,"суббота",1,1,1,"прямое","08:00","09:00",10),(route_id,"воскресенье",1,1,1,"прямое","09:00","10:00",10)]); con.commit()
    finally: con.close()
    visible = [cell.value for row in _workbook(_download(client,route_id))["Выходные дни"].iter_rows() for cell in row[:10]]
    assert visible.count("суббота · 1") == 1 and visible.count("воскресенье · 1") == 1
    other = _seed_route(weekend=False); con = db.connect()
    try:
        con.execute("""INSERT INTO route_trips(route_id,day_type,output_number,shift_number,trip_number,direction,dep_time,arr_time,distance_km) VALUES(?,?,?,?,?,?,?,?,?)""", (other,"weekend",1,1,1,"прямое","09:00","10:00",10)); con.commit()
    finally: con.close()
    simple_values = [cell.value for row in _workbook(_download(client,other))["Выходные дни"].iter_rows() for cell in row[:10]]
    assert 1 in simple_values and "weekend · 1" not in simple_values


def test_nested_consumer_preserves_chronology_and_chronometry_sections(tmp_path):
    _client(tmp_path); route_id = _seed_route(); workbook = _built_workbook(route_id)
    day = workbook["Рабочие дни"]
    assert day["C9"].value == 22*3600/86400 and day["E9"].value == 30*60/86400
    chrono = workbook["Хронометраж"]; values = [cell.value for row in chrono.iter_rows() for cell in row]
    assert all(label in values for label in ("Прямое направление","Обратное направление","Из парка","В парк"))
    assert any(isinstance(value,str) and value.startswith("=SUM($C$") for value in values)
    assert any(isinstance(value,str) and value.startswith("=SUM($D$") for value in values)
    assert any(isinstance(cell.value,(int,float)) and cell.number_format == "[h]:mm" for row in chrono.iter_rows() for cell in row)


def test_formula_like_user_text_is_stored_as_literal(tmp_path):
    _client(tmp_path)
    import app.db as db
    con = db.connect()
    try:
        route_id = con.execute("INSERT INTO routes(number,name,start_point,end_point,version) VALUES(?,?,?,?,?)", ("1","=ROUTE","+START","-END",1)).lastrowid
        malicious = (("@STOP","=EXT"),("=STOP","+EXT"),("+STOP","-EXT"),("-STOP","@EXT"))
        stop_ids = [con.execute("INSERT INTO stops(name,external_code) VALUES(?,?)", pair).lastrowid for pair in malicious]
        con.executemany("INSERT INTO route_stops(route_id,direction,stop_id,sequence) VALUES(?,?,?,?)", [(route_id,"forward",stop_id,index) for index,stop_id in enumerate(stop_ids,1)])
        con.commit()
    finally: con.close()
    workbook = _built_workbook(route_id)
    attacker_tokens = ("=ROUTE","+START","-END","@STOP","=STOP","+STOP","-STOP","=EXT","+EXT","-EXT","@EXT")
    rendered = [cell for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if isinstance(cell.value,str) and any(token in cell.value for token in attacker_tokens)]
    assert rendered and all(cell.data_type == "s" for cell in rendered)
    dangerous = [cell for cell in rendered if cell.value.lstrip("'").startswith(("=","+","-","@"))]
    assert dangerous and all(cell.value.startswith("'") for cell in dangerous)
    formulas = [cell.value for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if cell.data_type == "f"]
    assert formulas and all(not any(token in formula for token in attacker_tokens) for formula in formulas)
