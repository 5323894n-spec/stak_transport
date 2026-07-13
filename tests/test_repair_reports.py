# -*- coding: utf-8 -*-
import io
from openpyxl import load_workbook
from tests.test_repair_operations_api import create_order
from tests.test_repair_requests_api import make_client


def test_repair_excel_report_has_required_sheets_and_format(tmp_path):
    client, bus_id = make_client(tmp_path)
    create_order(client, bus_id)
    response = client.get("/api/repairs/reports/export.xlsx")
    assert response.status_code == 200, response.text
    wb = load_workbook(io.BytesIO(response.content), data_only=False)
    assert wb.sheetnames == ["Сводка", "Заказ-наряды", "Заявки", "История", "Планы ТО", "Склад"]
    assert wb["Сводка"]["A1"].value == "ОТЧЁТ ПО РЕМОНТУ И ТО"
    assert wb["Заказ-наряды"].freeze_panes == "A2"
    assert wb["Заказ-наряды"].auto_filter.ref
    assert wb["Заказ-наряды"]["A1"].font.bold is True
    assert wb["Заказ-наряды"].max_row == 2

def test_repair_excel_report_filters_orders_by_period_vehicle_and_status(tmp_path):
    client, first_bus = make_client(tmp_path)
    first = create_order(client, first_bus)

    import app.db as db
    con = db.connect()
    try:
        second_bus = con.execute(
            "INSERT INTO buses(garage_number,plate,odometer) VALUES(?,?,?)",
            ("Р-202", "А202АА69", 22000),
        ).lastrowid
        con.execute("UPDATE repair_orders SET created_at=? WHERE id=?", ("2026-07-10T10:00:00", first["id"]))
        con.commit()
    finally:
        con.close()

    second_request = client.post("/api/repairs/requests", json={
        "vehicle_id": second_bus, "odometer": 22000, "fault_description": "Ремонт второго автобуса",
    })
    repair_type_id = client.get("/api/repairs/references").json()["repair_types"][0]["id"]
    second = client.post("/api/repairs/orders", json={
        "request_id": second_request.json()["id"], "vehicle_id": second_bus,
        "repair_type_id": repair_type_id,
    })
    assert second.status_code == 201, second.text

    con = db.connect()
    try:
        con.execute(
            "UPDATE repair_orders SET created_at=?,status=? WHERE id=?",
            ("2026-08-10T10:00:00", "диагностика", second.json()["id"]),
        )
        con.execute(
            "UPDATE repair_requests SET created_at=? WHERE id=?",
            ("2026-08-10T09:00:00", second_request.json()["id"]),
        )
        con.commit()
    finally:
        con.close()

    response = client.get(
        f"/api/repairs/reports/export.xlsx?date_from=2026-08-01&date_to=2026-08-31&vehicle_id={second_bus}&status=диагностика"
    )
    assert response.status_code == 200, response.text
    wb = load_workbook(io.BytesIO(response.content), data_only=True)
    orders = wb["Заказ-наряды"]
    assert orders.max_row == 2
    assert orders["A2"].value == second.json()["order_number"]
    assert orders["B2"].value == "диагностика"
    assert orders["C2"].value == "Р-202"
    assert wb["Сводка"]["B4"].value == 1