# -*- coding: utf-8 -*-
from dataclasses import fields, is_dataclass
from typing import get_type_hints

import pytest
from openpyxl import Workbook

from app.route_document_data import (
    ScheduleTrip,
    load_route_document_data,
    parse_document_options,
)
from app.route_document_xlsx import (
    BLUE,
    NAVY,
    PALE_AMBER,
    PALE_BLUE,
    PALE_GREEN,
    _xlsx_download_response,
    apply_sheet_setup,
    apply_warning_cell,
    set_print_area,
    write_excel_time,
    write_section_header,
    write_table_header,
    write_title_band,
)


def _open_db(tmp_path):
    from app import db

    db.DB_PATH = str(tmp_path / "route-document-data.db")
    db.init_db()
    return db.connect()


def _seed_document(con):
    route_id = con.execute(
        """
        INSERT INTO routes(number,name,start_point,end_point,version)
        VALUES(?,?,?,?,?)
        """,
        ("42", "Центральный", "Вокзал", "Аэропорт", 7),
    ).lastrowid
    stop_ids = [
        con.execute(
            "INSERT INTO stops(name,external_code,address,latitude,longitude) "
            "VALUES(?,?,?,?,?)",
            stop,
        ).lastrowid
        for stop in (
            ("Вокзал", "STOP-1", "Привокзальная площадь", 56.123, 35.987),
            ("Площадь", "STOP-2", None, None, None),
            ("Аэропорт", "STOP-3", "Аэропорт, 1", 56.456, 36.123),
            ("Парк", "DEPOT-1", "Гаражная, 5", 56.789, 36.456),
        )
    ]
    con.executemany(
        """
        INSERT INTO route_stops(
          route_id,direction,stop_id,sequence,distance_from_prev_km,
          run_time_day_sec,run_time_night_sec
        ) VALUES(?,?,?,?,?,?,?)
        """,
        [
            (route_id, "forward", stop_ids[1], 2, 2.5, 180, 210),
            (route_id, "forward", stop_ids[0], 1, 0, 0, 0),
            (route_id, "backward", stop_ids[2], 1, 0, 0, 0),
        ],
    )
    con.executemany(
        """
        INSERT INTO route_depot_stops(
          route_id,direction,stop_id,sequence,distance_from_prev_km,
          run_time_day_sec,run_time_night_sec
        ) VALUES(?,?,?,?,?,?,?)
        """,
        [
            (route_id, "depot_in", stop_ids[3], 1, 1.25, 120, 150),
            (route_id, "depot_out", stop_ids[0], 2, 1.5, 90, 110),
            (route_id, "depot_out", stop_ids[3], 1, 0, 0, 0),
        ],
    )
    con.executemany(
        """
        INSERT INTO route_trips(
          route_id,day_type,output_number,shift_number,trip_number,
          direction,dep_time,arr_time,distance_km
        ) VALUES(?,?,?,?,?,?,?,?,?)
        """,
        [
            (route_id, "будни", 2, 1, 1, "прямое", "08:00", "08:30", 9.5),
            (route_id, "будни", 1, 1, 2, "обратное", "07:00", "07:35", 9.5),
            (route_id, "будни", 1, 1, 1, "прямое", "06:00", "06:35", 9.5),
            (route_id, "выходные", 1, 1, 1, "прямое", "09:00", "09:35", 9.5),
            (route_id, "workday", 1, 1, 1, "прямое", "12:00", "12:35", 9.5),
            (route_id, "weekend", 1, 1, 1, "прямое", "13:00", "13:35", 9.5),
            (route_id, "воскресенье", 1, 1, 1, "прямое", "11:00", "11:35", 9.5),
            (route_id, "суббота", 1, 1, 1, "прямое", "10:00", "10:35", 9.5),
            (route_id, "будни", 3, 2, None, "прямое", "08:00", "08:30", 9.5),
            (route_id, "будни", 3, None, 1, "прямое", "07:00", "07:30", 9.5),
            (route_id, "будни", 3, 1, 1, "прямое", "06:00", "06:30", 9.5),
            (route_id, "будни", 4, 1, 3, "прямое", "02:00", "02:30", 9.5),
            (route_id, "будни", 4, 1, 1, "прямое", "22:00", "22:30", 9.5),
            (route_id, "будни", 4, 1, 2, "прямое", "00:30", "01:00", 9.5),
        ],
    )
    con.commit()
    return route_id


def _walk(value):
    yield value
    if is_dataclass(value):
        for field in fields(value):
            yield from _walk(getattr(value, field.name))
    elif isinstance(value, dict):
        for key, item in value.items():
            yield from _walk(key)
            yield from _walk(item)
    elif isinstance(value, (list, tuple)):
        for item in value:
            yield from _walk(item)


def test_parse_document_options_winter():
    options = parse_document_options("winter", "2025-12-01")

    assert options.season_label == "ЗИМНИЙ ПЕРИОД"
    assert options.file_token == "ЗИМА"
    assert options.effective_date.isoformat() == "2025-12-01"


def test_parse_document_options_summer():
    options = parse_document_options("summer", "2025-06-01")

    assert options.season_label == "ЛЕТНИЙ ПЕРИОД"
    assert options.file_token == "ЛЕТО"
    assert options.effective_date.isoformat() == "2025-06-01"


@pytest.mark.parametrize(
    ("season", "effective_date"),
    [("autumn", "2025-12-01"), ("winter", "01.12.2025")],
)
def test_parse_document_options_rejects_invalid_values(season, effective_date):
    with pytest.raises(ValueError):
        parse_document_options(season, effective_date)


def test_load_route_document_data_is_ordered_and_neutral(tmp_path):
    con = _open_db(tmp_path)
    try:
        route_id = _seed_document(con)
        statements = []
        con.set_trace_callback(statements.append)
        document = load_route_document_data(con, route_id)
        con.set_trace_callback(None)
    finally:
        con.close()

    assert (document.route_id, document.route_number, document.route_name) == (
        route_id,
        "42",
        "Центральный",
    )
    assert (document.start_point, document.end_point, document.version) == (
        "Вокзал",
        "Аэропорт",
        7,
    )
    assert len([
        sql for sql in statements if sql.lstrip().upper().startswith("SELECT")
    ]) == 4
    assert [row["name"] for row in document.forward.stops] == [
        "Вокзал",
        "Площадь",
    ]
    assert [row["name"] for row in document.backward.stops] == ["Аэропорт"]
    assert [row["name"] for row in document.depot_out.stops] == [
        "Парк",
        "Вокзал",
    ]
    assert [row["name"] for row in document.depot_in.stops] == ["Парк"]
    assert all("stop_name" not in row for row in document.forward.stops)
    assert document.forward.stops[0]["external_code"] == "STOP-1"
    assert document.depot_out.stops[0]["external_code"] == "DEPOT-1"
    assert document.forward.stops[1]["run_time_day_sec"] == 180
    assert document.forward.stops[1]["run_time_night_sec"] == 210
    assert document.depot_out.stops[1]["distance_from_prev_km"] == 1.5
    assert document.forward.stops[0]["address"] == "Привокзальная площадь"
    assert document.forward.stops[0]["latitude"] == pytest.approx(56.123)
    assert isinstance(document.forward.stops[0]["longitude"], float)
    assert document.forward.stops[1]["address"] is None
    assert document.forward.stops[1]["latitude"] is None
    assert document.forward.stops[1]["longitude"] is None
    assert document.depot_in.stops[0]["address"] == "Гаражная, 5"
    assert isinstance(document.depot_in.stops[0]["latitude"], float)

    assert list(document.schedules) == ["workday", "weekend"]
    assert list(document.schedules["workday"]) == ["будни", "workday"]
    assert list(document.schedules["weekend"]) == [
        "выходные", "weekend", "суббота", "воскресенье",
    ]
    workday_outputs = document.schedules["workday"]["будни"]
    assert [output.output_number for output in workday_outputs] == [1, 2, 3, 4]
    assert [trip.trip_number for trip in workday_outputs[0].trips] == [1, 2]
    assert workday_outputs[0].trips[0].departure_sec == 6 * 3600

    for variant in document.schedules["weekend"].values():
        assert [output.output_number for output in variant] == [1]
    assert document.schedules["weekend"]["выходные"][0].trips[0].arrival_sec == 9 * 3600 + 35 * 60
    assert document.schedules["weekend"]["суббота"][0].trips[0].departure_sec == 10 * 3600
    assert document.schedules["weekend"]["воскресенье"][0].trips[0].departure_sec == 11 * 3600

    assert all(
        not type(value).__module__.startswith("openpyxl")
        for value in _walk(document)
    )


def test_schedule_trips_are_ordered_on_service_day_axis(tmp_path):
    con = _open_db(tmp_path)
    try:
        route_id = _seed_document(con)
        outputs = load_route_document_data(
            con, route_id
        ).schedules["workday"]["будни"]
    finally:
        con.close()

    mismatched = outputs[2].trips
    assert [trip.departure_sec for trip in mismatched] == [
        6 * 3600, 7 * 3600, 8 * 3600,
    ]
    assert [trip.trip_number for trip in mismatched] == [1, 1, None]
    assert mismatched[1].shift_number is None

    overnight = outputs[3].trips
    assert [trip.departure_sec for trip in overnight] == [
        22 * 3600, 30 * 60, 2 * 3600,
    ]


def test_schedule_trip_declares_nullable_shift_number():
    assert get_type_hints(ScheduleTrip)["shift_number"] == int | None


def test_load_route_document_data_rejects_unknown_route(tmp_path):
    con = _open_db(tmp_path)
    try:
        with pytest.raises(ValueError, match="Маршрут не найден"):
            load_route_document_data(con, 999999)
    finally:
        con.close()


def test_xlsx_style_surface_smoke():
    assert (NAVY, BLUE, PALE_BLUE, PALE_GREEN, PALE_AMBER) == (
        "17324D",
        "DCE8F3",
        "EEF4FB",
        "E9F6EE",
        "FFF3DD",
    )
    workbook = Workbook()
    sheet = workbook.active

    apply_sheet_setup(sheet)
    assert sheet.sheet_view.showGridLines is False
    assert sheet.page_setup.orientation == "landscape"
    assert str(sheet.page_setup.paperSize) == str(sheet.PAPERSIZE_A4)
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 0
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert sheet.oddFooter.center.text == "Страница &P из &N"

    write_title_band(sheet, 1, "Документ", end_col=4)
    write_section_header(sheet, 2, "Раздел", end_col=4)
    write_table_header(sheet, 3, ("A", "B"))
    apply_warning_cell(sheet.cell(4, 1, "Проверить"))
    assert sheet["A1"].value == "Документ"
    assert sheet["A1"].fill.fgColor.rgb.endswith(NAVY)
    assert sheet["A2"].fill.fgColor.rgb.endswith(BLUE)
    assert sheet["A3"].fill.fgColor.rgb.endswith(PALE_BLUE)
    assert sheet["A4"].fill.fgColor.rgb.endswith(PALE_AMBER)

    write_excel_time(sheet["B4"], 5400)
    write_excel_time(sheet["C4"], None)
    assert sheet["B4"].value == 5400 / 86400
    assert sheet["B4"].number_format == "[h]:mm"
    assert sheet["C4"].value is None
    assert sheet["C4"].number_format == "[h]:mm"

    set_print_area(sheet, min_row=1, min_col=1, max_row=4, max_col=4)
    assert sheet.print_area == "'Sheet'!$A$1:$D$4"

    response = _xlsx_download_response(workbook, "маршрут/42.xlsx")
    assert response.media_type.endswith("spreadsheetml.sheet")
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    assert "%D0%BC%D0%B0%D1%80%D1%88%D1%80%D1%83%D1%82%2F42.xlsx" in (
        response.headers["content-disposition"]
    )

    portrait_sheet = workbook.create_sheet("Portrait")
    apply_sheet_setup(portrait_sheet, landscape=False)
    assert portrait_sheet.page_setup.orientation == "portrait"
