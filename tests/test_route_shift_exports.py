# -*- coding: utf-8 -*-
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


DAY_TYPE = "будни"


def _client(tmp_path, *, authenticated=True):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    tmp_path.mkdir(parents=True, exist_ok=True)
    db.DB_PATH = str(tmp_path / "route-shift-export.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
        con.commit()
    finally:
        con.close()
    client = TestClient(app)
    if authenticated:
        token = client.post(
            "/api/login", json={"username": "admin", "password": "admin"}
        ).json()["token"]
        client.headers.update({"Authorization": "Bearer " + token})
    return client


def _seed_export_data():
    import app.db as db

    con = db.connect()
    try:
        route_id = con.execute(
            "INSERT INTO routes(number,name,trip_time_min,trip_time_back_min) "
            "VALUES('42','Вокзал — Аэропорт',60,60)"
        ).lastrowid
        trip_ids = []
        for output, trip_number, dep, arr in (
            (2, 7, "23:30", "25:00"),
            (2, 8, "25:10", "27:00"),
            (1, 1, "06:00", "09:00"),
            (1, 2, "09:10", "14:30"),
        ):
            trip_ids.append(con.execute(
                "INSERT INTO route_trips(route_id,day_type,output_number,shift_number,"
                "trip_number,direction,dep_time,arr_time) VALUES(?,?,?,?,?,?,?,?)",
                (route_id, DAY_TYPE, output, 1, trip_number, "прямое", dep, arr),
            ).lastrowid)
        shift_types = {
            row["code"]: row["id"]
            for row in con.execute("SELECT id,code FROM shift_types")
        }
        now = "2026-07-22T10:00:00"
        first_shift = con.execute(
            """INSERT INTO output_shifts(
              route_id,day_type,output_number,shift_number,shift_type_id,
              trip_from_id,trip_to_id,start_sec,end_sec,driver_slots,
              handover_after_min,source,is_manual_locked,manual_reason,
              created_at,updated_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (route_id, DAY_TYPE, 2, 1, shift_types["two_driver_long"],
             trip_ids[0], trip_ids[1], 84600, 97200, 2, 15,
             "generated", 0, None, now, now),
        ).lastrowid
        second_shift = con.execute(
            """INSERT INTO output_shifts(
              route_id,day_type,output_number,shift_number,shift_type_id,
              trip_from_id,trip_to_id,start_sec,end_sec,driver_slots,
              handover_after_min,source,is_manual_locked,manual_reason,
              created_at,updated_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (route_id, DAY_TYPE, 1, 2, shift_types["single_8h"],
             trip_ids[2], trip_ids[3], 21600, 52200, 1, 10,
             "manual", 1, "Закреплено диспетчером", now, now),
        ).lastrowid
        driver_ids = [con.execute(
            "INSERT INTO drivers(tab_number,fio) VALUES(?,?)",
            (f"E{index + 1}", f"Водитель {index + 1}"),
        ).lastrowid for index in range(3)]
        for driver_id, date, shift_id, output, shift in (
            (driver_ids[0], "2026-07-21", first_shift, 2, 1),
            (driver_ids[1], "2026-07-22", first_shift, 2, 1),
            (driver_ids[2], "2026-07-22", second_shift, 1, 2),
        ):
            con.execute(
                "INSERT INTO roster_assignments(driver_id,date,route_id,day_type,"
                "output_number,shift_number,output_shift_id) VALUES(?,?,?,?,?,?,?)",
                (driver_id, date, route_id, DAY_TYPE, output, shift, shift_id),
            )
        con.commit()
        return route_id
    finally:
        con.close()


def _export(client, route_id, suffix=""):
    response = client.get(
        f"/api/routes/{route_id}/output-shifts/export.xlsx?day_type={DAY_TYPE}{suffix}"
    )
    assert response.status_code == 200, response.text
    return response, load_workbook(io.BytesIO(response.content))


def test_output_shift_export_has_ordered_typed_rows_and_manual_highlight(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_export_data()
    response, workbook = _export(client, route_id)
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "route_42_" in response.headers["content-disposition"]
    assert workbook.sheetnames == ["Смены выходов"]
    sheet = workbook.active
    assert sheet["A1"].value == "Смены выходов маршрута № 42"
    assert "Вокзал — Аэропорт" in sheet["A2"].value
    assert "тип дня: будни" in sheet["A2"].value
    assert "все даты" in sheet["A2"].value
    assert [cell.value for cell in sheet[3]] == [
        "Выход", "Смена", "Тип смены", "Диапазон рейсов", "Начало",
        "Окончание", "Длительность, мин", "Длительность, ч",
        "Водительских мест", "Пересмена, мин", "Источник", "Ручная",
        "Причина ручной правки", "Назначений",
    ]
    assert [sheet.cell(row, 1).value for row in (4, 5)] == [1, 2]
    assert sheet["D4"].value == "1–2"
    assert sheet["D5"].value == "7–8"
    assert sheet["E5"].value == "23:30"
    assert sheet["F5"].value == "27:00"
    assert sheet["L4"].value == "Да"
    assert sheet["M4"].value == "Закреплено диспетчером"
    assert sheet["N4"].value == 1
    assert sheet["N5"].value == 2
    for coordinate in ("A4", "B4", "G4", "H4", "I4", "J4", "N4"):
        assert sheet[coordinate].data_type == "n"
    assert sheet["G4"].value == 510
    assert sheet["H4"].value == pytest.approx(8.5)
    assert sheet["I5"].value == 2
    assert sheet["A4"].fill.fgColor.rgb in ("00FFF2CC", "FFF2CC")
    assert sheet["A5"].fill.fgColor.rgb in ("00E2F0D9", "E2F0D9")


def test_output_shift_export_is_print_ready(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_export_data()
    _, workbook = _export(client, route_id)
    sheet = workbook.active
    assert sheet.freeze_panes == "A4"
    assert sheet.auto_filter.ref == "A3:N5"
    assert sheet.print_title_rows == "$1:$3"
    assert sheet.page_setup.orientation == "landscape"
    assert str(sheet.page_setup.paperSize) == str(sheet.PAPERSIZE_A4)
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert sheet.print_area == "'Смены выходов'!$A$1:$N$5"
    assert sheet.column_dimensions["C"].width >= 18
    assert sheet.column_dimensions["M"].width >= 25
    assert sheet["A1"].fill.fgColor.rgb in ("0017365D", "17365D")
    assert sheet["A3"].font.color.rgb in ("00FFFFFF", "FFFFFF")
    assert sheet["M3"].alignment.wrap_text is True
    assert sheet.page_margins.left < 0.5


def test_output_shift_export_scopes_assignment_counts_to_service_date(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_export_data()
    _, workbook = _export(client, route_id, "&service_date=2026-07-22")
    sheet = workbook.active
    assert "дата назначений: 22.07.2026" in sheet["A2"].value
    assert sheet["N4"].value == 1
    assert sheet["N5"].value == 1


def test_output_shift_export_empty_and_validation_behaviour(tmp_path):
    client = _client(tmp_path)
    import app.db as db
    con = db.connect()
    try:
        route_id = con.execute(
            "INSERT INTO routes(number,name) VALUES('0','Пустой маршрут')"
        ).lastrowid
        con.commit()
    finally:
        con.close()
    response, workbook = _export(client, route_id)
    sheet = workbook.active
    assert sheet.max_row == 3
    assert sheet.auto_filter.ref == "A3:N3"
    assert sheet.print_area == "'Смены выходов'!$A$1:$N$3"
    assert response.status_code == 200
    assert client.get(
        f"/api/routes/{route_id}/output-shifts/export.xlsx"
    ).status_code == 400
    assert client.get(
        "/api/routes/999999/output-shifts/export.xlsx?day_type=будни"
    ).status_code == 404
    anonymous = _client(tmp_path / "anonymous", authenticated=False)
    assert anonymous.get(
        f"/api/routes/{route_id}/output-shifts/export.xlsx?day_type=будни"
    ).status_code == 401


@pytest.mark.parametrize("service_date", [
    "20260722",
    "2026-W30-3",
    "2026-02-30",
])
def test_output_shift_export_rejects_noncanonical_or_impossible_date(
    tmp_path, service_date,
):
    client = _client(tmp_path)
    route_id = _seed_export_data()
    response = client.get(
        f"/api/routes/{route_id}/output-shifts/export.xlsx?day_type={DAY_TYPE}&service_date={service_date}"
    )
    assert response.status_code == 400


def test_output_shift_export_escapes_formula_like_text(tmp_path):
    client = _client(tmp_path)
    route_id = _seed_export_data()
    import app.db as db

    con = db.connect()
    try:
        con.execute(
            "UPDATE routes SET number=?,name=? WHERE id=?",
            ("=2+2", "  +SUM(1,1)", route_id),
        )
        con.execute(
            "UPDATE shift_types SET name='@dangerous_type' WHERE code='single_8h'"
        )
        con.execute(
            "UPDATE output_shifts SET source=?,manual_reason=? "
            "WHERE route_id=? AND is_manual_locked=1",
            ("-dangerous_source", " =HYPERLINK(\"bad\")", route_id),
        )
        con.commit()
    finally:
        con.close()

    _, workbook = _export(client, route_id)
    sheet = workbook.active
    assert sheet["A1"].data_type != "f"
    assert sheet["A2"].data_type != "f"
    assert sheet["A2"].value.startswith("'  +SUM")
    assert sheet["C4"].value == "'@dangerous_type"
    assert sheet["K4"].value == "'-dangerous_source"
    assert sheet["L4"].value == "Да"
    assert sheet["M4"].value == "' =HYPERLINK(\"bad\")"
    for coordinate in ("C4", "D4", "K4", "L4", "M4"):
        assert sheet[coordinate].data_type == "s"
