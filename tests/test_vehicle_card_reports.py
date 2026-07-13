# -*- coding: utf-8 -*-
import io

from openpyxl import load_workbook

from tests.test_repair_operations_api import create_order
from tests.test_repair_requests_api import make_client


def test_vehicle_dossier_print_contains_required_sections(tmp_path):
    client, bus_id = make_client(tmp_path)
    create_order(client, bus_id)
    response = client.get(
        f"/api/repairs/vehicles/{bus_id}/print"
        "?date_from=2026-01-01&date_to=2026-12-31"
    )
    assert response.status_code == 200, response.text
    for text in (
        "ТЕХНИЧЕСКОЕ ДОСЬЕ АВТОБУСА",
        "Ремонты",
        "Запчасти",
        "Исполнители",
        "ДТП и повреждения",
        "Ответственный мастер",
    ):
        assert text in response.text
    assert "@media print" in response.text


def test_vehicle_dossier_excel_has_all_sheets_and_money_formats(tmp_path):
    client, bus_id = make_client(tmp_path)
    create_order(client, bus_id)
    response = client.get(f"/api/repairs/vehicles/{bus_id}/export.xlsx")
    assert response.status_code == 200, response.text
    wb = load_workbook(io.BytesIO(response.content), data_only=True)
    assert wb.sheetnames == [
        "Паспорт",
        "Сводка",
        "Ремонты",
        "Операции",
        "Запчасти",
        "Исполнители",
        "ДТП",
        "Повреждения",
        "ТО",
        "Затраты",
        "Фотографии",
    ]
    assert wb["Паспорт"]["A1"].value == "ТЕХНИЧЕСКОЕ ДОСЬЕ АВТОБУСА"
    assert wb["Ремонты"].freeze_panes == "A2"
    assert wb["Ремонты"].auto_filter.ref


def test_vehicle_dossier_rejects_invalid_period(tmp_path):
    client, bus_id = make_client(tmp_path)
    response = client.get(
        f"/api/repairs/vehicles/{bus_id}/print"
        "?date_from=2026-12-31&date_to=2026-01-01"
    )
    assert response.status_code == 400
