# -*- coding: utf-8 -*-
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook


DATE = "2026-07-08"


def make_client_with_order(tmp_path):
    import app.db as db
    from app.auth import ensure_admin
    from app.main import app

    db.DB_PATH = str(tmp_path / "atp-order-excel.db")
    db.init_db()
    con = db.connect()
    try:
        ensure_admin(con)
        con.execute(
            "INSERT INTO settings(key,value) VALUES('org_name',?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            ("Тестовое автотранспортное предприятие",),
        )
        driver = con.execute(
            "INSERT INTO drivers(tab_number,fio,license_number,license_issued,license_expires,snils,status) "
            "VALUES(?,?,?,?,?,?,?)",
            ("8001", "Excel Driver", "77008001", "2020-01-01", "2035-01-01", "123-456-789 01", "работает"),
        )
        driver_id = driver.lastrowid
        bus = con.execute(
            "INSERT INTO buses(garage_number,plate,brand,model,status,odometer,fuel_balance,assigned_driver_id) "
            "VALUES(?,?,?,?,?,?,?,?)",
            ("EX-1", "А777АА77", "Test", "Bus", "исправен", 1000, 70, driver_id),
        )
        bus_id = bus.lastrowid
        route = con.execute(
            "INSERT INTO routes(number,name,comm_type,transport_type,start_point,end_point,length_km,length_back_km) "
            "VALUES(?,?,?,?,?,?,?,?)",
            ("15", "Вокзал - Центр", "городское", "Регулярные перевозки", "Вокзал", "Центр", 12, 12),
        )
        route_id = route.lastrowid
        order = con.execute("INSERT INTO orders(date,status,approved_by) VALUES(?,?,?)", (DATE, "утвержден", "admin"))
        order_id = order.lastrowid
        con.execute(
            "INSERT INTO order_lines(order_id,route_id,output_number,shift_number,driver_id,bus_id,"
            "report_time,depart_depot,start_line,end_line,return_depot,shift_hours,trips_count,distance_km,"
            "planned_fuel,dispatcher_note,status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                order_id,
                route_id,
                2,
                1,
                driver_id,
                bus_id,
                "05:20",
                "05:40",
                "06:00",
                "14:00",
                "14:20",
                8.5,
                12,
                144.0,
                50.4,
                "важная отметка",
                "план",
            ),
        )
        con.commit()
    finally:
        con.close()

    client = TestClient(app)
    token = client.post("/api/login", json={"username": "admin", "password": "admin"}).json()["token"]
    client.headers.update({"Authorization": "Bearer " + token})
    return client


def test_order_export_xlsx_is_printable_form(tmp_path):
    client = make_client_with_order(tmp_path)

    response = client.get(f"/api/orders/export.xlsx?date={DATE}")

    assert response.status_code == 200, response.text
    wb = load_workbook(io.BytesIO(response.content), data_only=False)
    ws = wb.active

    assert ws["A1"].value == "Тестовое автотранспортное предприятие"
    assert ws["A2"].value == "НАРЯД НА ВЫПУСК АВТОБУСОВ"
    assert "08.07.2026" in ws["A3"].value

    header_row = next(
        row for row in range(1, 20)
        if ws.cell(row=row, column=1).value == "Маршрут"
    )
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, 19)]
    assert headers[-1] == "Отметки"
    assert ws.freeze_panes == f"A{header_row + 1}"
    assert ws.auto_filter.ref == f"A{header_row}:R{header_row + 1}"
    assert ws.page_setup.orientation == "landscape"
    assert ws.print_title_rows == f"$1:${header_row}"
    assert ws.sheet_view.showGridLines is False

    data_row = header_row + 1
    assert ws.cell(row=data_row, column=1).value == "№ 15 Вокзал - Центр"
    assert ws.cell(row=data_row, column=4).value == "Excel Driver"
    assert ws.cell(row=data_row, column=6).value == "EX-1"
    assert ws.cell(row=data_row, column=17).value == "план"
    assert ws.cell(row=data_row, column=18).value == "важная отметка"

    all_text = "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Диспетчер" in all_text
    assert "Начальник эксплуатации" in all_text
    assert "Механик" in all_text
    assert "Медработник" in all_text

