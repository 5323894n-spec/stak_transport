# -*- coding: utf-8 -*-
"""Excel-отчёты модуля выручки."""
from openpyxl import Workbook

from .route_document_xlsx import (
    apply_sheet_setup, write_table_header, write_title_band,
)
from . import revenue_service as rs

_TITLES = {
    "route": ("ВЫРУЧКА ПО МАРШРУТАМ", "route_id", "Маршрут"),
    "driver": ("ВЫРУЧКА ПО ВОДИТЕЛЯМ", "driver_id", "Водитель"),
}


def build_revenue_report(con, *, date_from, date_to, group_by="route"):
    title, key, label = _TITLES.get(group_by, _TITLES["route"])
    sheets = [
        s for s in rs.list_sheets(con, date_from=date_from, date_to=date_to)
        if s["status"] != "аннулирован"
    ]
    totals = {}
    for s in sheets:
        bucket = totals.setdefault(s[key], {"expected": 0.0, "submitted": 0.0})
        bucket["expected"] += s["expected_amount"] or 0.0
        bucket["submitted"] += s["submitted_amount"] or 0.0
    wb = Workbook()
    ws = wb.active
    ws.title = "Выручка"
    apply_sheet_setup(ws)
    write_title_band(ws, 1, title, end_col=4)
    write_table_header(
        ws, 2, (label, "Ожидаемо, руб.", "Сдано, руб.", "Разница, руб.")
    )
    row = 3
    for ident, bucket in sorted(totals.items(), key=lambda kv: str(kv[0])):
        diff = round(bucket["submitted"] - bucket["expected"], 2)
        ws.cell(row, 1, ident)
        ws.cell(row, 2, round(bucket["expected"], 2))
        ws.cell(row, 3, round(bucket["submitted"], 2))
        ws.cell(row, 4, diff)
        row += 1
    return wb


def revenue_report_filename(date_from, date_to, group_by):
    return f"Выручка_{group_by}_{date_from}_{date_to}.xlsx"
