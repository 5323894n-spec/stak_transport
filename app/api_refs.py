# -*- coding: utf-8 -*-
"""Справочники: водители, автобусы, маршруты, календарь, отсутствия, нормативы, пользователи."""
import json, datetime
from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File
from . import db
from .auth import current_user, require_write, hash_password
from .erm_import import ErmImportError, parse_erm_route_workbook
from .route_depot import normalize_erm_depot_sections
from .route_migration import migrate_route
from .xl import xlsx_response

router = APIRouter(prefix="/api")

REF = {
    "drivers": dict(section="drivers", label="Водители", fields=[
        "tab_number","fio","birth_date","division","position","license_categories","license_number",
        "license_issued","license_expires","snils","inn","phone","address","employment_type",
        "default_schedule","assigned_route_id","assigned_bus_id","driver_class","bus_type_permits",
        "hired_date","fired_date","status","med_info","training_info","restrictions","notes"]),
    "buses": dict(section="buses", label="Автобусы", fields=[
        "garage_number","plate","vin","brand","model","year","bus_class","capacity","fuel_type",
        "fuel_rate","winter_coeff","column_name","assigned_driver_id","next_to_date","osago_expires",
        "diag_card_expires","status","odometer","tank_capacity","fuel_balance","equipment"]),
    "routes": dict(section="routes", label="Маршруты", fields=[
        "number","name","comm_type","transport_type","start_point","end_point","stops","stops_back",
        "length_km","length_back_km",
        "trip_time_min","trip_time_back_min","interval_min","outputs_count","bus_types","season","work_days","notes","version","active"]),
    "absence_types": dict(section="drivers", label="Виды отсутствий", fields=["code","name","code_1c","paid"]),
}

def _get_table(table):
    if table not in REF:
        raise HTTPException(404, "Нет такого справочника")
    return REF[table]

@router.get("/refs/{table}")
def ref_list(table: str, q: str = "", user=Depends(current_user)):
    meta = _get_table(table)
    con = db.connect()
    try:
        items = db.rows(con.execute(f"SELECT * FROM {table} ORDER BY id"))
        if q:
            ql = q.lower()
            items = [i for i in items if ql in json.dumps(i, ensure_ascii=False, default=str).lower()]
        return {"items": items, "label": meta["label"]}
    finally:
        con.close()

@router.post("/refs/{table}")
def ref_create(table: str, payload: dict = Body(...), user=Depends(current_user)):
    meta = _get_table(table)
    require_write(user, meta["section"])
    con = db.connect()
    try:
        fields = [f for f in meta["fields"] if f in payload]
        if not fields: raise HTTPException(400, "Нет данных")
        sql = f"INSERT INTO {table}({','.join(fields)}) VALUES({','.join('?' * len(fields))})"
        cur = con.execute(sql, [payload[f] for f in fields])
        db.audit(con, user["username"], "создание", table, cur.lastrowid, new=payload)
        con.commit()
        return {"id": cur.lastrowid}
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(400, f"Ошибка сохранения: {e}")
    finally:
        con.close()

@router.put("/refs/{table}/{item_id}")
def ref_update(table: str, item_id: int, payload: dict = Body(...), user=Depends(current_user)):
    meta = _get_table(table)
    require_write(user, meta["section"])
    con = db.connect()
    try:
        old = db.one(con.execute(f"SELECT * FROM {table} WHERE id=?", (item_id,)))
        if not old: raise HTTPException(404, "Запись не найдена")
        fields = [f for f in meta["fields"] if f in payload]
        con.execute(f"UPDATE {table} SET {','.join(f + '=?' for f in fields)} WHERE id=?",
                    [payload[f] for f in fields] + [item_id])
        db.audit(con, user["username"], "изменение", table, item_id,
                 old={k: old.get(k) for k in fields}, new={k: payload.get(k) for k in fields})
        con.commit()
        return {"ok": True}
    finally:
        con.close()

@router.delete("/refs/{table}/{item_id}")
def ref_delete(table: str, item_id: int, user=Depends(current_user)):
    meta = _get_table(table)
    require_write(user, meta["section"])
    con = db.connect()
    try:
        old = db.one(con.execute(f"SELECT * FROM {table} WHERE id=?", (item_id,)))
        con.execute(f"DELETE FROM {table} WHERE id=?", (item_id,))
        db.audit(con, user["username"], "удаление", table, item_id, old=old)
        con.commit()
        return {"ok": True}
    finally:
        con.close()

@router.get("/refs/{table}/export.xlsx")
def ref_export(table: str, user=Depends(current_user)):
    meta = _get_table(table)
    con = db.connect()
    try:
        items = db.rows(con.execute(f"SELECT * FROM {table} ORDER BY id"))
        headers = ["id"] + meta["fields"]
        return xlsx_response(meta["label"], headers, [[i.get(h) for h in headers] for i in items],
                             filename=f"{table}.xlsx")
    finally:
        con.close()


@router.post("/import/routes/erm")
async def route_erm_import(file: UploadFile = File(...), user=Depends(current_user)):
    """Импорт маршрутной информации из ЭРМ Excel."""
    require_write(user, "routes")
    data = await file.read()
    try:
        parsed = parse_erm_route_workbook(data)
    except ErmImportError as e:
        raise HTTPException(400, str(e))
    fields = [
        "number", "name", "comm_type", "transport_type", "start_point", "end_point",
        "stops", "stops_back", "length_km", "length_back_km",
        "trip_time_min", "trip_time_back_min",
    ]
    con = db.connect()
    try:
        old = db.one(con.execute("SELECT * FROM routes WHERE number=?", (parsed["number"],)))
        notes = json.dumps({
            "source": "ЭРМ",
            "file_name": file.filename,
            "imported_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "previous_notes": old.get("notes") if old else "",
            "details": parsed["details"],
        }, ensure_ascii=False, default=str)
        payload = {k: parsed.get(k) for k in fields}
        payload["notes"] = notes
        if old:
            payload["version"] = int(old.get("version") or 1) + 1
            update_fields = list(payload.keys())
            con.execute(
                f"UPDATE routes SET {','.join(f + '=?' for f in update_fields)} WHERE id=?",
                [payload[f] for f in update_fields] + [old["id"]])
            route_id = old["id"]
            created = False
            db.audit(con, user["username"], "импорт ЭРМ", "routes", route_id,
                     old={k: old.get(k) for k in update_fields if k in old}, new=payload,
                     comment=f"обновлён маршрут № {parsed['number']}")
        else:
            payload["version"] = 1
            payload["active"] = 1
            insert_fields = list(payload.keys())
            cur = con.execute(
                f"INSERT INTO routes({','.join(insert_fields)}) VALUES({','.join('?' * len(insert_fields))})",
                [payload[f] for f in insert_fields])
            route_id = cur.lastrowid
            created = True
            db.audit(con, user["username"], "импорт ЭРМ", "routes", route_id, new=payload,
                     comment=f"создан маршрут № {parsed['number']}")
        network_result = migrate_route(con, route_id)
        normalize_erm_depot_sections(con, route_id, parsed["details"])
        con.commit()
        return {
            "route_id": route_id,
            "created": created,
            "updated": not created,
            "route": {k: payload.get(k) for k in ["number", "name", "start_point", "end_point"]},
            "summary": parsed["details"]["summary"],
            "network": network_result,
        }
    finally:
        con.close()

@router.post("/import/{table}")
async def ref_import(table: str, file: UploadFile = File(...), user=Depends(current_user)):
    """Импорт из Excel: первая строка — заголовки, совпадающие с полями справочника."""
    meta = _get_table(table)
    require_write(user, meta["section"])
    import io
    from openpyxl import load_workbook
    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True)
    except Exception:
        raise HTTPException(400, "Не удалось прочитать файл. Нужен формат .xlsx")
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(it)]
    known = [h for h in headers if h in meta["fields"]]
    if not known:
        raise HTTPException(400, f"Не найдено ни одной знакомой колонки. Ожидаются: {', '.join(meta['fields'])}")
    key = meta["fields"][0]  # натуральный ключ — первое поле (tab_number/garage_number/number)
    con = db.connect()
    added, skipped, errors = 0, 0, []
    try:
        for rn, row in enumerate(it, start=2):
            rec = {h: row[i] for i, h in enumerate(headers) if h in meta["fields"] and i < len(row)}
            if not any(v is not None and str(v).strip() for v in rec.values()):
                continue
            try:
                if key in rec and con.execute(f"SELECT 1 FROM {table} WHERE {key}=?", (str(rec[key]),)).fetchone():
                    skipped += 1
                    continue
                fields = list(rec.keys())
                con.execute(f"INSERT INTO {table}({','.join(fields)}) VALUES({','.join('?' * len(fields))})",
                            [str(rec[f]) if rec[f] is not None else None for f in fields])
                added += 1
            except Exception as e:
                errors.append(f"строка {rn}: {e}")
        db.audit(con, user["username"], "импорт Excel", table, None,
                 comment=f"добавлено {added}, пропущено дублей {skipped}, ошибок {len(errors)}")
        con.commit()
    finally:
        con.close()
    return {"added": added, "skipped": skipped, "errors": errors[:50]}

# ---------- Производственный календарь ----------
RF_HOLIDAYS = {(1, d) for d in range(1, 9)} | {(2, 23), (3, 8), (5, 1), (5, 9), (6, 12), (11, 4)}

@router.get("/calendar")
def calendar_get(year: int, user=Depends(current_user)):
    con = db.connect()
    try:
        stored = {r["date"]: r["day_type"] for r in
                  con.execute("SELECT * FROM calendar WHERE date LIKE ?", (f"{year}-%",))}
        out = {}
        d = datetime.date(year, 1, 1)
        while d.year == year:
            iso = d.isoformat()
            if iso in stored:
                out[iso] = stored[iso]
            elif (d.month, d.day) in RF_HOLIDAYS:
                out[iso] = "праздник"
            elif d.weekday() >= 5:
                out[iso] = "выходной"
            else:
                out[iso] = "рабочий"
            d += datetime.timedelta(days=1)
        return out
    finally:
        con.close()

@router.post("/calendar")
def calendar_set(payload: dict = Body(...), user=Depends(current_user)):
    require_write(user, "routes")
    con = db.connect()
    try:
        con.execute("INSERT INTO calendar(date, day_type, comment) VALUES(?,?,?) "
                    "ON CONFLICT(date) DO UPDATE SET day_type=excluded.day_type, comment=excluded.comment",
                    (payload["date"], payload["day_type"], payload.get("comment", "")))
        db.audit(con, user["username"], "изменение календаря", "calendar", payload["date"], new=payload)
        con.commit()
        return {"ok": True}
    finally:
        con.close()

# ---------- Отсутствия ----------
@router.get("/absences")
def absences_list(date_from: str = "", date_to: str = "", driver_id: int = 0, user=Depends(current_user)):
    con = db.connect()
    try:
        q = ("SELECT a.*, d.fio, d.tab_number, t.name AS type_name FROM absences a "
             "JOIN drivers d ON d.id=a.driver_id LEFT JOIN absence_types t ON t.code=a.type_code WHERE 1=1")
        args = []
        if date_from: q += " AND a.date_to>=?"; args.append(date_from)
        if date_to: q += " AND a.date_from<=?"; args.append(date_to)
        if driver_id: q += " AND a.driver_id=?"; args.append(driver_id)
        return {"items": db.rows(con.execute(q + " ORDER BY a.date_from DESC", args))}
    finally:
        con.close()

@router.post("/absences")
def absence_create(payload: dict = Body(...), user=Depends(current_user)):
    if user["role"] not in ("админ", "кадры", "диспетчер"):
        raise HTTPException(403, "Нет права оформлять отсутствия")
    con = db.connect()
    try:
        cur = con.execute("INSERT INTO absences(driver_id,type_code,date_from,date_to,status,comment) VALUES(?,?,?,?,?,?)",
            (payload["driver_id"], payload["type_code"], payload["date_from"], payload["date_to"],
             payload.get("status", "утверждено"), payload.get("comment", "")))
        # исключаем водителя из графика на период
        tname = db.one(con.execute("SELECT name FROM absence_types WHERE code=?", (payload["type_code"],)))
        status = payload["type_code"]
        d = datetime.date.fromisoformat(payload["date_from"])
        d1 = datetime.date.fromisoformat(payload["date_to"])
        affected = 0
        while d <= d1:
            r = con.execute("SELECT id, status FROM roster WHERE driver_id=? AND date=?",
                            (payload["driver_id"], d.isoformat())).fetchone()
            if r and r["status"] == "работа": affected += 1
            con.execute("INSERT INTO roster(driver_id,date,status,comment) VALUES(?,?,?,?) "
                        "ON CONFLICT(driver_id,date) DO UPDATE SET status=excluded.status, "
                        "route_id=NULL, output_number=NULL, shift_number=NULL, start_time=NULL, end_time=NULL, "
                        "hours=0, night_hours=0, comment=excluded.comment",
                        (payload["driver_id"], d.isoformat(), status,
                         (tname or {}).get("name", status)))
            d += datetime.timedelta(days=1)
        drv = db.one(con.execute("SELECT fio FROM drivers WHERE id=?", (payload["driver_id"],)))
        if affected:
            db.notify(con, "warning", "график",
                      f"{drv['fio']}: отсутствие {payload['date_from']}—{payload['date_to']} сняло {affected} смен из графика — подберите замену")
        db.audit(con, user["username"], "оформление отсутствия", "absences", cur.lastrowid, new=payload)
        con.commit()
        return {"id": cur.lastrowid, "roster_affected": affected}
    finally:
        con.close()

@router.delete("/absences/{aid}")
def absence_delete(aid: int, user=Depends(current_user)):
    if user["role"] not in ("админ", "кадры", "диспетчер"):
        raise HTTPException(403, "Нет права")
    con = db.connect()
    try:
        old = db.one(con.execute("SELECT * FROM absences WHERE id=?", (aid,)))
        if old:
            d = datetime.date.fromisoformat(old["date_from"])
            d1 = datetime.date.fromisoformat(old["date_to"])
            while d <= d1:
                con.execute("UPDATE roster SET status='выходной', comment='' WHERE driver_id=? AND date=? AND status=?",
                            (old["driver_id"], d.isoformat(), old["type_code"]))
                d += datetime.timedelta(days=1)
        con.execute("DELETE FROM absences WHERE id=?", (aid,))
        db.audit(con, user["username"], "удаление отсутствия", "absences", aid, old=old)
        con.commit()
        return {"ok": True}
    finally:
        con.close()

# ---------- Нормативы ----------
@router.get("/norms")
def norms_list(user=Depends(current_user)):
    con = db.connect()
    try:
        items = db.rows(con.execute("SELECT * FROM norms ORDER BY valid_from DESC"))
        for i in items:
            i["params"] = json.loads(i["params"])
        return {"items": items, "defaults": db.DEFAULT_NORMS}
    finally:
        con.close()

@router.post("/norms")
def norms_save(payload: dict = Body(...), user=Depends(current_user)):
    if user["role"] != "админ":
        raise HTTPException(403, "Нормативы меняет только администратор")
    con = db.connect()
    try:
        params = json.dumps(payload.get("params", {}), ensure_ascii=False)
        if payload.get("id"):
            old = db.one(con.execute("SELECT * FROM norms WHERE id=?", (payload["id"],)))
            con.execute("UPDATE norms SET name=?, valid_from=?, valid_to=?, params=?, doc_ref=?, comment=?, active=? WHERE id=?",
                        (payload["name"], payload["valid_from"], payload["valid_to"], params,
                         payload.get("doc_ref", ""), payload.get("comment", ""), payload.get("active", 1), payload["id"]))
            db.audit(con, user["username"], "изменение нормативов", "norms", payload["id"], old=old, new=payload)
        else:
            cur = con.execute("INSERT INTO norms(name,valid_from,valid_to,params,doc_ref,comment,active) VALUES(?,?,?,?,?,?,?)",
                        (payload["name"], payload["valid_from"], payload["valid_to"], params,
                         payload.get("doc_ref", ""), payload.get("comment", ""), payload.get("active", 1)))
            db.audit(con, user["username"], "создание версии нормативов", "norms", cur.lastrowid, new=payload)
        con.commit()
        return {"ok": True}
    finally:
        con.close()

# ---------- Пользователи, настройки, коды времени, аудит, уведомления ----------
@router.get("/users")
def users_list(user=Depends(current_user)):
    if user["role"] != "админ": raise HTTPException(403, "Только администратор")
    con = db.connect()
    try:
        return {"items": db.rows(con.execute("SELECT id, username, full_name, role, active FROM users"))}
    finally:
        con.close()

@router.post("/users")
def user_save(payload: dict = Body(...), user=Depends(current_user)):
    if user["role"] != "админ": raise HTTPException(403, "Только администратор")
    con = db.connect()
    try:
        if payload.get("id"):
            con.execute("UPDATE users SET username=?, full_name=?, role=?, active=? WHERE id=?",
                        (payload["username"], payload.get("full_name", ""), payload["role"],
                         payload.get("active", 1), payload["id"]))
            if payload.get("password"):
                con.execute("UPDATE users SET password_hash=? WHERE id=?",
                            (hash_password(payload["password"]), payload["id"]))
            db.audit(con, user["username"], "изменение пользователя", "users", payload["id"])
        else:
            con.execute("INSERT INTO users(username,password_hash,full_name,role,active) VALUES(?,?,?,?,?)",
                        (payload["username"], hash_password(payload.get("password", "12345")),
                         payload.get("full_name", ""), payload["role"], payload.get("active", 1)))
            db.audit(con, user["username"], "создание пользователя", "users", payload["username"])
        con.commit()
        return {"ok": True}
    finally:
        con.close()

@router.get("/settings")
def settings_get(user=Depends(current_user)):
    con = db.connect()
    try:
        return db.get_settings(con)
    finally:
        con.close()

@router.post("/settings")
def settings_set(payload: dict = Body(...), user=Depends(current_user)):
    if user["role"] != "админ": raise HTTPException(403, "Только администратор")
    if "repair_repeat_days" in payload:
        try:
            repeat_days = int(payload["repair_repeat_days"])
        except (TypeError, ValueError):
            raise HTTPException(400, "Период повторной неисправности должен быть целым числом")
        if not 1 <= repeat_days <= 365:
            raise HTTPException(400, "Период повторной неисправности должен быть от 1 до 365 дней")
        payload["repair_repeat_days"] = repeat_days
    con = db.connect()
    try:
        for k, v in payload.items():
            con.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                        (k, str(v)))
        db.audit(con, user["username"], "изменение настроек", "settings", None, new=payload)
        con.commit()
        return {"ok": True}
    finally:
        con.close()

@router.get("/time-codes")
def time_codes(user=Depends(current_user)):
    con = db.connect()
    try:
        return {"items": db.rows(con.execute("SELECT * FROM time_codes"))}
    finally:
        con.close()

@router.post("/time-codes")
def time_codes_save(payload: dict = Body(...), user=Depends(current_user)):
    if user["role"] not in ("админ", "бухгалтер"): raise HTTPException(403, "Нет права")
    con = db.connect()
    try:
        con.execute("UPDATE time_codes SET code_1c=?, name=? WHERE code=?",
                    (payload["code_1c"], payload.get("name"), payload["code"]))
        db.audit(con, user["username"], "изменение кода времени", "time_codes", payload["code"], new=payload)
        con.commit()
        return {"ok": True}
    finally:
        con.close()

@router.get("/audit")
def audit_list(date_from: str = "", date_to: str = "", username: str = "", object_type: str = "",
               limit: int = 300, user=Depends(current_user)):
    if user["role"] not in ("админ", "руководитель"): raise HTTPException(403, "Нет права")
    con = db.connect()
    try:
        q, args = "SELECT * FROM audit_log WHERE 1=1", []
        if date_from: q += " AND ts>=?"; args.append(date_from)
        if date_to: q += " AND ts<=?"; args.append(date_to + "T23:59:59")
        if username: q += " AND username=?"; args.append(username)
        if object_type: q += " AND object_type=?"; args.append(object_type)
        return {"items": db.rows(con.execute(q + " ORDER BY id DESC LIMIT ?", args + [limit]))}
    finally:
        con.close()

@router.get("/notifications")
def notifications(user=Depends(current_user)):
    con = db.connect()
    try:
        return {"items": db.rows(con.execute("SELECT * FROM notifications ORDER BY id DESC LIMIT 100"))}
    finally:
        con.close()

@router.post("/notifications/seen")
def notifications_seen(user=Depends(current_user)):
    con = db.connect()
    try:
        con.execute("UPDATE notifications SET seen=1")
        con.commit()
        return {"ok": True}
    finally:
        con.close()
