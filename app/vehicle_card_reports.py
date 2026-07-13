# -*- coding: utf-8 -*-
"""Печатное техническое досье автобуса и его Excel-выгрузка."""
import datetime
import html
import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import db
from .auth import current_user
from .repair_service import audit_change, require_repair_action

router = APIRouter(prefix="/api/repairs/vehicles", tags=["vehicle-card-reports"])
HEADER = PatternFill("solid", fgColor="17365D")
THIN = Side(style="thin", color="B8C6D5")
MONEY = "#,##0.00"


def _period(date_from, date_to):
    try:
        start = datetime.date.fromisoformat(date_from) if date_from else None
        end = datetime.date.fromisoformat(date_to) if date_to else None
    except ValueError:
        raise HTTPException(400, "Неверный формат периода отчёта")
    if start and end and start > end:
        raise HTTPException(400, "Начало периода отчёта позже окончания")
    return start, end


def _date_filter(column, start, end):
    conditions, args = [], []
    if start:
        conditions.append(f"{column}>=?")
        args.append(start.isoformat() + "T00:00:00")
    if end:
        conditions.append(f"{column}<=?")
        args.append(end.isoformat() + "T23:59:59")
    return conditions, args


def collect_vehicle_dossier(con, bus_id, date_from="", date_to=""):
    start, end = _period(date_from, date_to)
    vehicle = db.one(con.execute(
        "SELECT b.*,r.name assigned_route_name FROM buses b "
        "LEFT JOIN routes r ON r.id=b.assigned_route_id WHERE b.id=?",
        (bus_id,),
    ))
    if not vehicle:
        raise HTTPException(404, "Автобус не найден")

    def selected(sql, bus_column, date_column, tail=""):
        conditions, args = [f"{bus_column}=?"], [bus_id]
        date_conditions, date_args = _date_filter(date_column, start, end)
        conditions.extend(date_conditions)
        args.extend(date_args)
        return db.rows(con.execute(
            sql + " WHERE " + " AND ".join(conditions) + tail,
            args,
        ))

    repairs = selected(
        "SELECT ro.*,ro.number order_number,rr.number request_number,"
        "rt.name repair_type_name,u.full_name master_name FROM repair_orders ro "
        "LEFT JOIN repair_requests rr ON rr.id=ro.request_id "
        "LEFT JOIN repair_types rt ON rt.id=ro.repair_type_id "
        "LEFT JOIN users u ON u.id=ro.responsible_master_id",
        "ro.bus_id", "ro.created_at", " ORDER BY ro.created_at DESC,ro.id DESC",
    )
    operations = selected(
        "SELECT ro.number order_number,o.sequence_no,o.name,o.status,o.norm_hours,"
        "o.actual_hours,o.price,o.result FROM repair_operations o "
        "JOIN repair_orders ro ON ro.id=o.order_id",
        "ro.bus_id", "ro.created_at", " ORDER BY ro.created_at DESC,o.sequence_no,o.id",
    )
    parts = selected(
        "SELECT ro.number order_number,p.code,p.name,p.unit,rp.requested_qty,"
        "rp.issued_qty,rp.installed_qty,rp.returned_qty,rp.unit_price,"
        "rp.installed_qty*rp.unit_price line_cost,rp.status FROM repair_parts rp "
        "JOIN repair_orders ro ON ro.id=rp.order_id JOIN parts p ON p.id=rp.part_id",
        "ro.bus_id", "ro.created_at", " ORDER BY ro.created_at DESC,rp.id",
    )
    workers = selected(
        "SELECT ro.number order_number,u.full_name,rw.role,rw.status,rw.planned_hours,"
        "rw.actual_hours,rw.hourly_rate,rw.actual_hours*rw.hourly_rate labor_cost "
        "FROM repair_order_workers rw JOIN repair_orders ro ON ro.id=rw.order_id "
        "JOIN users u ON u.id=rw.worker_id",
        "ro.bus_id", "ro.created_at", " ORDER BY ro.created_at DESC,rw.id",
    )
    incidents = selected(
        "SELECT vi.*,u.full_name responsible_name FROM vehicle_incidents vi "
        "LEFT JOIN users u ON u.id=vi.responsible_user_id",
        "vi.bus_id", "vi.occurred_at", " ORDER BY vi.occurred_at DESC,vi.id DESC",
    )
    damages = selected(
        "SELECT vd.*,vi.occurred_at,vi.incident_type FROM vehicle_damages vd "
        "JOIN vehicle_incidents vi ON vi.id=vd.incident_id",
        "vi.bus_id", "vi.occurred_at", " ORDER BY vi.occurred_at DESC,vd.id",
    )
    maintenance = db.rows(con.execute(
        "SELECT mp.name,rt.name repair_type,mp.last_date,mp.last_odometer,"
        "mp.next_date,mp.next_odometer,mp.interval_days,mp.interval_km,mp.active "
        "FROM maintenance_plans mp LEFT JOIN repair_types rt ON rt.id=mp.repair_type_id "
        "WHERE mp.bus_id=? ORDER BY mp.active DESC,mp.next_date",
        (bus_id,),
    ))
    media = selected(
        "SELECT id,category,caption,captured_at,original_name,mime_type,size_bytes,"
        "is_cover,uploaded_at FROM repair_attachments",
        "bus_id", "COALESCE(captured_at,uploaded_at)",
        " AND cancelled_at IS NULL ORDER BY is_cover DESC,COALESCE(captured_at,uploaded_at) DESC,id DESC",
    )
    for item in media:
        item["download_url"] = f"/api/repairs/attachments/{item['id']}/download"

    costs = [{
        "period": (row.get("closed_at") or row.get("created_at") or "")[:7],
        "order_number": row["order_number"],
        "labor_cost": row.get("labor_cost") or 0,
        "parts_cost": row.get("parts_cost") or 0,
        "external_cost": row.get("external_cost") or 0,
        "other_cost": row.get("other_cost") or 0,
        "total_cost": row.get("total_cost") or 0,
    } for row in repairs]
    summary = {
        "repairs": len(repairs),
        "total_cost": sum(float(row["total_cost"] or 0) for row in repairs),
        "downtime_hours": sum(float(row["downtime_hours"] or 0) for row in repairs),
        "incidents": len(incidents),
        "open_damages": sum(1 for row in damages if not row["resolved"]),
        "date_from": date_from,
        "date_to": date_to,
    }
    return {
        "vehicle": vehicle, "summary": summary, "repairs": repairs,
        "operations": operations, "parts": parts, "workers": workers,
        "incidents": incidents, "damages": damages, "maintenance": maintenance,
        "costs": costs, "media": media,
    }


def _e(value):
    return html.escape(str(value if value is not None else ""))


def _html_table(headers, rows):
    head = "".join(f"<th>{_e(header)}</th>" for header in headers)
    body = "".join(
        "<tr>" + "".join(f"<td>{_e(value)}</td>" for value in row) + "</tr>"
        for row in rows
    )
    if not body:
        body = f"<tr><td colspan='{len(headers)}'>Нет данных</td></tr>"
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def render_vehicle_dossier(data):
    v, s = data["vehicle"], data["summary"]
    period = " — ".join(x for x in (s["date_from"], s["date_to"]) if x) or "за всё время"
    cover = next((item for item in data["media"] if item["is_cover"]), None)
    cover_html = f"<img class='cover' src='{_e(cover['download_url'])}' alt='Фото автобуса'>" if cover else ""
    repairs = _html_table(
        ["Заказ-наряд", "Дата", "Вид", "Статус", "Ответственный мастер", "Результат", "Стоимость"],
        ([r["order_number"], r["created_at"], r["repair_type_name"], r["status"], r["master_name"], r["result"], r["total_cost"]] for r in data["repairs"]),
    )
    parts = _html_table(
        ["Заказ-наряд", "Код", "Запчасть", "Установлено", "Цена", "Сумма"],
        ([r["order_number"], r["code"], r["name"], r["installed_qty"], r["unit_price"], r["line_cost"]] for r in data["parts"]),
    )
    workers = _html_table(
        ["Заказ-наряд", "Исполнитель", "Роль", "Часы", "Стоимость работ"],
        ([r["order_number"], r["full_name"], r["role"], r["actual_hours"], r["labor_cost"]] for r in data["workers"]),
    )
    incidents = _html_table(
        ["Дата", "Тип", "Место", "Обстоятельства", "Ответственный", "Ущерб", "Статус"],
        ([r["occurred_at"], r["incident_type"], r["place"], r["circumstances"], r["responsible_name"], r["actual_damage_cost"], r["status"]] for r in data["incidents"]),
    )
    damages = _html_table(
        ["Дата", "Зона", "Описание", "Тяжесть", "Устранено"],
        ([r["occurred_at"], r["area"], r["description"], r["severity"], "Да" if r["resolved"] else "Нет"] for r in data["damages"]),
    )
    maintenance = _html_table(
        ["План", "Вид ТО", "Последняя дата", "Следующая дата", "Следующий пробег"],
        ([r["name"], r["repair_type"], r["last_date"], r["next_date"], r["next_odometer"]] for r in data["maintenance"]),
    )
    return f"""<!doctype html><html lang='ru'><head><meta charset='utf-8'>
<title>Техническое досье {_e(v['garage_number'])}</title><style>
@page{{size:A4 landscape;margin:10mm}}body{{font:11px Arial;color:#172033}}h1{{font-size:20px;color:#17365d}}h2{{font-size:15px;color:#17365d;border-bottom:2px solid #17365d;padding-bottom:3px}}.toolbar{{text-align:right}}.cover{{max-width:230px;max-height:150px;object-fit:cover;float:right;margin:0 0 12px 18px}}.passport{{display:grid;grid-template-columns:repeat(3,1fr);gap:6px 18px}}.summary{{display:flex;gap:10px;margin:14px 0}}.metric{{background:#eef4fb;padding:9px;border-radius:5px;min-width:110px}}table{{width:100%;border-collapse:collapse;font-size:9px;margin-bottom:12px}}th,td{{border:1px solid #9aa9ba;padding:4px;vertical-align:top}}th{{background:#17365d;color:white}}section,.event{{break-inside:avoid}}.signatures{{display:flex;justify-content:space-between;margin-top:30px}}@media print{{.toolbar{{display:none}}a{{color:inherit;text-decoration:none}}}}
</style></head><body><div class='toolbar'><button onclick='print()'>Печать / сохранить PDF</button></div>{cover_html}
<h1>ТЕХНИЧЕСКОЕ ДОСЬЕ АВТОБУСА</h1><p>Период отчёта: {_e(period)}</p>
<div class='passport'><div><b>Гаражный №:</b> {_e(v['garage_number'])}</div><div><b>Госномер:</b> {_e(v['plate'])}</div><div><b>VIN:</b> {_e(v['vin'])}</div><div><b>Марка/модель:</b> {_e(v['brand'])} {_e(v['model'])}</div><div><b>Пробег:</b> {_e(v['odometer'])}</div><div><b>Статус:</b> {_e(v['status'])}</div></div>
<div class='summary'><div class='metric'><b>Ремонтов</b><br>{s['repairs']}</div><div class='metric'><b>Затраты</b><br>{s['total_cost']:.2f}</div><div class='metric'><b>Простой, ч</b><br>{s['downtime_hours']:.2f}</div><div class='metric'><b>ДТП/событий</b><br>{s['incidents']}</div><div class='metric'><b>Открытых повреждений</b><br>{s['open_damages']}</div></div>
<section><h2>Ремонты</h2>{repairs}</section><section><h2>Запчасти</h2>{parts}</section><section><h2>Исполнители</h2>{workers}</section><section><h2>ДТП и повреждения</h2>{incidents}{damages}</section><section><h2>Техническое обслуживание</h2>{maintenance}</section>
<div class='signatures'><span>Ответственный мастер __________________</span><span>Проверил __________________</span><span>Дата __________________</span></div></body></html>"""


def _table_sheet(wb, title, headers, rows, money_columns=()):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for index in money_columns:
            row[index - 1].number_format = MONEY
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, 1):
        width = max((len(str(cell.value or "")) for cell in column), default=8) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(max(width, 10), 42)
    return ws


def build_vehicle_workbook(data):
    v, s = data["vehicle"], data["summary"]
    wb = Workbook()
    passport = wb.active
    passport.title = "Паспорт"
    passport["A1"] = "ТЕХНИЧЕСКОЕ ДОСЬЕ АВТОБУСА"
    passport["A1"].font = Font(size=16, bold=True, color="17365D")
    passport.append(["Поле", "Значение"])
    for label, value in (("Гаражный №", v["garage_number"]), ("Госномер", v["plate"]), ("Марка", v["brand"]), ("Модель", v["model"]), ("VIN", v["vin"]), ("Пробег", v["odometer"]), ("Статус", v["status"]), ("Маршрут", v["assigned_route_name"])):
        passport.append([label, value])
    passport.column_dimensions["A"].width = 25
    passport.column_dimensions["B"].width = 42
    summary = wb.create_sheet("Сводка")
    summary.append(["Показатель", "Значение"])
    for label, value in (("Период с", s["date_from"]), ("Период по", s["date_to"]), ("Ремонтов", s["repairs"]), ("Общие затраты", s["total_cost"]), ("Простой, часов", s["downtime_hours"]), ("ДТП/событий", s["incidents"]), ("Открытых повреждений", s["open_damages"])):
        summary.append([label, value])
    summary["B5"].number_format = MONEY
    _table_sheet(wb, "Ремонты", ["Заказ-наряд", "Заявка", "Создан", "Закрыт", "Вид ремонта", "Статус", "Ответственный мастер", "Работы", "Запчасти", "Внешние", "Прочие", "Итого", "Результат"], ([r["order_number"], r["request_number"], r["created_at"], r["closed_at"], r["repair_type_name"], r["status"], r["master_name"], r["labor_cost"], r["parts_cost"], r["external_cost"], r["other_cost"], r["total_cost"], r["result"]] for r in data["repairs"]), range(8, 13))
    _table_sheet(wb, "Операции", ["Заказ-наряд", "№", "Операция", "Статус", "Норма, ч", "Факт, ч", "Цена", "Результат"], ([r[k] for k in ("order_number", "sequence_no", "name", "status", "norm_hours", "actual_hours", "price", "result")] for r in data["operations"]), (5, 6, 7))
    _table_sheet(wb, "Запчасти", ["Заказ-наряд", "Код", "Запчасть", "Ед.", "Запрошено", "Выдано", "Установлено", "Возвращено", "Цена", "Сумма", "Статус"], ([r[k] for k in ("order_number", "code", "name", "unit", "requested_qty", "issued_qty", "installed_qty", "returned_qty", "unit_price", "line_cost", "status")] for r in data["parts"]), (9, 10))
    _table_sheet(wb, "Исполнители", ["Заказ-наряд", "Исполнитель", "Роль", "Статус", "План, ч", "Факт, ч", "Ставка", "Стоимость"], ([r[k] for k in ("order_number", "full_name", "role", "status", "planned_hours", "actual_hours", "hourly_rate", "labor_cost")] for r in data["workers"]), (5, 6, 7, 8))
    _table_sheet(wb, "ДТП", ["Дата", "Тип", "Место", "Обстоятельства", "Виновность", "Полиция", "Страховой случай", "Ответственный", "Оценка", "Факт", "Статус"], ([r[k] for k in ("occurred_at", "incident_type", "place", "circumstances", "fault_status", "police_document_number", "insurance_case_number", "responsible_name", "estimated_damage_cost", "actual_damage_cost", "status")] for r in data["incidents"]), (9, 10))
    _table_sheet(wb, "Повреждения", ["Дата", "Тип события", "Зона", "Описание", "Тяжесть", "Нужен ремонт", "Устранено", "Дата устранения"], ([r["occurred_at"], r["incident_type"], r["area"], r["description"], r["severity"], "Да" if r["repair_required"] else "Нет", "Да" if r["resolved"] else "Нет", r["resolved_at"]] for r in data["damages"]))
    _table_sheet(wb, "ТО", ["План", "Вид ТО", "Последняя дата", "Последний пробег", "Следующая дата", "Следующий пробег", "Интервал, дней", "Интервал, км", "Активен"], ([r[k] for k in ("name", "repair_type", "last_date", "last_odometer", "next_date", "next_odometer", "interval_days", "interval_km", "active")] for r in data["maintenance"]))
    _table_sheet(wb, "Затраты", ["Период", "Заказ-наряд", "Работы", "Запчасти", "Внешние", "Прочие", "Итого"], ([r[k] for k in ("period", "order_number", "labor_cost", "parts_cost", "external_cost", "other_cost", "total_cost")] for r in data["costs"]), (3, 4, 5, 6, 7))
    _table_sheet(wb, "Фотографии", ["Категория", "Подпись", "Дата съёмки", "Файл", "Тип", "Размер", "Обложка", "Ссылка"], ([r["category"], r["caption"], r["captured_at"], r["original_name"], r["mime_type"], r["size_bytes"], "Да" if r["is_cover"] else "Нет", r["download_url"]] for r in data["media"]))
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


@router.get("/{bus_id}/print", response_class=HTMLResponse)
def print_vehicle_card(bus_id: int, date_from: str = "", date_to: str = "", user=Depends(current_user)):
    require_repair_action(user, "read_reports")
    con = db.connect()
    try:
        data = collect_vehicle_dossier(con, bus_id, date_from, date_to)
        audit_change(
            con, user, "печать технического досье автобуса", "vehicle_dossier",
            bus_id, new={"date_from": date_from, "date_to": date_to},
        )
        con.commit()
    finally:
        con.close()
    return HTMLResponse(render_vehicle_dossier(data))


@router.get("/{bus_id}/export.xlsx")
def export_vehicle_card(bus_id: int, date_from: str = "", date_to: str = "", user=Depends(current_user)):
    require_repair_action(user, "read_reports")
    con = db.connect()
    try:
        data = collect_vehicle_dossier(con, bus_id, date_from, date_to)
        audit_change(
            con, user, "экспорт технического досье автобуса", "vehicle_dossier",
            bus_id, new={"date_from": date_from, "date_to": date_to},
        )
        con.commit()
    finally:
        con.close()
    stream = build_vehicle_workbook(data)
    filename = f"vehicle_{bus_id}_dossier.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
