# -*- coding: utf-8 -*-
"""Сохранённая матрица времени рейсов по остановкам."""

import datetime

import json
import secrets
import sqlite3
from fastapi import APIRouter, Body, Depends, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


from . import db
from .auth import current_user, require_write
from .route_periods import calculate_period_preview
from .route_timetable import (
    adjust_stop_times, build_schedule_preview, calculate_trip_stop_times,
    format_service_time,
)
from .xl import _xlsx_download_response


router = APIRouter(prefix="/api")

DIRECTION_TO_TRACE = {
    "прямое": "forward",
    "forward": "forward",
    "обратное": "backward",
    "backward": "backward",
}


def service_time_to_seconds(value):
    try:
        hours, minutes = map(int, str(value).strip().split(":"))
    except (TypeError, ValueError):
        raise ValueError("Некорректное время отправления рейса")
    if hours < 0 or hours >= 48 or not 0 <= minutes < 60:
        raise ValueError("Некорректное время отправления рейса")
    return hours * 3600 + minutes * 60


def _trace_rows(con, route_id, direction):
    return db.rows(
        con.execute(
            "SELECT rs.*,s.name AS stop_name,s.external_code AS stop_code "
            "FROM route_stops rs JOIN stops s ON s.id=rs.stop_id "
            "WHERE rs.route_id=? AND rs.direction=? ORDER BY rs.sequence,rs.id",
            (route_id, direction),
        )
    )


def _period_for_trip(con, trip, departure_sec):
    if trip.get("period_id"):
        return db.one(
            con.execute("SELECT * FROM day_periods WHERE id=?", (trip["period_id"],))
        )
    departure_min = departure_sec // 60
    return db.one(
        con.execute(
            "SELECT * FROM day_periods WHERE route_id=? AND day_type=? "
            "AND active=1 AND start_min<=? AND end_min>? "
            "ORDER BY priority,start_min,id LIMIT 1",
            (trip["route_id"], trip["day_type"], departure_min, departure_min),
        )
    )


def _runtime_overrides(con, period_id):
    if not period_id:
        return {}
    return {
        row["route_stop_id"]: row["run_time_sec"]
        for row in con.execute(
            "SELECT route_stop_id,run_time_sec FROM route_stop_runtimes "
            "WHERE period_id=?",
            (period_id,),
        )
    }


def recalculate_trip_stop_times_in_connection(con, trip_id, preserve_manual=True):
    trip = db.one(con.execute("SELECT * FROM route_trips WHERE id=?", (trip_id,)))
    if not trip:
        raise HTTPException(404, "Рейс не найден")
    existing = db.rows(
        con.execute(
            "SELECT * FROM trip_stop_times WHERE trip_id=? ORDER BY sequence,id",
            (trip_id,),
        )
    )
    if preserve_manual and any(row["is_manual_override"] for row in existing):
        return existing

    trace_direction = DIRECTION_TO_TRACE.get(trip.get("direction"))
    if not trace_direction:
        raise ValueError("Неизвестное направление рейса")
    trace = _trace_rows(con, trip["route_id"], trace_direction)
    departure_sec = service_time_to_seconds(trip.get("dep_time"))
    period = _period_for_trip(con, trip, departure_sec)
    rows = calculate_trip_stop_times(
        trace,
        departure_sec=departure_sec,
        runtime_factor=(period or {}).get("travel_time_factor", 1),
        runtime_overrides=_runtime_overrides(con, (period or {}).get("id")),
    )

    timestamp = datetime.datetime.now().isoformat(timespec="seconds")
    con.execute("DELETE FROM trip_stop_times WHERE trip_id=?", (trip_id,))
    for row in rows:
        con.execute(
            "INSERT INTO trip_stop_times(trip_id,route_stop_id,sequence,arrival_sec,"
            "departure_sec,is_timing_point,is_manual_override,created_at,updated_at) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            (
                trip_id,
                row["route_stop_id"],
                row["sequence"],
                row["arrival_sec"],
                row["departure_sec"],
                row["is_timing_point"],
                0,
                timestamp,
                timestamp,
            ),
        )
    if period and not trip.get("period_id"):
        con.execute("UPDATE route_trips SET period_id=? WHERE id=?", (period["id"], trip_id))
    con.execute(
        "UPDATE route_trips SET arr_time=? WHERE id=?",
        (format_service_time(rows[-1]["arrival_sec"]), trip_id),
    )
    return db.rows(
        con.execute(
            "SELECT * FROM trip_stop_times WHERE trip_id=? ORDER BY sequence,id",
            (trip_id,),
        )
    )


def _serialized_time_row(row):
    return {
        **row,
        "arrival_time": format_service_time(row["arrival_sec"]),
        "departure_time": format_service_time(row["departure_sec"]),
        "is_timing_point": bool(row["is_timing_point"]),
        "is_manual_override": bool(row["is_manual_override"]),
    }


def _even_distribution(trace, trip, departure_sec):
    """Spread the trip's dep→arr span evenly across stops when leg run times are
    not defined, so the per-stop schedule still renders."""
    arrival_sec = service_time_to_seconds(trip.get("arr_time"))
    span = (
        arrival_sec - departure_sec
        if arrival_sec is not None and arrival_sec > departure_sec
        else 0
    )
    last = len(trace) - 1
    rows = []
    for index, stop in enumerate(trace):
        offset = round(span * index / last) if last > 0 and span else 0
        moment = departure_sec + offset
        rows.append({
            "route_stop_id": stop["id"],
            "sequence": int(stop["sequence"]),
            "arrival_sec": moment,
            "departure_sec": moment,
            "is_timing_point": 1 if stop.get("is_timing_point") else 0,
        })
    return rows


def _display_stop_times(con, trip, trace):
    """Return per-stop times for the matrix: stored rows (preserving manual
    edits) when present, otherwise computed on the fly so the schedule shows."""
    stored = db.rows(
        con.execute(
            "SELECT * FROM trip_stop_times WHERE trip_id=? ORDER BY sequence,id",
            (trip["id"],),
        )
    )
    if stored:
        return [_serialized_time_row(row) for row in stored]
    if not trace:
        return []
    departure_sec = service_time_to_seconds(trip.get("dep_time"))
    if departure_sec is None:
        return []
    period = _period_for_trip(con, trip, departure_sec)
    try:
        rows = calculate_trip_stop_times(
            trace,
            departure_sec=departure_sec,
            runtime_factor=(period or {}).get("travel_time_factor", 1),
            runtime_overrides=_runtime_overrides(con, (period or {}).get("id")),
        )
    except ValueError:
        rows = _even_distribution(trace, trip, departure_sec)
    return [_serialized_time_row({**row, "is_manual_override": 0}) for row in rows]


def _trip_time_rows(con, trip_id):
    return db.rows(con.execute(
        "SELECT * FROM trip_stop_times WHERE trip_id=? ORDER BY sequence,id",
        (trip_id,),
    ))


@router.patch("/trips/{trip_id}/stop-times/{route_stop_id}")
def trip_stop_time_adjust(
    trip_id: int, route_stop_id: int, payload: dict = Body(...),
    user=Depends(current_user),
):
    require_write(user, "trips")
    reason = str(payload.get("reason") or "").strip()
    if not reason:
        raise HTTPException(400, "Укажите причину ручной корректировки")
    strategy = str(payload.get("strategy") or "selected_only").strip()
    con = db.connect()
    try:
        trip = db.one(con.execute("SELECT * FROM route_trips WHERE id=?", (trip_id,)))
        if not trip:
            raise HTTPException(404, "Рейс не найден")
        rows = _trip_time_rows(con, trip_id)
        if not rows:
            raise HTTPException(400, "Сначала рассчитайте время по остановкам")
        departure_sec = service_time_to_seconds(payload.get("departure_time"))
        if departure_sec + 12 * 3600 < int(rows[0]["arrival_sec"]):
            departure_sec += 24 * 3600
        adjusted = adjust_stop_times(
            rows, route_stop_id=route_stop_id,
            departure_sec=departure_sec, strategy=strategy,
        )
        timestamp = datetime.datetime.now().isoformat(timespec="seconds")
        for row in adjusted:
            selected = int(row["route_stop_id"]) == route_stop_id
            con.execute(
                "UPDATE trip_stop_times SET arrival_sec=?,departure_sec=?,"
                "is_manual_override=?,override_strategy=?,override_reason=?,updated_at=? "
                "WHERE id=?",
                (row["arrival_sec"], row["departure_sec"],
                 1 if selected else row.get("is_manual_override", 0),
                 strategy if selected else row.get("override_strategy"),
                 reason if selected else row.get("override_reason"),
                 timestamp, row["id"]),
            )
        con.execute(
            "UPDATE route_trips SET dep_time=?,arr_time=? WHERE id=?",
            (format_service_time(adjusted[0]["departure_sec"]),
             format_service_time(adjusted[-1]["arrival_sec"]), trip_id),
        )
        db.audit(
            con, user["username"], "ручная корректировка времени остановки",
            "route_trips", trip_id,
            old={"route_stop_id": route_stop_id, "rows": rows},
            new={"route_stop_id": route_stop_id, "strategy": strategy,
                 "reason": reason, "rows": adjusted},
        )
        con.commit()
        return {"ok": True, "items": [_serialized_time_row(row) for row in adjusted]}
    except ValueError as exc:
        con.rollback()
        raise HTTPException(400, str(exc))
    finally:
        con.close()


@router.post("/routes/{route_id}/stop-times/reset-manual")
def route_stop_times_reset_manual(
    route_id: int, payload: dict = Body(...), user=Depends(current_user),
):
    require_write(user, "trips")
    day_type = str(payload.get("day_type") or "").strip()
    if not day_type:
        raise HTTPException(400, "Не указан тип дня")
    con = db.connect()
    selected_scopes = sum(
        payload.get(name) is not None for name in ("trip_id", "output_number")
    )
    if selected_scopes > 1:
        raise HTTPException(400, "Укажите только одну область сброса")
    try:
        query = "SELECT id FROM route_trips WHERE route_id=? AND day_type=?"
        args = [route_id, day_type]
        if payload.get("trip_id") is not None:
            query += " AND id=?"
            args.append(int(payload["trip_id"]))
        if payload.get("output_number") is not None:
            query += " AND output_number=?"
            args.append(int(payload["output_number"]))
        trip_ids = [row[0] for row in con.execute(query, args)]
        for selected_trip_id in trip_ids:
            recalculate_trip_stop_times_in_connection(
                con, selected_trip_id, preserve_manual=False
            )
        db.audit(
            con, user["username"], "сброс ручных корректировок времени",
            "routes", route_id,
            new={"day_type": day_type, "trips": len(trip_ids)},
        )
        con.commit()
        return {"ok": True, "updated": len(trip_ids)}
    except (TypeError, ValueError) as exc:
        con.rollback()
        raise HTTPException(400, str(exc))
    finally:
        con.close()


@router.post("/trips/{trip_id}/stop-times/recalculate")
def trip_stop_times_recalculate(trip_id: int, user=Depends(current_user)):
    require_write(user, "trips")
    con = db.connect()
    try:
        rows = recalculate_trip_stop_times_in_connection(
            con, trip_id, preserve_manual=False
        )
        db.audit(
            con,
            user["username"],
            "пересчёт поостановочного расписания",
            "route_trips",
            trip_id,
            new={"stop_times": len(rows)},
        )
        con.commit()
        return {"ok": True, "items": [_serialized_time_row(row) for row in rows]}
    except ValueError as exc:
        con.rollback()
        raise HTTPException(400, str(exc))
    finally:
        con.close()


@router.get("/routes/{route_id}/stop-times")
def stop_time_matrix(
    route_id: int,
    day_type: str,
    direction: str = "",
    output_number: int = 0,
    user=Depends(current_user),
):
    con = db.connect()
    try:
        if not con.execute("SELECT 1 FROM routes WHERE id=?", (route_id,)).fetchone():
            raise HTTPException(404, "Маршрут не найден")
        stops = {
            "forward": _trace_rows(con, route_id, "forward"),
            "backward": _trace_rows(con, route_id, "backward"),
        }
        query = (
            "SELECT * FROM route_trips WHERE route_id=? AND day_type=?"
        )
        args = [route_id, day_type]
        if direction:
            query += " AND direction=?"
            args.append(direction)
        if output_number:
            query += " AND output_number=?"
            args.append(output_number)
        query += " ORDER BY output_number,dep_time,trip_number,id"
        trips = []
        for trip in db.rows(con.execute(query, args)):
            trace = stops.get(DIRECTION_TO_TRACE.get(trip["direction"], ""), [])
            trips.append(
                {
                    "trip_id": trip["id"],
                    "output_number": trip["output_number"],
                    "shift_number": trip["shift_number"],
                    "trip_number": trip["trip_number"],
                    "direction": trip["direction"],
                    "period_id": trip.get("period_id"),
                    "dep_time": trip["dep_time"],
                    "arr_time": trip["arr_time"],
                    "times": _display_stop_times(con, trip, trace),
                }
            )
        return {"stops": stops, "trips": trips}
    finally:
        con.close()


def _active_periods(con, route_id, day_type):
    return db.rows(
        con.execute(
            "SELECT * FROM day_periods WHERE route_id=? AND day_type=? AND active=1 "
            "ORDER BY start_min,priority,id",
            (route_id, day_type),
        )
    )


def _all_runtime_overrides(con, period_ids):
    result = {period_id: {} for period_id in period_ids}
    if not period_ids:
        return result
    placeholders = ",".join("?" for _ in period_ids)
    for row in con.execute(
        "SELECT period_id,route_stop_id,run_time_sec FROM route_stop_runtimes "
        f"WHERE period_id IN ({placeholders})",
        period_ids,
    ):
        result.setdefault(row["period_id"], {})[row["route_stop_id"]] = row["run_time_sec"]
    return result


@router.post("/routes/{route_id}/schedule-generation/preview")
def schedule_generation_preview(
    route_id: int,
    payload: dict = Body(...),
    user=Depends(current_user),
):
    require_write(user, "trips")
    day_type = str(payload.get("day_type") or "будни")
    con = db.connect()
    try:
        route = db.one(con.execute("SELECT * FROM routes WHERE id=?", (route_id,)))
        if not route:
            raise HTTPException(404, "Маршрут не найден")
        periods = _active_periods(con, route_id, day_type)
        if not periods:
            raise HTTPException(400, "Для маршрута не заданы периоды движения")
        forward_trace = _trace_rows(con, route_id, "forward")
        backward_trace = _trace_rows(con, route_id, "backward")
        if not forward_trace or not backward_trace:
            raise HTTPException(400, "Для генерации нужны остановки обоих направлений")
        forward_min = int(route.get("trip_time_min") or 0)
        backward_min = int(route.get("trip_time_back_min") or 0)
        if forward_min <= 0 or backward_min <= 0:
            raise HTTPException(400, "Для маршрута не задано время движения")
        outputs = int(payload.get("outputs") or route.get("outputs_count") or 1)
        terminal_layover_min = int(payload.get("terminal_layover_min", 6))
        interval_preview = calculate_period_preview(
            periods,
            forward_min=forward_min,
            backward_min=backward_min,
            terminal_layover_min=terminal_layover_min,
        )
        trips = build_schedule_preview(
            departures=interval_preview["departures"],
            periods=periods,
            forward_trace=forward_trace,
            backward_trace=backward_trace,
            runtime_overrides=_all_runtime_overrides(
                con, [period["id"] for period in periods]
            ),
            outputs=outputs,
            terminal_layover_sec=terminal_layover_min * 60,
        )
        old_trip_count = con.execute(
            "SELECT COUNT(*) FROM route_trips WHERE route_id=? AND day_type=?",
            (route_id, day_type),
        ).fetchone()[0]
        old_stop_time_count = con.execute(
            "SELECT COUNT(*) FROM trip_stop_times tst JOIN route_trips rt "
            "ON rt.id=tst.trip_id WHERE rt.route_id=? AND rt.day_type=?",
            (route_id, day_type),
        ).fetchone()[0]
        diff = {
            "old_trip_count": old_trip_count,
            "new_trip_count": len(trips),
            "old_stop_time_count": old_stop_time_count,
            "new_stop_time_count": sum(len(trip["stop_times"]) for trip in trips),
        }
        plan = {
            "kind": "stop_schedule_generation",
            "route_id": route_id,
            "day_type": day_type,
            "outputs": outputs,
            "terminal_layover_min": terminal_layover_min,
            "trips": trips,
            "periods": interval_preview["periods"],
            "warnings": interval_preview["warnings"],
            "max_buses_required": interval_preview["max_buses_required"],
            "diff": diff,
        }
        token = secrets.token_hex(16)
        now = datetime.datetime.now()
        expires = now + datetime.timedelta(minutes=30)
        con.execute(
            "INSERT INTO schedule_generation_previews(token,route_id,day_type,username,"
            "payload_json,created_at,expires_at) VALUES(?,?,?,?,?,?,?)",
            (token, route_id, day_type, user["username"],
             json.dumps(plan, ensure_ascii=False), now.isoformat(timespec="seconds"),
             expires.isoformat(timespec="seconds")),
        )
        con.commit()
        return {"preview_token": token, "expires_at": expires.isoformat(timespec="seconds"), **plan}
    except ValueError as exc:
        con.rollback()
        raise HTTPException(400, str(exc))
    finally:
        con.close()


def _insert_generated_stop_times(con, trip_id, stop_times, timestamp):
    for row in stop_times:
        con.execute(
            "INSERT INTO trip_stop_times(trip_id,route_stop_id,sequence,arrival_sec,"
            "departure_sec,is_timing_point,is_manual_override,override_strategy,"
            "override_reason,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (
                trip_id,
                row["route_stop_id"],
                row["sequence"],
                row["arrival_sec"],
                row["departure_sec"],
                1 if row.get("is_timing_point") else 0,
                0,
                None,
                None,
                timestamp,
                timestamp,
            ),
        )


@router.post("/routes/{route_id}/schedule-generation/apply")
def schedule_generation_apply(
    route_id: int,
    payload: dict = Body(...),
    user=Depends(current_user),
):
    require_write(user, "trips")
    day_type = str(payload.get("day_type") or "будни")
    token = str(payload.get("preview_token") or "").strip()
    if not token:
        raise HTTPException(400, "Не указан токен предпросмотра")
    con = db.connect()
    try:
        preview = con.execute(
            "SELECT * FROM schedule_generation_previews WHERE token=? "
            "AND route_id=? AND day_type=? AND username=?",
            (token, route_id, day_type, user["username"]),
        ).fetchone()
        if not preview:
            raise HTTPException(404, "Предпросмотр не найден")
        preview = dict(preview)
        if preview["applied_at"]:
            raise HTTPException(409, "Предпросмотр уже применён")
        if datetime.datetime.fromisoformat(preview["expires_at"]) < datetime.datetime.now():
            raise HTTPException(410, "Срок предпросмотра истёк")
        plan = json.loads(preview["payload_json"])
        if (
            plan.get("kind") != "stop_schedule_generation"
            or int(plan.get("route_id")) != route_id
            or plan.get("day_type") != day_type
        ):
            raise HTTPException(400, "Некорректный план генерации")

        old_trip_count = con.execute(
            "SELECT COUNT(*) FROM route_trips WHERE route_id=? AND day_type=?",
            (route_id, day_type),
        ).fetchone()[0]
        con.execute(
            "DELETE FROM route_trips WHERE route_id=? AND day_type=?",
            (route_id, day_type),
        )
        timestamp = datetime.datetime.now().isoformat(timespec="seconds")
        inserted = 0
        for source in plan.get("trips") or []:
            cursor = con.execute(
                "INSERT INTO route_trips(route_id,day_type,output_number,shift_number,"
                "trip_number,direction,dep_time,arr_time,distance_km,break_after_min,"
                "break_type,period_id,source,generation_key) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    route_id,
                    day_type,
                    source["output_number"],
                    source.get("shift_number", 1),
                    source["trip_number"],
                    source["direction"],
                    source["dep_time"],
                    source["arr_time"],
                    source.get("distance_km", 0),
                    source.get("break_after_min", 0),
                    source.get("break_type", ""),
                    source.get("period_id"),
                    "period_generation",
                    token,
                ),
            )
            _insert_generated_stop_times(
                con, cursor.lastrowid, source.get("stop_times") or [], timestamp
            )
            inserted += 1
        updated = con.execute(
            "UPDATE schedule_generation_previews SET applied_at=? "
            "WHERE token=? AND applied_at IS NULL",
            (timestamp, token),
        )
        if updated.rowcount != 1:
            raise HTTPException(409, "Предпросмотр уже применён")
        db.audit(
            con,
            user["username"],
            "применение поостановочного расписания",
            "routes",
            route_id,
            old={"day_type": day_type, "trips": old_trip_count},
            new={"day_type": day_type, "trips": inserted, "generation_key": token},
        )
        con.commit()
        return {"ok": True, "trips": inserted, "generation_key": token}
    except (KeyError, TypeError, ValueError, sqlite3.IntegrityError) as exc:
        con.rollback()
        raise HTTPException(400, f"Не удалось применить расписание: {exc}")
    finally:
        con.close()

EXPORT_DARK = "17365D"
EXPORT_BLUE = "DDEBF7"
EXPORT_ALT = "F5F8FC"
EXPORT_MANUAL = "FFF2CC"


def _prepare_timetable_sheet(ws, *, title, metadata, headers, widths,
                             landscape=False, freeze="A4"):
    """Apply the common print-ready timetable style."""
    last_column = max(1, len(headers))
    ws.title = "Расписание"
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=EXPORT_DARK)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 27
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    ws.cell(2, 1, metadata)
    ws.cell(2, 1).font = Font(italic=True, color="44546A")
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22
    thin = Side(style="thin", color="B4C6E7")
    for column, header in enumerate(headers, 1):
        cell = ws.cell(3, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=EXPORT_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin, right=thin)
        ws.column_dimensions[get_column_letter(column)].width = widths[column - 1]
    ws.row_dimensions[3].height = 34
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A3:{get_column_letter(last_column)}{max(3, ws.max_row)}"
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def _finish_timetable_sheet(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row_index, row in enumerate(ws.iter_rows(min_row=4), 4):
        fill = PatternFill("solid", fgColor=EXPORT_ALT) if row_index % 2 == 0 else None
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if cell.column == 1 else "center",
                vertical="center", wrap_text=True,
            )
            if fill and cell.fill.fill_type is None:
                cell.fill = fill
        ws.row_dimensions[row_index].height = 24
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"


def _trip_direction_key(direction):
    return DIRECTION_TO_TRACE.get(direction, direction)


@router.get("/routes/{route_id}/stop-times/export.xlsx")
def route_stop_times_export(
    route_id: int, day_type: str, direction: str = "", output_number: int = 0,
    user=Depends(current_user),
):
    con = db.connect()
    try:
        route = db.one(con.execute("SELECT * FROM routes WHERE id=?", (route_id,)))
        if not route:
            raise HTTPException(404, "Маршрут не найден")
        query = "SELECT * FROM route_trips WHERE route_id=? AND day_type=?"
        args = [route_id, day_type]
        if direction:
            query += " AND direction=?"
            args.append(direction)
        if output_number:
            query += " AND output_number=?"
            args.append(output_number)
        query += " ORDER BY output_number,dep_time,trip_number,id"
        trips = db.rows(con.execute(query, args))
        headers = ["Остановка"] + [
            f"Вых. {trip['output_number']} · рейс {trip['trip_number']}\n"
            f"{trip['dep_time']}–{trip['arr_time']}" for trip in trips
        ]
        wb = Workbook()
        ws = wb.active
        _prepare_timetable_sheet(
            ws,
            title=f"Поостановочное расписание маршрута № {route['number']}",
            metadata=f"{route.get('name') or ''} · тип дня: {day_type}",
            headers=headers,
            widths=[34] + [16] * len(trips),
            landscape=True,
            freeze="B4",
        )
        time_maps = {
            trip["id"]: {
                row["route_stop_id"]: row
                for row in db.rows(con.execute(
                    "SELECT * FROM trip_stop_times WHERE trip_id=?", (trip["id"],)
                ))
            } for trip in trips
        }
        for trace_direction, label in (("forward", "Прямое"), ("backward", "Обратное")):
            stops = _trace_rows(con, route_id, trace_direction)
            for stop in stops:
                values = [f"{label} · {stop['sequence']}. {stop['stop_name']}"]
                manual_columns = []
                for column, trip in enumerate(trips, 2):
                    row = time_maps[trip["id"]].get(stop["id"])
                    if not row or _trip_direction_key(trip["direction"]) != trace_direction:
                        values.append("")
                        continue
                    arrival = format_service_time(row["arrival_sec"])
                    departure = format_service_time(row["departure_sec"])
                    values.append(arrival if arrival == departure else f"{arrival} / {departure}")
                    if row["is_manual_override"]:
                        manual_columns.append(column)
                ws.append(values)
                for column in manual_columns:
                    ws.cell(ws.max_row, column).fill = PatternFill("solid", fgColor=EXPORT_MANUAL)
        _finish_timetable_sheet(ws)
        return _xlsx_download_response(
            wb, f"route_{route['number']}_{day_type}_stop_times.xlsx"
        )
    finally:
        con.close()


@router.get("/trips/{trip_id}/stop-times/export.xlsx")
def trip_stop_times_export(trip_id: int, user=Depends(current_user)):
    con = db.connect()
    try:
        trip = db.one(con.execute(
            "SELECT rt.*,r.number AS route_number,r.name AS route_name "
            "FROM route_trips rt JOIN routes r ON r.id=rt.route_id WHERE rt.id=?",
            (trip_id,),
        ))
        if not trip:
            raise HTTPException(404, "Рейс не найден")
        rows = db.rows(con.execute(
            "SELECT tst.*,s.name AS stop_name FROM trip_stop_times tst "
            "JOIN route_stops rs ON rs.id=tst.route_stop_id "
            "JOIN stops s ON s.id=rs.stop_id WHERE tst.trip_id=? "
            "ORDER BY tst.sequence,tst.id",
            (trip_id,),
        ))
        wb = Workbook()
        ws = wb.active
        _prepare_timetable_sheet(
            ws,
            title=f"Лист рейса № {trip['trip_number']} · маршрут № {trip['route_number']}",
            metadata=(f"{trip.get('route_name') or ''} · {trip['day_type']} · "
                      f"выход {trip['output_number']} · {trip['dep_time']}–{trip['arr_time']}"),
            headers=["№", "Остановка", "Прибытие", "Отправление", "Ручная корректировка"],
            widths=[7, 38, 15, 15, 32],
        )
        for row in rows:
            ws.append([
                row["sequence"], row["stop_name"],
                format_service_time(row["arrival_sec"]),
                format_service_time(row["departure_sec"]),
                row.get("override_reason") or "",
            ])
            if row["is_manual_override"]:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=EXPORT_MANUAL)
        _finish_timetable_sheet(ws)
        return _xlsx_download_response(wb, f"trip_{trip_id}_stop_times.xlsx")
    finally:
        con.close()


@router.get("/stops/{stop_id}/timetable.xlsx")
def stop_pavilion_timetable_export(
    stop_id: int, day_type: str, user=Depends(current_user),
):
    con = db.connect()
    try:
        stop = db.one(con.execute("SELECT * FROM stops WHERE id=?", (stop_id,)))
        if not stop:
            raise HTTPException(404, "Остановка не найдена")
        rows = db.rows(con.execute(
            "SELECT r.number AS route_number,r.name AS route_name,rt.direction,"
            "rt.output_number,rt.trip_number,tst.arrival_sec,tst.departure_sec "
            "FROM trip_stop_times tst "
            "JOIN route_stops rs ON rs.id=tst.route_stop_id "
            "JOIN route_trips rt ON rt.id=tst.trip_id "
            "JOIN routes r ON r.id=rt.route_id "
            "WHERE rs.stop_id=? AND rt.day_type=? "
            "ORDER BY tst.departure_sec,r.number,rt.output_number,rt.trip_number",
            (stop_id, day_type),
        ))
        wb = Workbook()
        ws = wb.active
        _prepare_timetable_sheet(
            ws,
            title=f"Расписание остановки «{stop['name']}»",
            metadata=f"Тип дня: {day_type} · отправлений: {len(rows)}",
            headers=["Время", "Маршрут", "Направление", "Выход", "Рейс", "Прибытие"],
            widths=[14, 22, 22, 10, 10, 14],
        )
        for row in rows:
            ws.append([
                format_service_time(row["departure_sec"]),
                f"№ {row['route_number']} · {row.get('route_name') or ''}",
                row["direction"], row["output_number"], row["trip_number"],
                format_service_time(row["arrival_sec"]),
            ])
        _finish_timetable_sheet(ws)
        return _xlsx_download_response(wb, f"stop_{stop_id}_{day_type}.xlsx")
    finally:
        con.close()
