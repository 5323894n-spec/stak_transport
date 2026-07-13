# -*- coding: utf-8 -*-
"""Excel-отчёты модуля ремонта и ТО."""
import datetime
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from . import db
from .auth import current_user

router = APIRouter(prefix="/api/repairs/reports", tags=["repair-reports"])
HEADER = PatternFill("solid", fgColor="17365D")
THIN = Side(style="thin", color="B8C6D5")

def add_table(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows: ws.append(list(row))
    for cell in ws[1]:
        cell.fill = HEADER; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows():
        for cell in row: cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN); cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, 1):
        width = max((len(str(cell.value or "")) for cell in column), default=8) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(max(width, 10), 42)
    return ws

@router.get("/export.xlsx")
def export_repairs(date_from: str = "", date_to: str = "", vehicle_id: int = 0, status: str = "", user=Depends(current_user)):
    try:
        start_date = datetime.date.fromisoformat(date_from) if date_from else None
        end_date = datetime.date.fromisoformat(date_to) if date_to else None
    except ValueError:
        raise HTTPException(400, "Неверный формат периода отчёта")
    if start_date and end_date and start_date > end_date:
        raise HTTPException(400, "Начало периода отчёта позже окончания")

    def filters(alias, date_column, *, include_status=False):
        where, args = [], []
        if start_date:
            where.append(f"{alias}.{date_column}>=?"); args.append(start_date.isoformat() + "T00:00:00")
        if end_date:
            where.append(f"{alias}.{date_column}<=?"); args.append(end_date.isoformat() + "T23:59:59")
        if vehicle_id:
            where.append(f"{alias}.bus_id=?"); args.append(vehicle_id)
        if include_status and status:
            where.append(f"{alias}.status=?"); args.append(status)
        return (" WHERE " + " AND ".join(where)) if where else "", args

    con = db.connect()
    try:
        where, args = filters("ro", "created_at", include_status=True)
        orders = db.rows(con.execute(
            "SELECT ro.number,ro.status,b.garage_number,b.plate,rt.name repair_type,ro.created_at,ro.planned_end,ro.closed_at,ro.actual_hours,ro.parts_cost,ro.total_cost,ro.downtime_hours,ro.result "
            "FROM repair_orders ro JOIN buses b ON b.id=ro.bus_id LEFT JOIN repair_types rt ON rt.id=ro.repair_type_id" + where + " ORDER BY ro.created_at DESC", args))
        where, args = filters("rr", "created_at")
        requests = db.rows(con.execute(
            "SELECT rr.number,rr.status,b.garage_number,b.plate,rr.source,rr.priority,rr.odometer,rr.description,rr.created_at,rr.closed_at FROM repair_requests rr JOIN buses b ON b.id=rr.bus_id" + where + " ORDER BY rr.created_at DESC", args))
        where, args = filters("vrh", "closed_at")
        history = db.rows(con.execute(
            "SELECT vrh.order_number,vrh.request_number,vrh.bus_id,vrh.opened_at,vrh.closed_at,vrh.odometer,vrh.result,vrh.total_cost,vrh.downtime_hours FROM vehicle_repair_history vrh" + where + " ORDER BY vrh.closed_at DESC", args))
        plan_where, plan_args = (" WHERE mp.active=1", [])
        if vehicle_id:
            plan_where += " AND mp.bus_id=?"; plan_args.append(vehicle_id)
        plans = db.rows(con.execute(
            "SELECT b.garage_number,b.plate,mp.name,rt.name repair_type,mp.next_date,mp.next_odometer,b.odometer current_odometer,mp.warning_days,mp.warning_km FROM maintenance_plans mp JOIN buses b ON b.id=mp.bus_id JOIN repair_types rt ON rt.id=mp.repair_type_id" + plan_where + " ORDER BY mp.next_date", plan_args))
        stock = db.rows(con.execute("SELECT p.code,p.name,p.unit,w.name warehouse,p.stock_qty,p.reserved_qty,p.min_qty,p.unit_price FROM parts p LEFT JOIN warehouses w ON w.id=p.warehouse_id WHERE p.active=1 ORDER BY p.name"))
        active = sum(1 for row in orders if row["status"] not in ("завершен", "отменен"))
        total_cost = sum(float(row["total_cost"] or 0) for row in orders)
        downtime = sum(float(row["downtime_hours"] or 0) for row in orders)
    finally: con.close()
    wb = Workbook(); ws = wb.active; ws.title = "Сводка"
    ws["A1"] = "ОТЧЁТ ПО РЕМОНТУ И ТО"; ws["A1"].font = Font(size=16, bold=True, color="17365D")
    ws.append(["Сформирован", datetime.datetime.now()]); ws.append(["Активных заказ-нарядов", active]); ws.append(["Всего заказ-нарядов", len(orders)]); ws.append(["Общая стоимость", total_cost]); ws.append(["Простой, часов", downtime])
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 24; ws["B5"].number_format = '#,##0.00'; ws["B6"].number_format = '#,##0.00'
    order_ws = add_table(wb, "Заказ-наряды", ["Номер","Статус","Гаражный №","Госномер","Вид ремонта","Создан","План окончания","Закрыт","Часы","Запчасти","Всего","Простой","Результат"], ([r[k] for k in r] for r in orders))
    for row in order_ws.iter_rows(min_row=2):
        for col in (9,10,11,12): row[col-1].number_format = '#,##0.00'
    add_table(wb, "Заявки", ["Номер","Статус","Гаражный №","Госномер","Источник","Приоритет","Пробег","Описание","Создана","Закрыта"], ([r[k] for k in r] for r in requests))
    add_table(wb, "История", ["Заказ-наряд","Заявка","ID автобуса","Открыт","Закрыт","Пробег","Результат","Стоимость","Простой"], ([r[k] for k in r] for r in history))
    add_table(wb, "Планы ТО", ["Гаражный №","Госномер","План","Вид ТО","Следующая дата","Следующий пробег","Текущий пробег","Дней предупреждения","Км предупреждения"], ([r[k] for k in r] for r in plans))
    add_table(wb, "Склад", ["Код","Наименование","Ед.","Склад","Остаток","Резерв","Минимум","Цена"], ([r[k] for k in r] for r in stock))
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    headers = {"Content-Disposition": "attachment; filename=repair_report.xlsx"}
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)