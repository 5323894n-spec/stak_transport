# -*- coding: utf-8 -*-
"""Экспорт в Excel."""
import datetime
import io
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import Response


def xlsx_response(title, headers, data_rows, filename="export.xlsx", col_widths=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Лист1"
    thin = Border(*[Side(style="thin")] * 4)
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(headers)
    for c in ws[2]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDE7F3")
        c.border = thin
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in data_rows:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=3):
        for c in row:
            c.border = thin
    for i, h in enumerate(headers, 1):
        w = (col_widths[i-1] if col_widths and i <= len(col_widths) else max(10, min(35, len(str(h)) + 4)))
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    encoded_name = quote(filename, safe="")
    return Response(buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"})


def _xlsx_download_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    encoded_name = quote(filename, safe="")
    return Response(buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"})


def _date_label(date_text):
    try:
        return datetime.date.fromisoformat(date_text).strftime("%d.%m.%Y")
    except Exception:
        return date_text or ""


def _num(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _merge_value(ws, row, col1, col2, value, *, fill=None, font=None, alignment=None, border=None):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        for col in range(col1, col2 + 1):
            ws.cell(row=row, column=col).border = border
    return cell


def order_xlsx_response(order, settings, lines, filename="naryad.xlsx"):
    headers = [
        "Маршрут", "Выход", "Смена", "Водитель", "Таб.№", "Автобус (гар.)", "Госномер", "Явка", "Выезд",
        "Начало", "Окончание", "Заезд", "Часы", "Рейсов", "Пробег, км", "План. топливо, л", "Статус", "Отметки",
    ]
    header_row = 9
    data_start = header_row + 1
    last_col = len(headers)
    last_letter = get_column_letter(last_col)
    date_text = order.get("date", "")
    sheet_title = f"Наряд {date_text}"[:31] or "Наряд"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False

    dark = "1F4E78"
    mid = "D9EAF7"
    light = "EEF5FB"
    card_fill = PatternFill("solid", fgColor="EAF2F8")
    header_fill = PatternFill("solid", fgColor=dark)
    sub_fill = PatternFill("solid", fgColor=mid)
    neutral_fill = PatternFill("solid", fgColor="F6F8FA")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    bad_fill = PatternFill("solid", fgColor="F4CCCC")
    white_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color=dark)
    org_font = Font(name="Calibri", size=12, bold=True, color="22313F")
    bold_font = Font(name="Calibri", size=10, bold=True, color="22313F")
    body_font = Font(name="Calibri", size=10, color="22313F")
    small_font = Font(name="Calibri", size=9, color="566573")
    thin_side = Side(style="thin", color="B7C9D6")
    medium_side = Side(style="medium", color="6C8EAD")
    thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=medium_side, bottom=medium_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    org_name = settings.get("org_name") or "Автотранспортное предприятие"
    approved = order.get("approved_by") or ""
    status = order.get("status") or ""
    date_label = _date_label(date_text)
    now_label = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")

    _merge_value(ws, 1, 1, last_col, org_name, font=org_font, alignment=center)
    _merge_value(ws, 2, 1, last_col, "НАРЯД НА ВЫПУСК АВТОБУСОВ", font=title_font, alignment=center)
    meta = f"на {date_label}   |   статус: {status or '—'}"
    if approved:
        meta += f"   |   утвердил: {approved}"
    _merge_value(ws, 3, 1, last_col, meta, font=small_font, alignment=center)

    data_end = data_start + len(lines) - 1 if lines else data_start
    card_pairs = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12), (13, 14), (15, 16), (17, 18)]
    summary = [
        ("Всего строк", f"=COUNTA(Q{data_start}:Q{data_end})"),
        ("В работе", f"=COUNTIFS(Q{data_start}:Q{data_end},\"<>отменен\",Q{data_start}:Q{data_end},\"<>\")"),
        ("Отменено", f"=COUNTIF(Q{data_start}:Q{data_end},\"отменен\")"),
        ("Водителей", f"=COUNTA(D{data_start}:D{data_end})"),
        ("Автобусов", f"=COUNTA(F{data_start}:F{data_end})"),
        ("Часов", f"=SUM(M{data_start}:M{data_end})"),
        ("Рейсов", f"=SUM(N{data_start}:N{data_end})"),
        ("Пробег, км", f"=SUM(O{data_start}:O{data_end})"),
        ("Топливо, л", f"=SUM(P{data_start}:P{data_end})"),
    ]
    for (col1, col2), (label, formula) in zip(card_pairs, summary):
        _merge_value(ws, 5, col1, col2, label, fill=card_fill, font=small_font, alignment=center, border=thin)
        cell = _merge_value(ws, 6, col1, col2, formula, fill=PatternFill("solid", fgColor="FFFFFF"), font=bold_font, alignment=center, border=thin)
        if label in {"Часов", "Пробег, км", "Топливо, л"}:
            cell.number_format = "0.0"
        else:
            cell.number_format = "0"

    _merge_value(ws, 8, 1, last_col, "Детализация выпусков", fill=sub_fill, font=bold_font, alignment=left, border=thin)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = header_border

    for offset, line in enumerate(lines):
        row = data_start + offset
        route = f"№ {line.get('rn') or ''} {line.get('route_name') or ''}".strip()
        values = [
            route,
            line.get("output_number"),
            line.get("shift_number"),
            line.get("fio") or "",
            line.get("tab_number") or "",
            line.get("garage_number") or "",
            line.get("plate") or "",
            line.get("report_time") or "",
            line.get("depart_depot") or "",
            line.get("start_line") or "",
            line.get("end_line") or "",
            line.get("return_depot") or "",
            _num(line.get("shift_hours")),
            int(_num(line.get("trips_count"))),
            _num(line.get("distance_km")),
            _num(line.get("planned_fuel")),
            line.get("status") or "",
            line.get("dispatcher_note") or "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = body_font
            cell.border = thin
            cell.alignment = right if col in (2, 3, 13, 14, 15, 16) else left
            if col in (13, 15, 16):
                cell.number_format = "0.0"
            elif col == 14:
                cell.number_format = "0"
        if not line.get("fio") or not line.get("garage_number"):
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).fill = warn_fill
        status_value = (line.get("status") or "").lower()
        status_cell = ws.cell(row=row, column=17)
        if status_value in ("выдан", "выполнен"):
            status_cell.fill = ok_fill
        elif status_value == "отменен":
            status_cell.fill = bad_fill
        else:
            status_cell.fill = neutral_fill

    if not lines:
        for col in range(1, last_col + 1):
            cell = ws.cell(row=data_start, column=col, value="")
            cell.border = thin
            cell.font = body_font
            cell.alignment = left

    table_ref = f"A{header_row}:{last_letter}{data_end}"
    ws.auto_filter.ref = table_ref
    ws.freeze_panes = f"A{data_start}"
    ws.print_title_rows = f"$1:${header_row}"

    widths = [24, 8, 8, 26, 10, 14, 14, 10, 10, 10, 10, 10, 9, 9, 12, 14, 12, 24]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(1, data_end + 1):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[header_row].height = 30

    cf_range = f"A{data_start}:{last_letter}{data_end}"
    ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'$Q{data_start}="отменен"'], fill=bad_fill))
    ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'OR($D{data_start}="",$F{data_start}="")'], fill=warn_fill))

    sign_row = data_end + 3
    _merge_value(ws, sign_row, 1, last_col, "Подписи ответственных лиц", fill=sub_fill, font=bold_font, alignment=left, border=thin)
    sign_labels = [
        ("Диспетчер", 1, 4),
        ("Начальник эксплуатации", 5, 9),
        ("Механик", 10, 13),
        ("Медработник", 14, 18),
    ]
    for label, col1, col2 in sign_labels:
        _merge_value(ws, sign_row + 1, col1, col2, f"{label}: __________________ /__________________/", font=body_font, alignment=left, border=thin)
    _merge_value(ws, sign_row + 3, 1, last_col, f"Файл сформирован: {now_label}", font=small_font, alignment=left)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    return _xlsx_download_response(wb, filename)


SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="17365D")
SUMMARY_HEADER_FONT = Font(color="FFFFFF", bold=True)
SUMMARY_BORDER = Border(
    left=Side(style="thin", color="B7C9D9"),
    right=Side(style="thin", color="B7C9D9"),
    top=Side(style="thin", color="B7C9D9"),
    bottom=Side(style="thin", color="B7C9D9"),
)


def _summary_write_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = SUMMARY_HEADER_FILL
        cell.font = SUMMARY_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = SUMMARY_BORDER
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = SUMMARY_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        values = [ws.cell(row=row, column=col).value for row in range(1, ws.max_row + 1)]
        width = min(42, max(10, max(len(str(v or "")) for v in values) + 2))
        ws.column_dimensions[letter].width = width


def summary_schedule_xlsx_response(settings, summary, views, lines, errors, filename="Сводное_расписание_за_период.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Титульный лист"
    org = settings.get("org_name", "")
    period = summary["period_start"] if summary["period_start"] == summary["period_end"] else f"{summary['period_start']} — {summary['period_end']}"
    ws["A1"] = org
    ws["A2"] = "СВОДНОЕ РАСПИСАНИЕ"
    ws["A3"] = f"Период: {period}"
    ws["A4"] = f"Дата формирования: {summary.get('created_at') or ''}"
    ws["A6"] = "Маршрутов"; ws["B6"] = summary["routes_count"]
    ws["A7"] = "Рейсов"; ws["B7"] = summary["trips_count"]
    ws["A8"] = "Выходов"; ws["B8"] = summary["runs_count"]
    ws["A9"] = "Автобусов"; ws["B9"] = summary["vehicles_count"]
    ws["A10"] = "Водителей"; ws["B10"] = summary["drivers_count"]
    ws["A11"] = "Ошибок"; ws["B11"] = summary["errors_count"]
    ws["A12"] = "Предупреждений"; ws["B12"] = summary["warnings_count"]
    ws["A14"] = "Ответственный пользователь"; ws["B14"] = summary.get("created_by") or ""
    ws["A16"] = "Подпись ответственного пользователя"; ws["B16"] = "____________________ /____________________/"
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws["A2"].font = Font(size=16, bold=True, color="17365D")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34

    sheet_defs = [
        ("Сводка", ["Показатель", "Значение"], [
            ["Количество маршрутов", summary["routes_count"]],
            ["Количество рейсов", summary["trips_count"]],
            ["Количество выходов", summary["runs_count"]],
            ["Количество автобусов", summary["vehicles_count"]],
            ["Количество водителей", summary["drivers_count"]],
            ["Общий пробег", round(sum(float(x.get("distance_km") or 0) for x in lines), 1)],
            ["Количество ошибок", summary["errors_count"]],
            ["Количество предупреждений", summary["warnings_count"]],
            ["Дата формирования", summary.get("created_at") or ""],
            ["Период", period],
        ]),
        ("По маршрутам", ["Дата", "Маршрут", "Направление", "Рейс", "Выход", "Отправление", "Прибытие", "Время рейса", "Автобус", "Водитель", "Примечание"], [
            [x["service_date"], f"№ {x['route_number']} {x['route_name']}", x["direction"], x["trip_number"], x["run_number"], x["departure_time"], x["arrival_time"], x["trip_duration"], x["garage_number"] or x["vehicle_number"], x["driver_name"], x["comment"]] for x in views["by_routes"]
        ]),
        ("По выходам", ["Дата", "Маршрут", "Выход", "Смена", "Автобус", "Водитель", "Начало", "Окончание", "Рейсов", "Пробег", "Время на линии"], [
            [x["service_date"], f"№ {x['route_number']} {x['route_name']}", x["run_number"], x["shift_number"], x["garage_number"] or x["vehicle_number"], x["driver_name"], x["start_time"], x["end_time"], x["trips_count"], x["distance_km"], round((x["line_minutes"] or 0) / 60, 2)] for x in views["by_outputs"]
        ]),
        ("По водителям", ["Дата", "Водитель", "Табельный", "Маршрут", "Выход", "Автобус", "Начало", "Окончание", "Рабочее время"], [
            [x["service_date"], x["driver_name"], x["driver_tab_number"], f"№ {x['route_number']}", x["run_number"], x["garage_number"] or x["vehicle_number"], x["start_time"], x["end_time"], round((x["line_minutes"] or 0) / 60, 2)] for x in views["by_drivers"]
        ]),
        ("По автобусам", ["Дата", "Автобус", "Госномер", "Маршрут", "Выход", "Водитель", "Выезд", "Заезд", "Рейсов", "Пробег"], [
            [x["service_date"], x["garage_number"], x["vehicle_number"], f"№ {x['route_number']}", x["run_number"], x["driver_name"], x["start_time"], x["end_time"], x["trips_count"], x["distance_km"]] for x in views["by_buses"]
        ]),
        ("По времени", ["Дата", "Время", "Маршрут", "Направление", "Рейс", "Выход", "Автобус", "Водитель", "Событие"], [
            [x.get("service_date"), x.get("time"), f"№ {x.get('route_number', '')}", x.get("direction", ""), x.get("trip_number", ""), x.get("run_number", ""), x.get("garage_number", "") or x.get("vehicle_number", ""), x.get("driver_name", ""), x.get("event", "")] for x in views["by_time"]
        ]),
        ("Ошибки", ["Уровень", "Маршрут", "Выход", "Рейс", "Объект", "Описание ошибки", "Рекомендация"], [
            [e["level"], e["route_number"], e["run_number"], e["trip_number"], e["object_label"], e["message"], e["recommendation"]] for e in errors
        ]),
        ("Исходные данные", list(lines[0].keys()) if lines else ["Нет данных"], [[x.get(k) for k in lines[0].keys()] for x in lines] if lines else []),
    ]
    for title, headers, rows in sheet_defs:
        sh = wb.create_sheet(title)
        _summary_write_table(sh, headers, rows)
    return _xlsx_download_response(wb, filename)


def _month_label(month):
    names = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    y, m = map(int, month.split("-"))
    return f"{names[m]} {y}", y, m


def roster_xlsx_response(month, settings, drivers, roster_rows, assignments, warnings, routes, norms, filename="roster.xlsx"):
    import calendar
    wb = Workbook()
    ws = wb.active
    ws.title = "График месяца"
    label, year, month_num = _month_label(month)
    days = calendar.monthrange(year, month_num)[1]
    ws["A1"] = settings.get("org_name", "")
    ws["A2"] = "ГРАФИК РАБОТЫ ВОДИТЕЛЕЙ"
    ws["A3"] = label
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days + 4)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=days + 4)
    ws["A2"].font = Font(size=15, bold=True, color="17365D")
    header_row = 7
    headers = ["Таб.№", "Водитель"] + [str(d) for d in range(1, days + 1)] + ["Итого"]
    for col, value in enumerate(headers, 1):
        c = ws.cell(header_row, col, value)
        c.fill = SUMMARY_HEADER_FILL
        c.font = SUMMARY_HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = SUMMARY_BORDER
    by_driver_date = {}
    for a in assignments:
        by_driver_date.setdefault((a.get("driver_id"), a.get("date")), []).append(a)
    roster_by_driver_date = {(r.get("driver_id"), r.get("date")): r for r in roster_rows}
    row = header_row + 1
    for driver in drivers:
        ws.cell(row, 1, driver.get("tab_number") or "")
        ws.cell(row, 2, driver.get("fio") or "")
        total = 0.0
        for d in range(1, days + 1):
            iso = f"{year:04d}-{month_num:02d}-{d:02d}"
            items = by_driver_date.get((driver.get("id"), iso), [])
            text_parts = []
            if items:
                for a in items:
                    total += float(a.get("hours") or 0)
                    text_parts.append(
                        f"№{a.get('route_number') or ''}\n{a.get('output_number') or ''}/{a.get('shift_number') or ''}\n"
                        f"{a.get('start_time') or ''}-{a.get('end_time') or ''}\nПЗВ{int(norms.get('prep_final_minutes') or 0)}м\n{round(float(a.get('hours') or 0), 2)} ч"
                    )
            else:
                r = roster_by_driver_date.get((driver.get("id"), iso))
                if r and r.get("status") and r.get("status") != "работа":
                    text_parts.append(r.get("status"))
                elif r and r.get("status") == "работа":
                    text_parts.append("работа")
                    total += float(r.get("hours") or 0)
            c = ws.cell(row, d + 2, "\n".join(text_parts))
            c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True, shrink_to_fit=True)
            c.border = SUMMARY_BORDER
            if items:
                c.fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(row, days + 3, round(total, 2))
        for col in range(1, days + 4):
            ws.cell(row, col).border = SUMMARY_BORDER
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True, shrink_to_fit=True)
        ws.row_dimensions[row].height = 64
        row += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    for col in range(3, days + 3):
        ws.column_dimensions[get_column_letter(col)].width = 10.5
    ws.column_dimensions[get_column_letter(days + 3)].width = 10
    ws.freeze_panes = "C8"

    sh = wb.create_sheet("Назначения")
    sh["A1"] = settings.get("org_name", "")
    sh["A2"] = "Назначения водителей по графику"
    assign_headers = ["Дата", "Таб.№", "Водитель", "Статус дня", "Маршрут", "Выход", "Смена", "Рейсы", "Начало", "Окончание", "ПЗВ, мин", "Обед/перерыв, мин", "Линейное время, ч", "Рабочее время с ПЗВ, ч", "Ночные часы", "Комментарий"]
    for col, value in enumerate(assign_headers, 1):
        c = sh.cell(4, col, value)
        c.fill = SUMMARY_HEADER_FILL
        c.font = SUMMARY_HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = SUMMARY_BORDER
    prep = int(norms.get("prep_final_minutes") or 0)
    for idx, a in enumerate(assignments, 5):
        values = [
            a.get("date"), a.get("tab_number"), a.get("fio"), "работа",
            f"№{a.get('route_number') or ''}", a.get("output_number"), a.get("shift_number"),
            f"{a.get('trip_from') or ''}-{a.get('trip_to') or ''}", a.get("start_time"), a.get("end_time"),
            prep, int(a.get("break_min") or 0), f"=ROUND((J{idx}-I{idx})*24,2)",
            f"=M{idx}+K{idx}/60", a.get("night_hours") or 0, a.get("comment") or "",
        ]
        for col, value in enumerate(values, 1):
            c = sh.cell(idx, col, value)
            c.border = SUMMARY_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 11:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
    sh.auto_filter.ref = f"A4:P{max(5, 4 + len(assignments))}"
    sh.freeze_panes = "A5"
    sh.protection.sheet = False
    for col in range(1, 17):
        sh.column_dimensions[get_column_letter(col)].width = 16

    warn_ws = wb.create_sheet("Проверки 424")
    warn_ws["A1"] = "Проверки 424"
    warn_headers = ["Уровень", "Дата", "Водитель", "Тип", "Норма", "Факт", "Рекомендация"]
    for col, value in enumerate(warn_headers, 1):
        warn_ws.cell(3, col, value)
    for idx, w in enumerate(warnings, 4):
        vals = [w.get("severity"), w.get("date"), w.get("driver"), w.get("type"), w.get("norm_value"), w.get("fact_value"), w.get("recommendation")]
        for col, value in enumerate(vals, 1):
            warn_ws.cell(idx, col, value)
    if not warnings:
        warn_ws.cell(4, 1, "Нарушений не найдено")
    for col in range(1, 8):
        warn_ws.column_dimensions[get_column_letter(col)].width = 24

    refs = wb.create_sheet("Справочники")
    refs["A1"] = "ПЗВ по нормам, мин"
    refs["B1"] = prep
    refs["A2"] = "Месяц"
    refs["B2"] = label
    refs["A4"] = "Маршруты"
    for idx, r in enumerate(routes, 5):
        refs.cell(idx, 1, r.get("number"))
        refs.cell(idx, 2, r.get("name"))
    return _xlsx_download_response(wb, filename)
