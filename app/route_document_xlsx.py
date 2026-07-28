# -*- coding: utf-8 -*-
"""OpenPyXL builders and presentation helpers for route documents."""

import datetime
from io import BytesIO
import re
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "17324D"
BLUE = "DCE8F3"
PALE_BLUE = "EEF4FB"
PALE_GREEN = "E9F6EE"
PALE_AMBER = "FFF3DD"
_WHITE = "FFFFFF"
_BORDER = Border(**{
    side: Side(style="thin", color=NAVY)
    for side in ("left", "right", "top", "bottom")
})


def _safe_text(value):
    """Return user-controlled text as a literal Excel string."""
    if not isinstance(value, str):
        return value
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def apply_sheet_setup(ws, *, landscape=True):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "Страница &P из &N"
    return ws


def write_excel_time(cell, seconds):
    cell.value = None if seconds is None else seconds / 86400
    cell.number_format = "[h]:mm"
    return cell


def _write_band(ws, row, text, *, start_col, end_col, color, font_color, size):
    if end_col < start_col:
        raise ValueError("Последний столбец полосы не может быть меньше первого")
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, text)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color=font_color, size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _BORDER
    for column in range(start_col + 1, end_col + 1):
        ws.cell(row, column).fill = PatternFill("solid", fgColor=color)
        ws.cell(row, column).border = _BORDER
    return cell


def write_title_band(ws, row, text, *, start_col=1, end_col=8):
    return _write_band(ws, row, text, start_col=start_col, end_col=end_col,
                       color=NAVY, font_color=_WHITE, size=14)


def write_section_header(ws, row, text, *, start_col=1, end_col=8):
    return _write_band(ws, row, text, start_col=start_col, end_col=end_col,
                       color=BLUE, font_color=NAVY, size=11)


def write_table_header(ws, row, headers, *, start_col=1):
    result = []
    for offset, header in enumerate(headers):
        cell = ws.cell(row, start_col + offset, header)
        cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        cell.font = Font(bold=True, color=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        result.append(cell)
    return tuple(result)


def apply_warning_cell(cell):
    cell.fill = PatternFill("solid", fgColor=PALE_AMBER)
    cell.font = Font(color=NAVY)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _BORDER
    return cell


def set_print_area(ws, *, min_row=1, min_col=1, max_row=None, max_col=None):
    max_row = ws.max_row if max_row is None else max_row
    max_col = ws.max_column if max_col is None else max_col
    ws.print_area = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return ws.print_area


def _route_token(route_number):
    value = str(route_number or "").strip()
    value = re.sub(r"^[MМmм](?=\d)", "", value, count=1)
    numeric = re.fullmatch(r"(\d+)", value)
    if numeric:
        return f"М{int(numeric.group(1)):03d}"
    safe = re.sub(r"[^0-9A-Za-zА-Яа-яЁё]+", "_", value).strip("_")
    return f"М{safe or 'БЕЗ_НОМЕРА'}"


def schedule_filename(data, options):
    return f"Расписание_{_route_token(data.route_number)}_{options.effective_date:%Y%m%d}_{options.file_token}.xlsx"


def _style_cell(cell, *, center=True):
    cell.border = _BORDER
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    return cell


def _flatten(schedule):
    return [(variant, output) for variant, outputs in schedule.items() for output in outputs]


def _service_metrics(trips):
    if not trips:
        return None, None, 0, 0.0, 0, 0
    first = next((trip.departure_sec for trip in trips if trip.departure_sec is not None), None)
    prior_departure = None
    day_offset = 0
    last = None
    for trip in trips:
        departure = trip.departure_sec
        if departure is not None:
            if prior_departure is not None and departure < prior_departure:
                day_offset += 86400
            prior_departure = departure
        if trip.arrival_sec is not None:
            last = trip.arrival_sec + day_offset
            if departure is not None and trip.arrival_sec < departure:
                last += 86400
    breaks = sum(max(0, trip.break_after_sec) for trip in trips)
    distance = sum(float(trip.distance_km or 0) for trip in trips)
    duration = max(0, last - first) if first is not None and last is not None else 0
    return first, last, breaks, distance, len(trips), duration


def _section_runtime(section):
    if not section.stops:
        return None
    return sum(int(stop.get("run_time_day_sec") or 0) for stop in section.stops)


def _document_header(ws, data):
    ws.oddHeader.left.text = f"Маршрут {_route_token(data.route_number)}"
    ws.oddHeader.right.text = f"Версия {data.version}"
    ws.oddFooter.left.text = f"Сформировано {datetime.date.today():%d.%m.%Y} · версия {data.version}"
    ws.oddFooter.center.text = "Страница &P из &N"


def _visible_footer(ws, row, max_col, data):
    ws.cell(row, 1, f"Сформировано {datetime.date.today():%d.%m.%Y} · Версия {data.version}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    ws.cell(row, 1).font = Font(italic=True, color=NAVY)
    ws.cell(row, 1).alignment = Alignment(horizontal="right")


def _trip_terminals(direction, start_point, end_point):
    normalized = str(direction or "").strip().lower()
    if normalized in {"backward", "back", "reverse", "обратное"}:
        return end_point, start_point
    return start_point, end_point


def _write_day_sheet(ws, data, options, outer_key, day_label):
    outputs = _flatten(data.schedules.get(outer_key, {}))
    variants = {variant for variant, _ in outputs}
    max_trips = max((len(output.trips) for _, output in outputs), default=1)
    headers = ["Выход / тип дня", "Из парка"]
    for index in range(1, max_trips + 1):
        representative = next(
            (output.trips[index - 1] for _, output in outputs
             if len(output.trips) >= index),
            None,
        )
        departure, arrival = _trip_terminals(
            representative.direction if representative else None,
            data.start_point,
            data.end_point,
        )
        headers.extend((f"Рейс {index}\nотправление ({departure})",
                        f"Рейс {index}\nприбытие ({arrival})"))
    headers.extend(("В парк", "Перерывы", "Всего", "Пробег, км"))
    visible_max = max(len(headers), 10)
    helper_start = visible_max + 2
    apply_sheet_setup(ws)
    _document_header(ws, data)
    write_title_band(ws, 1, "МАРШРУТНОЕ РАСПИСАНИЕ", end_col=visible_max)
    ws.cell(2, 1, f"Маршрут: {_route_token(data.route_number)} — {data.route_name}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=visible_max)
    ws.cell(3, 1, _safe_text(f"{data.start_point} — {data.end_point} · {options.season_label} · действует с {options.effective_date:%d.%m.%Y}"))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=visible_max)
    ws.cell(4, 1, day_label)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=visible_max)
    for row in (2, 3, 4):
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="center")
    labels = ("Количество выходов", "Количество рейсов", "Общий пробег, км",
              "Сумма перерывов", "Продолжительность работы")
    for column, label in zip((1, 3, 5, 7, 9), labels):
        for row in (5, 6):
            ws.cell(row, column).fill = PatternFill("solid", fgColor=PALE_GREEN)
            ws.cell(row, column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(5, column, label).font = Font(bold=True, color=NAVY)
        ws.cell(6, column).font = Font(bold=True, color=NAVY, size=12)
    write_table_header(ws, 8, headers)
    helper_headers = ("identity", "first", "last", "break_sec", "distance", "trips", "duration_sec")
    for offset, header in enumerate(helper_headers):
        ws.cell(8, helper_start + offset, header)
        ws.column_dimensions[get_column_letter(helper_start + offset)].hidden = True
    first_row = 9
    depot_out_runtime = _section_runtime(data.depot_out)
    depot_in_runtime = _section_runtime(data.depot_in)
    if not outputs:
        ws.merge_cells(start_row=first_row, start_column=1, end_row=first_row, end_column=visible_max)
        apply_warning_cell(ws.cell(first_row, 1, "Расписание не заполнено"))
        last_row = first_row
    else:
        for row, (variant, output) in enumerate(outputs, start=first_row):
            label = f"{variant} · {output.output_number}" if len(variants) > 1 else output.output_number
            _style_cell(ws.cell(row, 1, _safe_text(label)))
            first, last, breaks, distance, trip_count, duration = _service_metrics(output.trips)
            depot_out = (first - depot_out_runtime) % 86400 if first is not None and depot_out_runtime is not None else None
            write_excel_time(ws.cell(row, 2), depot_out)
            _style_cell(ws.cell(row, 2))
            for index, trip in enumerate(output.trips):
                column = 3 + index * 2
                departure_cell = write_excel_time(ws.cell(row, column), trip.departure_sec)
                arrival_cell = write_excel_time(ws.cell(row, column + 1), trip.arrival_sec)
                departure, arrival = _trip_terminals(
                    trip.direction, data.start_point, data.end_point
                )
                departure_cell.comment = Comment(f"Отправление: {departure}", "ATP")
                arrival_cell.comment = Comment(f"Прибытие: {arrival}", "ATP")
                _style_cell(departure_cell); _style_cell(arrival_cell)
            for column in range(3 + len(output.trips) * 2, 3 + max_trips * 2):
                _style_cell(ws.cell(row, column))
            depot_in_col = 3 + max_trips * 2
            depot_in = (last + depot_in_runtime) % 86400 if last is not None and depot_in_runtime is not None else None
            write_excel_time(ws.cell(row, depot_in_col), depot_in)
            _style_cell(ws.cell(row, depot_in_col))
            park_duration = duration + (depot_out_runtime or 0) + (depot_in_runtime or 0)
            for offset, value in enumerate((1, first, last, breaks, distance, trip_count, park_duration)):
                ws.cell(row, helper_start + offset, value)
            break_letter = get_column_letter(helper_start + 3)
            distance_letter = get_column_letter(helper_start + 4)
            duration_letter = get_column_letter(helper_start + 6)
            formula_cells = (ws.cell(row, depot_in_col + 1, f"={break_letter}{row}/86400"),
                             ws.cell(row, depot_in_col + 2, f"={duration_letter}{row}/86400"))
            for cell in formula_cells:
                cell.number_format = "[h]:mm"; _style_cell(cell)
            distance_cell = ws.cell(row, depot_in_col + 3, f"={distance_letter}{row}")
            distance_cell.number_format = "0.0"; _style_cell(distance_cell)
        last_row = first_row + len(outputs) - 1
    letters = [get_column_letter(helper_start + offset) for offset in range(7)]
    ws.cell(6, 1, f"=COUNT({letters[0]}{first_row}:{letters[0]}{last_row})")
    ws.cell(6, 3, f"=SUM({letters[5]}{first_row}:{letters[5]}{last_row})")
    ws.cell(6, 5, f"=SUM({letters[4]}{first_row}:{letters[4]}{last_row})")
    ws.cell(6, 7, f"=SUM({letters[3]}{first_row}:{letters[3]}{last_row})/86400")
    ws.cell(6, 9, f"=SUM({letters[6]}{first_row}:{letters[6]}{last_row})/86400")
    ws.cell(6, 5).number_format = "0.0"; ws.cell(6, 7).number_format = "[h]:mm"; ws.cell(6, 9).number_format = "[h]:mm"
    footer_row = last_row + 2
    _visible_footer(ws, footer_row, visible_max, data)
    ws.freeze_panes = "C9"; ws.print_title_rows = "1:8"
    set_print_area(ws, max_row=footer_row, max_col=visible_max)
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 12
    for column in range(3, visible_max + 1): ws.column_dimensions[get_column_letter(column)].width = 14
    for row in range(1, footer_row + 1): ws.row_dimensions[row].height = 38 if row == 8 else 24
    ws.row_dimensions[2].height = min(72, 30 + len(data.route_name) // 35 * 12)
    ws.row_dimensions[3].height = 36; ws.row_dimensions[4].height = 30


_SECTION_LABELS = {"forward": "Прямое направление", "backward": "Обратное направление",
                   "depot_out": "Из парка", "depot_in": "В парк"}


def _write_chronometry_sheet(ws, data, options):
    apply_sheet_setup(ws); _document_header(ws, data)
    headers = ("Остановка / внешний ID", "Участок, км", "День", "Ночь",
               "Накоплено день", "Накоплено ночь", "Накоплено, км", "Примечание")
    visible_max = len(headers)
    write_title_band(ws, 1, "ХРОНОМЕТРАЖ МАРШРУТА", end_col=visible_max)
    ws.cell(2, 1, f"Маршрут: {_route_token(data.route_number)} — {data.route_name}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=visible_max)
    ws.cell(3, 1, f"{options.season_label} · действует с {options.effective_date:%d.%m.%Y}")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=visible_max)
    row = 5
    for section in (data.forward, data.backward, data.depot_out, data.depot_in):
        write_section_header(ws, row, _SECTION_LABELS[section.direction], end_col=visible_max)
        row += 1; write_table_header(ws, row, headers); row += 1; first = row
        if not section.stops:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=visible_max)
            apply_warning_cell(ws.cell(row, 1, "Данные хронометража не заполнены")); row += 2; continue
        for stop in section.stops:
            external_id = str(stop.get("external_code") or "").strip()
            name = str(stop.get("name") or "")
            ws.cell(row, 1, _safe_text(f"{name} · {external_id}" if external_id else name))
            ws.cell(row, 2, float(stop.get("distance_from_prev_km") or 0))
            write_excel_time(ws.cell(row, 3), int(stop.get("run_time_day_sec") or 0))
            write_excel_time(ws.cell(row, 4), int(stop.get("run_time_night_sec") or 0))
            ws.cell(row, 5, f"=SUM($C${first}:C{row})").number_format = "[h]:mm"
            ws.cell(row, 6, f"=SUM($D${first}:D{row})").number_format = "[h]:mm"
            ws.cell(row, 7, f"=SUM($B${first}:B{row})").number_format = "0.000"
            for column in range(1, visible_max + 1): _style_cell(ws.cell(row, column), center=column != 1)
            row += 1
        row += 1
    footer_row = row
    _visible_footer(ws, footer_row, visible_max, data)
    helper_col = visible_max + 2
    ws.cell(1, helper_col, "route_version"); ws.cell(2, helper_col, data.version)
    ws.column_dimensions[get_column_letter(helper_col)].hidden = True
    ws.freeze_panes = "B7"; ws.print_title_rows = "1:3"
    set_print_area(ws, max_row=footer_row, max_col=visible_max)
    for index, width in enumerate((38, 12, 12, 12, 17, 17, 17, 20), start=1): ws.column_dimensions[get_column_letter(index)].width = width
    for row in range(1, footer_row + 1): ws.row_dimensions[row].height = 24
    ws.row_dimensions[2].height = min(72, 30 + len(data.route_name) // 35 * 12)


def build_schedule_workbook(data, options):
    workbook = Workbook(); workbook.remove(workbook.active)
    _write_day_sheet(workbook.create_sheet("Рабочие дни"), data, options, "workday", "Рабочие дни")
    _write_day_sheet(workbook.create_sheet("Выходные дни"), data, options, "weekend", "Выходные дни")
    _write_chronometry_sheet(workbook.create_sheet("Хронометраж"), data, options)
    workbook.calculation.fullCalcOnLoad = True; workbook.calculation.forceFullCalc = True
    return workbook


def _xlsx_download_response(workbook, filename):
    stream = BytesIO(); workbook.save(stream); stream.seek(0)
    safe_filename = str(filename).replace("\r", "").replace("\n", "")
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe_filename, safe='')}"})
