# -*- coding: utf-8 -*-
"""Разбор, предпросмотр и применение табличного импорта трассы."""
import csv
import io
from pathlib import Path

from openpyxl import load_workbook

from .route_network import normalize_stop_name, recalculate_trace


REQUIRED_COLUMNS = {"direction", "sequence", "name"}
FIELD_NAMES = (
    "direction", "sequence", "external_code", "name", "latitude", "longitude",
    "address", "distance_km", "run_time_sec", "dwell_time_sec",
)


def _as_int(value, field):
    try:
        return int(value)
    except (TypeError, ValueError):
        raise ValueError(f"Поле {field} должно быть целым числом")


def _as_float(value, field, default=None):
    if value is None or str(value).strip() == "":
        return default
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        raise ValueError(f"Поле {field} должно быть числом")


def _direction(value):
    text = str(value or "").strip().lower()
    aliases = {
        "forward": "forward", "прямое": "forward", "прямое направление": "forward",
        "backward": "backward", "обратное": "backward", "обратное направление": "backward",
    }
    if text not in aliases:
        raise ValueError("direction должно быть forward/backward или прямое/обратное")
    return aliases[text]


def _read_csv(data):
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return reader.fieldnames or [], list(reader)


def _read_xlsx(data):
    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать файл Excel") from exc
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(iterator)]
    except StopIteration:
        return [], []
    rows = [dict(zip(headers, values)) for values in iterator if any(value is not None for value in values)]
    return headers, rows


def parse_network_file(filename, data):
    extension = Path(filename or "").suffix.lower()
    if extension == ".csv":
        headers, raw_rows = _read_csv(data)
    elif extension == ".xlsx":
        headers, raw_rows = _read_xlsx(data)
    else:
        raise ValueError("Поддерживаются только файлы .csv и .xlsx")
    missing = sorted(REQUIRED_COLUMNS - set(headers))
    if missing:
        raise ValueError("Отсутствуют обязательные столбцы: " + ", ".join(missing))

    grouped = {"forward": [], "backward": []}
    for line_number, raw in enumerate(raw_rows, start=2):
        name = str(raw.get("name") or "").strip()
        if not name:
            raise ValueError(f"Строка {line_number}: не заполнено name")
        direction = _direction(raw.get("direction"))
        row = {
            "direction": direction,
            "sequence": _as_int(raw.get("sequence"), "sequence"),
            "external_code": str(raw.get("external_code") or "").strip() or None,
            "name": name,
            "latitude": _as_float(raw.get("latitude"), "latitude"),
            "longitude": _as_float(raw.get("longitude"), "longitude"),
            "address": str(raw.get("address") or "").strip() or None,
            "distance_from_prev_km": _as_float(raw.get("distance_km"), "distance_km", 0.0),
            "run_time_sec": _as_int(raw.get("run_time_sec") or 0, "run_time_sec"),
            "dwell_time_sec": _as_int(raw.get("dwell_time_sec") or 0, "dwell_time_sec"),
        }
        if row["latitude"] is not None:
            row["latitude"] = round(row["latitude"], 6)
        if row["longitude"] is not None:
            row["longitude"] = round(row["longitude"], 6)
        if row["run_time_sec"] < 0 or row["dwell_time_sec"] < 0:
            raise ValueError(f"Строка {line_number}: время не может быть отрицательным")
        grouped[direction].append(row)

    for direction in grouped:
        grouped[direction] = recalculate_trace(grouped[direction])
    return grouped


def build_import_plan(con, route_id, grouped):
    if not con.execute("SELECT 1 FROM routes WHERE id=?", (route_id,)).fetchone():
        raise ValueError("Маршрут не найден")
    existing = [dict(row) for row in con.execute("SELECT * FROM stops ORDER BY id")]
    new_stops = {}
    conflicts = []
    matched_count = 0
    planned_rows = {"forward": [], "backward": []}

    for direction in ("forward", "backward"):
        for source in grouped[direction]:
            row = dict(source)
            code = str(row.get("external_code") or "")
            key = code or normalize_stop_name(row["name"])
            coded = [stop for stop in existing if code and str(stop.get("external_code") or "") == code]
            named = [
                stop for stop in existing
                if normalize_stop_name(stop.get("name") or "") == normalize_stop_name(row["name"])
            ]
            if len(coded) == 1:
                row["stop_ref"] = {"kind": "existing", "id": coded[0]["id"]}
                matched_count += 1
            elif len(coded) > 1 or (not code and len(named) > 1):
                candidates = coded if coded else named
                conflicts.append({
                    "direction": direction,
                    "sequence": row["sequence"],
                    "name": row["name"],
                    "candidate_ids": [stop["id"] for stop in candidates],
                })
                row["stop_ref"] = None
            elif not code and len(named) == 1:
                row["stop_ref"] = {"kind": "existing", "id": named[0]["id"]}
                matched_count += 1
            elif key in new_stops:
                row["stop_ref"] = {"kind": "new", "key": key}
                matched_count += 1
            else:
                new_stops[key] = {
                    "external_code": row.get("external_code"),
                    "name": row["name"],
                    "latitude": row.get("latitude"),
                    "longitude": row.get("longitude"),
                    "address": row.get("address"),
                }
                row["stop_ref"] = {"kind": "new", "key": key}
            planned_rows[direction].append(row)

    summary = {
        "created_stops": len(new_stops),
        "matched_stops": matched_count,
        "conflicts": len(conflicts),
        "forward_stops": len(planned_rows["forward"]),
        "backward_stops": len(planned_rows["backward"]),
    }
    return {
        "route_id": route_id,
        "rows": planned_rows,
        "new_stops": new_stops,
        "conflicts": conflicts,
        "summary": summary,
    }


def apply_import_plan(con, route_id, plan, timestamp):
    if int(plan.get("route_id") or 0) != route_id:
        raise ValueError("Предпросмотр создан для другого маршрута")
    if plan.get("conflicts"):
        raise ValueError("В предпросмотре есть неразрешённые конфликты остановок")

    new_ids = {}
    for key, stop in plan["new_stops"].items():
        cur = con.execute(
            "INSERT INTO stops(external_code,name,latitude,longitude,address,source,created_at,updated_at) "
            "VALUES(?,?,?,?,?,'table_import',?,?)",
            (stop.get("external_code"), stop["name"], stop.get("latitude"),
             stop.get("longitude"), stop.get("address"), timestamp, timestamp),
        )
        new_ids[key] = cur.lastrowid

    con.execute("DELETE FROM route_stops WHERE route_id=?", (route_id,))
    totals = {}
    names = {}
    for direction in ("forward", "backward"):
        saved_names = []
        for row in plan["rows"][direction]:
            ref = row["stop_ref"]
            stop_id = ref["id"] if ref["kind"] == "existing" else new_ids[ref["key"]]
            stop = con.execute("SELECT * FROM stops WHERE id=?", (stop_id,)).fetchone()
            saved_names.append(stop["name"])
            con.execute(
                "INSERT INTO route_stops("
                "route_id,direction,stop_id,sequence,distance_from_prev_km,cumulative_km,"
                "run_time_sec,dwell_time_sec,distance_source,created_at,updated_at"
                ") VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                (route_id, direction, stop_id, row["sequence"], row["distance_from_prev_km"],
                 row["cumulative_km"], row["run_time_sec"], row["dwell_time_sec"],
                 "table_import", timestamp, timestamp),
            )
        names[direction] = ", ".join(saved_names)
        rows = plan["rows"][direction]
        totals[direction] = rows[-1]["cumulative_km"] if rows else 0

    con.execute(
        "UPDATE routes SET stops=?,stops_back=?,length_km=?,length_back_km=? WHERE id=?",
        (names["forward"], names["backward"], totals["forward"], totals["backward"], route_id),
    )
    return {"summary": plan["summary"], "totals": totals}
