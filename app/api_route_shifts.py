# -*- coding: utf-8 -*-
"""Shift types and per-route structural shift settings."""

import datetime
import json
import re
import secrets
import sqlite3
from collections import defaultdict

from fastapi import APIRouter, Body, Depends, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import db
from .auth import current_user, require_write
from .route_shifts import build_output_shifts, validate_output_shift_plan
from .xl import _xlsx_download_response


router = APIRouter(prefix="/api")
SHIFT_CODE_RE = re.compile(r"^[a-z0-9_]+$")
EXPORT_DARK = "17365D"
EXPORT_ALT = "F5F8FC"
EXPORT_MANUAL = "FFF2CC"
EXPORT_TWO_DRIVER = "E2F0D9"


def _shift_type_payload(row):
    if not row:
        return None
    item = dict(row)
    item["allow_split"] = bool(item["allow_split"])
    item["active"] = bool(item["active"])
    return item


def _shift_type_by_id(con, shift_type_id):
    if shift_type_id is None:
        return None
    return db.one(
        con.execute("SELECT * FROM shift_types WHERE id=?", (shift_type_id,))
    )


def _validated_shift_type(payload):
    code = str(payload.get("code") or "").strip().lower()
    name = str(payload.get("name") or "").strip()
    if not code or not SHIFT_CODE_RE.fullmatch(code):
        raise ValueError("Код типа смены: только латинские буквы, цифры и подчёркивание")
    if not name:
        raise ValueError("Укажите название типа смены")
    try:
        planned = int(payload.get("planned_duration_min"))
        maximum = int(payload.get("max_duration_min"))
        driver_slots = int(payload.get("driver_slots", 1))
    except (TypeError, ValueError):
        raise ValueError("Длительность и количество водителей должны быть числами")
    if planned <= 0:
        raise ValueError("Плановая длительность должна быть больше нуля")
    if maximum < planned:
        raise ValueError("Максимальная длительность не может быть меньше плановой")
    if driver_slots not in (1, 2):
        raise ValueError("Количество водительских мест должно быть 1 или 2")
    color = str(payload.get("color") or "#2563eb").strip()
    if not re.fullmatch(r"#[0-9a-fA-F]{6}", color):
        raise ValueError("Цвет должен быть задан в формате #RRGGBB")
    return {
        "code": code,
        "name": name,
        "work_pattern": str(payload.get("work_pattern") or "custom").strip(),
        "planned_duration_min": planned,
        "max_duration_min": maximum,
        "driver_slots": driver_slots,
        "allow_split": 1 if payload.get("allow_split") else 0,
        "color": color.lower(),
        "active": 1 if payload.get("active", True) else 0,
    }


def _route_or_404(con, route_id):
    route = db.one(con.execute("SELECT * FROM routes WHERE id=?", (route_id,)))
    if not route:
        raise HTTPException(404, "Маршрут не найден")
    return route


def _active_type_or_error(con, shift_type_id, field_name, *, optional=False):
    if shift_type_id in (None, ""):
        if optional:
            return None
        raise ValueError(f"Укажите {field_name}")
    try:
        shift_type_id = int(shift_type_id)
    except (TypeError, ValueError):
        raise ValueError(f"Некорректный {field_name}")
    row = _shift_type_by_id(con, shift_type_id)
    if not row:
        raise ValueError(f"{field_name.capitalize()} не найден")
    if not row["active"]:
        raise ValueError(f"{field_name.capitalize()} должен быть активным")
    return row


def _settings_payload(con, route_id, day_type):
    settings = db.one(
        con.execute(
            "SELECT * FROM route_shift_settings WHERE route_id=? AND day_type=?",
            (route_id, day_type),
        )
    )
    persisted = settings is not None
    if not settings:
        default_type = db.one(
            con.execute("SELECT * FROM shift_types WHERE code='single_8h'")
        )
        long_type = db.one(
            con.execute("SELECT * FROM shift_types WHERE code='two_driver_long'")
        )
        settings = {
            "id": None,
            "route_id": route_id,
            "day_type": day_type,
            "default_shift_type_id": default_type["id"],
            "long_shift_type_id": long_type["id"],
            "handover_min": 10,
            "long_run_threshold_min": 720,
            "auto_split": 1,
            "updated_at": None,
        }
    default_type = _shift_type_by_id(con, settings["default_shift_type_id"])
    long_type = _shift_type_by_id(con, settings.get("long_shift_type_id"))
    return {
        **settings,
        "auto_split": bool(settings["auto_split"]),
        "persisted": persisted,
        "default_shift_type": _shift_type_payload(default_type),
        "long_shift_type": _shift_type_payload(long_type),
    }


@router.get("/shift-types")
def shift_types_list(active_only: bool = True, user=Depends(current_user)):
    con = db.connect()
    try:
        where = "WHERE active=1" if active_only else ""
        rows = db.rows(
            con.execute(f"SELECT * FROM shift_types {where} ORDER BY name,code,id")
        )
        return {"items": [_shift_type_payload(row) for row in rows]}
    finally:
        con.close()


@router.post("/shift-types")
def shift_type_save(payload: dict = Body(...), user=Depends(current_user)):
    require_write(user, "trips")
    try:
        values = _validated_shift_type(payload)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    con = db.connect()
    try:
        timestamp = datetime.datetime.now().isoformat(timespec="seconds")
        shift_type_id = payload.get("id")
        old = None
        if shift_type_id:
            old = _shift_type_by_id(con, shift_type_id)
            if not old:
                raise HTTPException(404, "Тип смены не найден")
            con.execute(
                """
                UPDATE shift_types
                SET code=?,name=?,work_pattern=?,planned_duration_min=?,
                    max_duration_min=?,driver_slots=?,allow_split=?,color=?,
                    active=?,updated_at=? WHERE id=?
                """,
                (*values.values(), timestamp, shift_type_id),
            )
            action = "изменение типа смены"
        else:
            cursor = con.execute(
                """
                INSERT INTO shift_types(
                  code,name,work_pattern,planned_duration_min,max_duration_min,
                  driver_slots,allow_split,color,active,created_at,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
                """,
                (*values.values(), timestamp, timestamp),
            )
            shift_type_id = cursor.lastrowid
            action = "создание типа смены"
        saved = _shift_type_payload(_shift_type_by_id(con, shift_type_id))
        db.audit(con, user["username"], action, "shift_types", shift_type_id,
                 old=old, new=saved)
        con.commit()
        return saved
    except sqlite3.IntegrityError:
        con.rollback()
        raise HTTPException(400, "Тип смены с таким кодом уже существует")
    finally:
        con.close()


@router.get("/routes/{route_id}/shift-settings/{day_type}")
def route_shift_settings_get(route_id: int, day_type: str,
                             user=Depends(current_user)):
    con = db.connect()
    try:
        _route_or_404(con, route_id)
        return _settings_payload(con, route_id, day_type)
    finally:
        con.close()


@router.put("/routes/{route_id}/shift-settings/{day_type}")
def route_shift_settings_save(route_id: int, day_type: str,
                              payload: dict = Body(...),
                              user=Depends(current_user)):
    require_write(user, "trips")
    con = db.connect()
    try:
        _route_or_404(con, route_id)
        default_type = _active_type_or_error(
            con, payload.get("default_shift_type_id"), "основной тип смены"
        )
        long_type = _active_type_or_error(
            con, payload.get("long_shift_type_id"), "тип длинной смены",
            optional=True,
        )
        try:
            handover_min = int(payload.get("handover_min", 10))
            threshold = int(payload.get("long_run_threshold_min", 720))
        except (TypeError, ValueError):
            raise ValueError("Параметры длительности должны быть числами")
        if handover_min < 0:
            raise ValueError("Время пересмены не может быть отрицательным")
        if threshold <= 0:
            raise ValueError("Порог длинного выпуска должен быть больше нуля")
        old = db.one(con.execute(
            "SELECT * FROM route_shift_settings WHERE route_id=? AND day_type=?",
            (route_id, day_type),
        ))
        timestamp = datetime.datetime.now().isoformat(timespec="seconds")
        con.execute(
            """
            INSERT INTO route_shift_settings(
              route_id,day_type,default_shift_type_id,long_shift_type_id,
              handover_min,long_run_threshold_min,auto_split,updated_at
            ) VALUES(?,?,?,?,?,?,?,?)
            ON CONFLICT(route_id,day_type) DO UPDATE SET
              default_shift_type_id=excluded.default_shift_type_id,
              long_shift_type_id=excluded.long_shift_type_id,
              handover_min=excluded.handover_min,
              long_run_threshold_min=excluded.long_run_threshold_min,
              auto_split=excluded.auto_split,updated_at=excluded.updated_at
            """,
            (route_id, day_type, default_type["id"],
             long_type["id"] if long_type else None, handover_min, threshold,
             1 if payload.get("auto_split", True) else 0, timestamp),
        )
        saved = _settings_payload(con, route_id, day_type)
        db.audit(con, user["username"], "настройка смен маршрута",
                 "route_shift_settings", saved["id"], old=old, new=saved)
        con.commit()
        return saved
    except ValueError as exc:
        con.rollback()
        raise HTTPException(400, str(exc))
    finally:
        con.close()


@router.get("/routes/{route_id}/output-shifts")
def route_output_shifts_list(route_id: int, day_type: str = "",
                             user=Depends(current_user)):
    day_type = str(day_type or "").strip()
    if not day_type:
        raise HTTPException(400, "Укажите тип дня")
    con = db.connect()
    try:
        _route_or_404(con, route_id)
        items = db.rows(con.execute(
            """
            SELECT os.id, os.route_id, os.day_type, os.output_number,
                   os.shift_number, os.shift_type_id,
                   st.code AS shift_type_code, st.name AS shift_type_name,
                   st.color AS shift_type_color,
                   os.trip_from_id, first_trip.trip_number AS trip_from_number,
                   os.trip_to_id, last_trip.trip_number AS trip_to_number,
                   os.start_sec, os.end_sec,
                   os.end_sec - os.start_sec AS duration_sec,
                   os.driver_slots, os.handover_after_min, os.source,
                   os.is_manual_locked, os.manual_reason,
                   COALESCE(assignments.assignment_count, 0) AS assignment_count
            FROM output_shifts os
            JOIN shift_types st ON st.id=os.shift_type_id
            JOIN route_trips first_trip ON first_trip.id=os.trip_from_id
            JOIN route_trips last_trip ON last_trip.id=os.trip_to_id
            LEFT JOIN (
              SELECT output_shift_id, COUNT(*) AS assignment_count
              FROM roster_assignments
              WHERE output_shift_id IS NOT NULL
              GROUP BY output_shift_id
            ) assignments ON assignments.output_shift_id=os.id
            WHERE os.route_id=? AND os.day_type=?
            ORDER BY os.output_number,os.shift_number,os.id
            """,
            (route_id, day_type),
        ))
        for item in items:
            item["is_manual_locked"] = bool(item["is_manual_locked"])
        return {
            "route_id": route_id,
            "day_type": day_type,
            "assignment_count_scope": "all_dates",
            "items": items,
        }
    finally:
        con.close()


def _service_clock(seconds):
    seconds = int(seconds)
    return f"{seconds // 3600:02d}:{seconds % 3600 // 60:02d}"


def _prepare_shift_export_sheet(ws, *, title, metadata, headers, widths):
    last_column = len(headers)
    ws.title = "Смены выходов"
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=EXPORT_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 27
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    metadata_cell = ws.cell(2, 1, metadata)
    metadata_cell.font = Font(italic=True, color="44546A")
    metadata_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22
    thin = Side(style="thin", color="B4C6E7")
    for column, header in enumerate(headers, 1):
        cell = ws.cell(3, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=EXPORT_DARK)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=thin, right=thin)
        ws.column_dimensions[get_column_letter(column)].width = widths[column - 1]
    ws.row_dimensions[3].height = 34
    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def _finish_shift_export_sheet(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row_index, row in enumerate(ws.iter_rows(min_row=4), 4):
        zebra = PatternFill("solid", fgColor=EXPORT_ALT) if row_index % 2 == 0 else None
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if zebra and cell.fill.fill_type is None:
                cell.fill = zebra
        ws.row_dimensions[row_index].height = 30
    end_row = max(3, ws.max_row)
    last_column = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A3:{last_column}{end_row}"
    ws.print_area = f"A1:{last_column}{end_row}"


@router.get("/routes/{route_id}/output-shifts/export.xlsx")
def route_output_shifts_export(route_id: int, day_type: str = "",
                               service_date: str = "",
                               user=Depends(current_user)):
    day_type = str(day_type or "").strip()
    if not day_type:
        raise HTTPException(400, "Укажите тип дня")
    service_date = str(service_date or "").strip()
    parsed_date = None
    if service_date:
        try:
            parsed_date = datetime.date.fromisoformat(service_date)
        except ValueError:
            raise HTTPException(400, "Дата должна быть в формате ГГГГ-ММ-ДД")
    con = db.connect()
    try:
        route = _route_or_404(con, route_id)
        count_where = "WHERE output_shift_id IS NOT NULL"
        query_args = []
        if service_date:
            count_where += " AND date=?"
            query_args.append(service_date)
        rows = db.rows(con.execute(
            f"""
            SELECT os.id, os.output_number, os.shift_number,
                   st.name AS shift_type_name,
                   first_trip.trip_number AS trip_from_number,
                   last_trip.trip_number AS trip_to_number,
                   os.start_sec, os.end_sec, os.driver_slots,
                   os.handover_after_min, os.source, os.is_manual_locked,
                   os.manual_reason,
                   COALESCE(assignments.assignment_count, 0) AS assignment_count
            FROM output_shifts os
            JOIN shift_types st ON st.id=os.shift_type_id
            JOIN route_trips first_trip ON first_trip.id=os.trip_from_id
            JOIN route_trips last_trip ON last_trip.id=os.trip_to_id
            LEFT JOIN (
              SELECT output_shift_id, COUNT(*) AS assignment_count
              FROM roster_assignments
              {count_where}
              GROUP BY output_shift_id
            ) assignments ON assignments.output_shift_id=os.id
            WHERE os.route_id=? AND os.day_type=?
            ORDER BY os.output_number,os.shift_number,os.id
            """,
            (*query_args, route_id, day_type),
        ))
        if parsed_date:
            count_scope = f"дата назначений: {parsed_date.strftime('%d.%m.%Y')}"
        else:
            count_scope = "назначения: все даты"
        headers = [
            "Выход", "Смена", "Тип смены", "Диапазон рейсов", "Начало",
            "Окончание", "Длительность, мин", "Длительность, ч",
            "Водительских мест", "Пересмена, мин", "Источник", "Ручная",
            "Причина ручной правки", "Назначений",
        ]
        widths = [9, 9, 22, 18, 12, 12, 17, 16, 17, 16, 15, 11, 32, 14]
        workbook = Workbook()
        sheet = workbook.active
        _prepare_shift_export_sheet(
            sheet,
            title=f"Смены выходов маршрута № {route['number']}",
            metadata=(f"{route.get('name') or 'Без наименования'} · "
                      f"тип дня: {day_type} · {count_scope}"),
            headers=headers,
            widths=widths,
        )
        for row_index, item in enumerate(rows, 4):
            duration_min = (int(item["end_sec"]) - int(item["start_sec"])) // 60
            sheet.append([
                int(item["output_number"]), int(item["shift_number"]),
                item["shift_type_name"],
                f"{item['trip_from_number']}–{item['trip_to_number']}",
                _service_clock(item["start_sec"]), _service_clock(item["end_sec"]),
                duration_min, duration_min / 60, int(item["driver_slots"]),
                int(item["handover_after_min"]), item["source"],
                "Да" if item["is_manual_locked"] else "Нет",
                item["manual_reason"] or "", int(item["assignment_count"]),
            ])
            fill = None
            if item["is_manual_locked"]:
                fill = PatternFill("solid", fgColor=EXPORT_MANUAL)
            elif int(item["driver_slots"]) == 2:
                fill = PatternFill("solid", fgColor=EXPORT_TWO_DRIVER)
            if fill:
                for cell in sheet[row_index]:
                    cell.fill = fill
            sheet.cell(row_index, 7).number_format = "#,##0"
            sheet.cell(row_index, 8).number_format = "0.00"
            for column in (1, 2, 9, 10, 14):
                sheet.cell(row_index, column).number_format = "#,##0"
        _finish_shift_export_sheet(sheet)
        return _xlsx_download_response(
            workbook, f"route_{route['number']}_{day_type}_output_shifts.xlsx"
        )
    finally:
        con.close()


def _clock_seconds(value):
    parts = str(value or "").strip().split(":")
    if len(parts) != 2:
        raise ValueError(f"Некорректное время: {value}")
    try:
        hours, minutes = (int(part) for part in parts)
    except ValueError:
        raise ValueError(f"Некорректное время: {value}")
    if hours < 0 or hours >= 48 or minutes < 0 or minutes >= 60:
        raise ValueError(f"Некорректное время: {value}")
    return hours * 3600 + minutes * 60


def _generation_trips(con, route_id, day_type):
    rows = db.rows(con.execute(
        "SELECT * FROM route_trips WHERE route_id=? AND day_type=? "
        "ORDER BY output_number,COALESCE(trip_number,id),id",
        (route_id, day_type),
    ))
    grouped = defaultdict(list)
    for row in rows:
        grouped[int(row["output_number"])].append(dict(row))
    result = []
    for output_number in sorted(grouped):
        previous_dep = None
        enriched = []
        for row in grouped[output_number]:
            dep_sec = _clock_seconds(row["dep_time"])
            if previous_dep is not None:
                while dep_sec < previous_dep:
                    dep_sec += 24 * 3600
            arr_sec = _clock_seconds(row["arr_time"])
            while arr_sec <= dep_sec:
                arr_sec += 24 * 3600
            enriched.append({**row, "dep_sec": dep_sec, "arr_sec": arr_sec})
            previous_dep = dep_sec
        result.extend(sorted(enriched, key=lambda item: (item["dep_sec"], item["id"])))
    return result


def _locked_shift_payload(row):
    return {
        "id": int(row["id"]),
        "shift_number": int(row["shift_number"]),
        "shift_type_id": int(row["shift_type_id"]),
        "output_number": int(row["output_number"]),
        "trip_from_id": int(row["trip_from_id"]),
        "trip_to_id": int(row["trip_to_id"]),
        "start_sec": int(row["start_sec"]),
        "end_sec": int(row["end_sec"]),
        "driver_slots": int(row["driver_slots"]),
        "handover_after_min": int(row["handover_after_min"]),
        "is_manual_locked": True,
    }


def _plan_conflict(output_number, exc, code="generation_failed"):
    return {
        "code": code,
        "output_number": output_number,
        "message": str(exc),
    }


class LockedPlanChanged(ValueError):
    pass


class PreviewStateChanged(ValueError):
    pass


def _shift_type_for_output(settings, output_trips):
    duration = output_trips[-1]["arr_sec"] - output_trips[0]["dep_sec"]
    long_type = settings.get("long_shift_type")
    if (
        duration >= int(settings["long_run_threshold_min"]) * 60
        and long_type
        and long_type.get("active")
    ):
        return long_type
    return settings["default_shift_type"]


def _require_active_generated_type(shift_type):
    if not shift_type.get("active"):
        raise ValueError("Выбранный тип смены неактивен")
    return shift_type


def _build_around_locked(output_trips, locked, *, shift_type, handover_min):
    conflicts = [
        conflict
        for conflict in validate_output_shift_plan(output_trips, locked)
        if conflict["code"] != "uncovered_trip"
    ]
    if conflicts:
        raise ValueError(
            "Заблокированные смены нельзя совместить с рейсами выпуска"
        )
    positions = {
        int(trip["id"]): index for index, trip in enumerate(output_trips)
    }
    ranges = sorted(
        (
            positions[int(shift["trip_from_id"])],
            positions[int(shift["trip_to_id"])],
        )
        for shift in locked
    )
    generated = []
    start = 0
    for first, last in ranges:
        if start < first:
            _require_active_generated_type(shift_type)
            generated.extend(build_output_shifts(
                output_trips[start:first],
                shift_type=shift_type,
                handover_min=handover_min,
            ))
        start = last + 1
    if start < len(output_trips):
        _require_active_generated_type(shift_type)
        generated.extend(build_output_shifts(
            output_trips[start:],
            shift_type=shift_type,
            handover_min=handover_min,
        ))

    used_numbers = {int(shift["shift_number"]) for shift in locked}
    next_number = 1
    for shift in sorted(generated, key=lambda item: (item["start_sec"], item["end_sec"])):
        while next_number in used_numbers:
            next_number += 1
        shift["shift_number"] = next_number
        used_numbers.add(next_number)
        next_number += 1
    combined = sorted(
        [*locked, *generated],
        key=lambda item: (item["start_sec"], item["end_sec"], item["shift_number"]),
    )
    handover_sec = int(handover_min) * 60
    for previous, current in zip(combined, combined[1:]):
        gap = int(current["start_sec"]) - int(previous["end_sec"])
        if gap < handover_sec:
            raise ValueError(
                "Нет допустимого времени пересмены рядом с заблокированной сменой"
            )
        if not previous.get("is_manual_locked"):
            previous["handover_after_min"] = gap // 60
    if validate_output_shift_plan(output_trips, combined):
        raise ValueError("План с заблокированными сменами не покрывает выпуск")
    return combined


def _generation_state(con, settings, shift_type_ids):
    types = []
    for shift_type_id in sorted(set(shift_type_ids)):
        row = _shift_type_by_id(con, shift_type_id)
        if not row:
            raise ValueError("План ссылается на неизвестный тип смены")
        types.append({
            "id": int(row["id"]),
            "active": bool(row["active"]),
            "planned_duration_min": int(row["planned_duration_min"]),
            "max_duration_min": int(row["max_duration_min"]),
            "driver_slots": int(row["driver_slots"]),
        })
    return {
        "settings": {
            "default_shift_type_id": int(settings["default_shift_type_id"]),
            "long_shift_type_id": (
                int(settings["long_shift_type_id"])
                if settings.get("long_shift_type_id") is not None else None
            ),
            "handover_min": int(settings["handover_min"]),
            "long_run_threshold_min": int(settings["long_run_threshold_min"]),
            "auto_split": bool(settings["auto_split"]),
        },
        "shift_types": types,
    }


@router.post("/routes/{route_id}/shift-generation/preview")
def route_shift_generation_preview(route_id: int, payload: dict = Body(...),
                                   user=Depends(current_user)):
    require_write(user, "trips")
    day_type = str(payload.get("day_type") or "").strip()
    if not day_type:
        raise HTTPException(400, "Укажите тип дня")
    preserve_locked = bool(payload.get("preserve_locked", True))
    con = db.connect()
    try:
        _route_or_404(con, route_id)
        trips = _generation_trips(con, route_id, day_type)
        if not trips:
            raise HTTPException(400, "Для маршрута и типа дня нет рейсов")
        settings = _settings_payload(con, route_id, day_type)
        by_output = defaultdict(list)
        for trip in trips:
            by_output[int(trip["output_number"])].append(trip)
        locked_by_output = defaultdict(list)
        locked_rows = db.rows(con.execute(
            "SELECT * FROM output_shifts WHERE route_id=? AND day_type=? "
            "AND is_manual_locked=1 ORDER BY output_number,shift_number,id",
            (route_id, day_type),
        ))
        for row in locked_rows:
            locked_by_output[int(row["output_number"])].append(
                _locked_shift_payload(row)
            )

        outputs = []
        conflicts = []
        for output_number in sorted(by_output):
            output_trips = by_output[output_number]
            locked = locked_by_output.get(output_number, [])
            shifts = []
            output_conflicts = []
            if locked:
                if not preserve_locked:
                    output_conflicts.append(_plan_conflict(
                        output_number,
                        "Выпуск содержит заблокированные ручные смены",
                        "locked_shifts_present",
                    ))
                else:
                    try:
                        shifts = _build_around_locked(
                            output_trips,
                            locked,
                            shift_type=_shift_type_for_output(
                                settings, output_trips
                            ),
                            handover_min=settings["handover_min"],
                        )
                    except ValueError as exc:
                        output_conflicts.append(_plan_conflict(
                            output_number, exc, "locked_shift_conflict"
                        ))
            else:
                try:
                    shift_type = _require_active_generated_type(
                        _shift_type_for_output(settings, output_trips)
                    )
                    shifts = build_output_shifts(
                        output_trips,
                        shift_type=shift_type,
                        handover_min=settings["handover_min"],
                    )
                except ValueError as exc:
                    output_conflicts.append(_plan_conflict(output_number, exc))
            conflicts.extend(output_conflicts)
            outputs.append({
                "output_number": output_number,
                "shifts": shifts,
                "conflicts": output_conflicts,
            })

        old_rows = db.rows(con.execute(
            "SELECT driver_slots FROM output_shifts WHERE route_id=? AND day_type=?",
            (route_id, day_type),
        ))
        new_shifts = [
            shift for output in outputs for shift in output["shifts"]
        ]
        diff = {
            "old_shift_count": len(old_rows),
            "new_shift_count": len(new_shifts),
            "old_driver_slots": sum(int(row["driver_slots"]) for row in old_rows),
            "new_driver_slots": sum(int(row["driver_slots"]) for row in new_shifts),
        }
        token = secrets.token_hex(16)
        now = datetime.datetime.now()
        expires_at = now + datetime.timedelta(minutes=30)
        plan = {
            "scope": {
                "route_id": route_id,
                "day_type": day_type,
                "username": user["username"],
            },
            "route_id": route_id,
            "day_type": day_type,
            "preserve_locked": preserve_locked,
            "outputs": outputs,
            "conflicts": conflicts,
            "diff": diff,
            "locked_shifts": sorted(
                (_locked_shift_payload(row) for row in locked_rows),
                key=lambda item: item["id"],
            ),
        }
        used_type_ids = {
            int(shift["shift_type_id"])
            for output in outputs for shift in output["shifts"]
        }
        plan["generation_state"] = _generation_state(con, settings, used_type_ids)
        con.execute(
            "INSERT INTO shift_generation_previews("
            "token,route_id,day_type,username,payload_json,created_at,expires_at"
            ") VALUES(?,?,?,?,?,?,?)",
            (token, route_id, day_type, user["username"],
             json.dumps(plan, ensure_ascii=False),
             now.isoformat(timespec="seconds"),
             expires_at.isoformat(timespec="seconds")),
        )
        con.commit()
        return {
            "preview_token": token,
            "expires_at": expires_at.isoformat(timespec="seconds"),
            **plan,
        }
    except ValueError as exc:
        con.rollback()
        raise HTTPException(400, str(exc))
    finally:
        con.close()


SQLITE_INT_MAX = 2 ** 63 - 1
MAX_SERVICE_SECONDS = 7 * 24 * 3600


def _required_dict(value, label):
    if not isinstance(value, dict):
        raise ValueError(f"{label}: ожидается объект")
    return value


def _required_list(value, label):
    if not isinstance(value, list):
        raise ValueError(f"{label}: ожидается список")
    return value


def _required_int(value, label, *, minimum=0, maximum=SQLITE_INT_MAX):
    if isinstance(value, bool) or not isinstance(value, int):
        raise ValueError(f"{label}: ожидается целое число")
    if value < minimum or value > maximum:
        raise ValueError(f"{label}: число вне допустимого диапазона")
    return value


def _validate_shift_json(shift, label, *, require_locked=False):
    shift = _required_dict(shift, label)
    _required_int(shift.get("shift_number"), f"{label}.shift_number", minimum=1)
    _required_int(shift.get("shift_type_id"), f"{label}.shift_type_id", minimum=1)
    _required_int(shift.get("output_number"), f"{label}.output_number", minimum=1)
    _required_int(shift.get("trip_from_id"), f"{label}.trip_from_id", minimum=1)
    _required_int(shift.get("trip_to_id"), f"{label}.trip_to_id", minimum=1)
    start = _required_int(
        shift.get("start_sec"), f"{label}.start_sec",
        maximum=MAX_SERVICE_SECONDS,
    )
    end = _required_int(
        shift.get("end_sec"), f"{label}.end_sec",
        minimum=1, maximum=MAX_SERVICE_SECONDS,
    )
    if end <= start:
        raise ValueError(f"{label}: окончание должно быть позже начала")
    slots = _required_int(
        shift.get("driver_slots"), f"{label}.driver_slots", minimum=1,
        maximum=2,
    )
    if slots not in (1, 2):
        raise ValueError(f"{label}.driver_slots: допустимо 1 или 2")
    _required_int(
        shift.get("handover_after_min", 0), f"{label}.handover_after_min",
        maximum=7 * 24 * 60,
    )
    locked = shift.get("is_manual_locked", False)
    if not isinstance(locked, bool):
        raise ValueError(f"{label}.is_manual_locked: ожидается boolean")
    if require_locked and not locked:
        raise ValueError(f"{label}: ожидается заблокированная смена")
    if locked:
        _required_int(shift.get("id"), f"{label}.id", minimum=1)
    return shift


def _validate_generation_state_json(state):
    state = _required_dict(state, "generation_state")
    settings = _required_dict(state.get("settings"), "generation_state.settings")
    _required_int(
        settings.get("default_shift_type_id"),
        "generation_state.settings.default_shift_type_id", minimum=1,
    )
    long_id = settings.get("long_shift_type_id")
    if long_id is not None:
        _required_int(
            long_id, "generation_state.settings.long_shift_type_id", minimum=1
        )
    _required_int(
        settings.get("handover_min"), "generation_state.settings.handover_min",
        maximum=7 * 24 * 60,
    )
    _required_int(
        settings.get("long_run_threshold_min"),
        "generation_state.settings.long_run_threshold_min", minimum=1,
        maximum=7 * 24 * 60,
    )
    if not isinstance(settings.get("auto_split"), bool):
        raise ValueError("generation_state.settings.auto_split: ожидается boolean")
    types = _required_list(state.get("shift_types"), "generation_state.shift_types")
    type_ids = set()
    for index, item in enumerate(types):
        label = f"generation_state.shift_types[{index}]"
        item = _required_dict(item, label)
        type_id = _required_int(item.get("id"), f"{label}.id", minimum=1)
        if type_id in type_ids:
            raise ValueError("generation_state содержит повторный тип смены")
        type_ids.add(type_id)
        if not isinstance(item.get("active"), bool):
            raise ValueError(f"{label}.active: ожидается boolean")
        planned = _required_int(
            item.get("planned_duration_min"), f"{label}.planned_duration_min",
            minimum=1, maximum=7 * 24 * 60,
        )
        maximum = _required_int(
            item.get("max_duration_min"), f"{label}.max_duration_min",
            minimum=1, maximum=7 * 24 * 60,
        )
        if maximum < planned:
            raise ValueError(f"{label}: max_duration меньше planned_duration")
        slots = _required_int(
            item.get("driver_slots"), f"{label}.driver_slots", minimum=1,
            maximum=2,
        )
        if slots not in (1, 2):
            raise ValueError(f"{label}.driver_slots: допустимо 1 или 2")
    return type_ids


def _validate_plan_json_shape(plan):
    plan = _required_dict(plan, "payload_json")
    _required_int(plan.get("route_id"), "route_id", minimum=1)
    if not isinstance(plan.get("day_type"), str) or not plan["day_type"]:
        raise ValueError("day_type: ожидается непустая строка")
    if not isinstance(plan.get("preserve_locked"), bool):
        raise ValueError("preserve_locked: ожидается boolean")
    scope = _required_dict(plan.get("scope"), "scope")
    _required_int(scope.get("route_id"), "scope.route_id", minimum=1)
    if not isinstance(scope.get("day_type"), str):
        raise ValueError("scope.day_type: ожидается строка")
    if not isinstance(scope.get("username"), str) or not scope["username"]:
        raise ValueError("scope.username: ожидается непустая строка")
    conflicts = _required_list(plan.get("conflicts"), "conflicts")
    for index, conflict in enumerate(conflicts):
        _required_dict(conflict, f"conflicts[{index}]")
    diff = _required_dict(plan.get("diff"), "diff")
    for field in (
        "old_shift_count", "new_shift_count",
        "old_driver_slots", "new_driver_slots",
    ):
        _required_int(diff.get(field), f"diff.{field}")
    locked = _required_list(plan.get("locked_shifts"), "locked_shifts")
    for index, shift in enumerate(locked):
        _validate_shift_json(
            shift, f"locked_shifts[{index}]", require_locked=True
        )
    outputs = _required_list(plan.get("outputs"), "outputs")
    for output_index, output in enumerate(outputs):
        label = f"outputs[{output_index}]"
        output = _required_dict(output, label)
        _required_int(
            output.get("output_number"), f"{label}.output_number", minimum=1
        )
        output_conflicts = _required_list(
            output.get("conflicts"), f"{label}.conflicts"
        )
        for index, conflict in enumerate(output_conflicts):
            _required_dict(conflict, f"{label}.conflicts[{index}]")
        shifts = _required_list(output.get("shifts"), f"{label}.shifts")
        for shift_index, shift in enumerate(shifts):
            _validate_shift_json(shift, f"{label}.shifts[{shift_index}]")
    state_type_ids = _validate_generation_state_json(
        plan.get("generation_state")
    )
    return plan, state_type_ids


def _validated_stored_plan(con, preview, route_id, day_type):
    try:
        plan = json.loads(preview["payload_json"])
    except (TypeError, json.JSONDecodeError):
        raise ValueError("Повреждён сохранённый план смен")
    plan, state_type_ids = _validate_plan_json_shape(plan)
    if plan.get("route_id") != route_id or plan.get("day_type") != day_type:
        raise ValueError("План смен не соответствует маршруту или типу дня")
    if plan.get("conflicts"):
        raise ValueError("План смен содержит конфликты")
    scope = plan.get("scope")
    if scope and scope != {
        "route_id": route_id,
        "day_type": day_type,
        "username": preview["username"],
    }:
        raise ValueError("Scope сохранённого плана был изменён")

    preview_locked = plan.get("locked_shifts")
    if preview_locked is None:
        preview_locked = [
            shift
            for output in plan.get("outputs", [])
            for shift in output.get("shifts", [])
            if shift.get("is_manual_locked")
        ]
    try:
        preview_locked = sorted(preview_locked, key=lambda item: int(item["id"]))
    except (TypeError, KeyError, ValueError):
        raise ValueError("Повреждён снимок заблокированных смен")
    current_locked = sorted(
        (
            _locked_shift_payload(row)
            for row in db.rows(con.execute(
                "SELECT * FROM output_shifts WHERE route_id=? AND day_type=? "
                "AND is_manual_locked=1 ORDER BY id",
                (route_id, day_type),
            ))
        ),
        key=lambda item: item["id"],
    )
    if preview_locked != current_locked:
        raise LockedPlanChanged(
            "Набор заблокированных смен изменился после preview"
        )
    output_locked = sorted(
        (
            shift
            for output in plan.get("outputs", [])
            for shift in output.get("shifts", [])
            if shift.get("is_manual_locked")
        ),
        key=lambda item: int(item["id"]),
    )
    if output_locked != preview_locked:
        raise LockedPlanChanged("Заблокированные смены исключены из плана")

    trips = _generation_trips(con, route_id, day_type)
    by_output = defaultdict(list)
    for trip in trips:
        by_output[int(trip["output_number"])].append(trip)
    outputs = plan.get("outputs")
    used_type_ids = set()
    if not isinstance(outputs, list):
        raise ValueError("Повреждён сохранённый план смен")
    planned_numbers = set()
    for output in outputs:
        output_number = int(output.get("output_number"))
        if output_number in planned_numbers or output_number not in by_output:
            raise ValueError("План содержит неизвестный или повторный выпуск")
        planned_numbers.add(output_number)
        shifts = output.get("shifts")
        if not isinstance(shifts, list):
            raise ValueError("Повреждён сохранённый план смен")
        conflicts = validate_output_shift_plan(by_output[output_number], shifts)
        if conflicts:
            raise ValueError("План смен не покрывает рейсы выпуска")
        for shift in shifts:
            shift_type = _shift_type_by_id(con, shift.get("shift_type_id"))
            used_type_ids.add(int(shift["shift_type_id"]))
            if not shift_type:
                raise ValueError("План ссылается на неизвестный тип смены")
            if shift.get("is_manual_locked"):
                locked = db.one(con.execute(
                    "SELECT * FROM output_shifts WHERE id=? AND route_id=? "
                    "AND day_type=? AND output_number=? AND is_manual_locked=1",
                    (shift.get("id"), route_id, day_type, output_number),
                ))
                if not locked:
                    raise ValueError("Заблокированная смена была изменена")
                expected = _locked_shift_payload(locked)
                if any(expected[key] != shift.get(key) for key in expected):
                    raise ValueError("Заблокированная смена была изменена")
    if planned_numbers != set(by_output):
        raise ValueError("План смен покрывает не все выпуски")
    if used_type_ids != state_type_ids:
        raise ValueError("Набор типов смен в плане не соответствует снимку")
    settings = _settings_payload(con, route_id, day_type)
    current_state = _generation_state(con, settings, used_type_ids)
    if current_state != plan["generation_state"]:
        raise PreviewStateChanged(
            "Настройки или типы смен изменились после preview"
        )
    types_by_id = {
        int(item["id"]): item for item in current_state["shift_types"]
    }
    handover_sec = int(settings["handover_min"]) * 60
    for output in outputs:
        ordered_shifts = sorted(
            output["shifts"], key=lambda item: (item["start_sec"], item["end_sec"])
        )
        for shift in ordered_shifts:
            if shift.get("is_manual_locked"):
                continue
            duration = int(shift["end_sec"]) - int(shift["start_sec"])
            shift_type = types_by_id[int(shift["shift_type_id"])]
            if int(shift["driver_slots"]) != int(shift_type["driver_slots"]):
                raise ValueError(
                    "Количество водителей не соответствует типу смены"
                )
            if duration > int(shift_type["max_duration_min"]) * 60:
                raise ValueError("Длительность смены превышает максимум типа")
            if not shift_type["active"]:
                raise ValueError("План использует неактивный тип смены")
        for previous, current in zip(ordered_shifts, ordered_shifts[1:]):
            if int(current["start_sec"]) - int(previous["end_sec"]) < handover_sec:
                raise ValueError("Между сменами недостаточно времени пересмены")
    return plan, by_output


def _validate_persisted_scope(con, route_id, day_type, by_output):
    shift_rows = db.rows(con.execute(
        "SELECT * FROM output_shifts WHERE route_id=? AND day_type=? "
        "ORDER BY output_number,start_sec,id",
        (route_id, day_type),
    ))
    shifts_by_output = defaultdict(list)
    for row in shift_rows:
        shifts_by_output[int(row["output_number"])].append(dict(row))
    if set(shifts_by_output) != set(by_output):
        raise ValueError("Сохранённые смены не соответствуют выпускам")

    expected_links = {}
    for output_number, trips in by_output.items():
        shifts = shifts_by_output[output_number]
        if validate_output_shift_plan(trips, shifts):
            raise ValueError("Сохранённые смены пересекаются или не покрывают выпуск")
        positions = {int(trip["id"]): index for index, trip in enumerate(trips)}
        for shift in shifts:
            first = positions[int(shift["trip_from_id"])]
            last = positions[int(shift["trip_to_id"])]
            for trip in trips[first:last + 1]:
                trip_id = int(trip["id"])
                if trip_id in expected_links:
                    raise ValueError("Рейс связан с пересекающимися сменами")
                expected_links[trip_id] = (
                    int(shift["id"]), int(shift["shift_number"])
                )
    actual = db.rows(con.execute(
        "SELECT id,output_shift_id,shift_number FROM route_trips "
        "WHERE route_id=? AND day_type=?",
        (route_id, day_type),
    ))
    if len(actual) != len(expected_links):
        raise ValueError("Не все рейсы покрыты сохранёнными сменами")
    for trip in actual:
        if expected_links.get(int(trip["id"])) != (
            trip["output_shift_id"], int(trip["shift_number"])
        ):
            raise ValueError("Рейс связан с неправильной сменой")


@router.post("/routes/{route_id}/shift-generation/apply")
def route_shift_generation_apply(route_id: int, payload: dict = Body(...),
                                 user=Depends(current_user)):
    require_write(user, "trips")
    day_type = str(payload.get("day_type") or "").strip()
    token = str(payload.get("preview_token") or payload.get("token") or "").strip()
    if not day_type or not token:
        raise HTTPException(400, "Укажите тип дня и токен preview")
    con = db.connect()
    try:
        _route_or_404(con, route_id)
        preview = db.one(con.execute(
            "SELECT * FROM shift_generation_previews WHERE token=?",
            (token,),
        ))
        if not preview:
            raise HTTPException(400, "Preview не найден")
        if preview["applied_at"]:
            raise HTTPException(409, "Preview уже применён")
        if (
            int(preview["route_id"]) != route_id
            or preview["day_type"] != day_type
            or preview["username"] != user["username"]
        ):
            raise HTTPException(400, "Preview не соответствует маршруту, дню или пользователю")
        if datetime.datetime.fromisoformat(preview["expires_at"]) <= datetime.datetime.now():
            raise HTTPException(409, "Срок действия preview истёк")
        plan, by_output = _validated_stored_plan(
            con, preview, route_id, day_type
        )

        replaceable_ids = [
            int(row["id"]) for row in db.rows(con.execute(
                "SELECT id FROM output_shifts WHERE route_id=? AND day_type=? "
                "AND is_manual_locked=0",
                (route_id, day_type),
            ))
        ]
        con.execute(
            "UPDATE route_trips SET output_shift_id=NULL "
            "WHERE route_id=? AND day_type=?",
            (route_id, day_type),
        )
        if replaceable_ids:
            marks = ",".join("?" for _ in replaceable_ids)
            con.execute(
                f"UPDATE roster_assignments SET output_shift_id=NULL "
                f"WHERE output_shift_id IN ({marks})",
                replaceable_ids,
            )
        con.execute(
            "DELETE FROM output_shifts WHERE route_id=? AND day_type=? "
            "AND is_manual_locked=0",
            (route_id, day_type),
        )

        timestamp = datetime.datetime.now().isoformat(timespec="seconds")
        inserted = 0
        for output in plan["outputs"]:
            output_number = int(output["output_number"])
            ordered = by_output[output_number]
            positions = {int(row["id"]): index for index, row in enumerate(ordered)}
            for shift in output["shifts"]:
                if shift.get("is_manual_locked"):
                    shift_id = int(shift["id"])
                else:
                    cursor = con.execute(
                        """
                        INSERT INTO output_shifts(
                          route_id,day_type,output_number,shift_number,
                          shift_type_id,trip_from_id,trip_to_id,start_sec,end_sec,
                          driver_slots,handover_after_min,source,is_manual_locked,
                          generation_key,created_at,updated_at
                        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        (route_id, day_type, output_number,
                         int(shift["shift_number"]), int(shift["shift_type_id"]),
                         int(shift["trip_from_id"]), int(shift["trip_to_id"]),
                         int(shift["start_sec"]), int(shift["end_sec"]),
                         int(shift["driver_slots"]),
                         int(shift.get("handover_after_min", 0)), "generated", 0,
                         token, timestamp, timestamp),
                    )
                    shift_id = cursor.lastrowid
                    inserted += 1
                first = positions[int(shift["trip_from_id"])]
                last = positions[int(shift["trip_to_id"])]
                trip_ids = [int(row["id"]) for row in ordered[first:last + 1]]
                marks = ",".join("?" for _ in trip_ids)
                con.execute(
                    f"UPDATE route_trips SET shift_number=?,output_shift_id=? "
                    f"WHERE id IN ({marks})",
                    (int(shift["shift_number"]), shift_id, *trip_ids),
                )

        invalid = con.execute(
            """
            SELECT COUNT(*) FROM route_trips rt
            LEFT JOIN output_shifts os ON os.id=rt.output_shift_id
            WHERE rt.route_id=? AND rt.day_type=? AND (
              os.id IS NULL OR os.route_id<>rt.route_id OR os.day_type<>rt.day_type
              OR os.output_number<>rt.output_number OR os.shift_number<>rt.shift_number
            )
            """,
            (route_id, day_type),
        ).fetchone()[0]
        if invalid:
            raise ValueError("Не все рейсы связаны ровно с одной допустимой сменой")
        _validate_persisted_scope(con, route_id, day_type, by_output)
        con.execute(
            "UPDATE shift_generation_previews SET applied_at=? WHERE token=? "
            "AND applied_at IS NULL",
            (timestamp, token),
        )
        db.audit(
            con, user["username"], "применение плана смен", "output_shifts",
            route_id, new={"day_type": day_type, "preview_token": token,
                           "shift_count": sum(len(item["shifts"]) for item in plan["outputs"])},
        )
        con.commit()
        return {
            "route_id": route_id,
            "day_type": day_type,
            "preview_token": token,
            "shift_count": sum(len(item["shifts"]) for item in plan["outputs"]),
            "inserted_shift_count": inserted,
        }
    except HTTPException:
        con.rollback()
        raise
    except LockedPlanChanged as exc:
        con.rollback()
        raise HTTPException(409, str(exc))
    except PreviewStateChanged as exc:
        con.rollback()
        raise HTTPException(409, str(exc))
    except (AttributeError, KeyError, OverflowError, TypeError, ValueError,
            sqlite3.IntegrityError) as exc:
        con.rollback()
        raise HTTPException(400, str(exc) or "Не удалось применить план смен")
    finally:
        con.close()


def _manual_output_rows(con, route_id, day_type, output_number):
    return db.rows(con.execute(
        "SELECT * FROM output_shifts WHERE route_id=? AND day_type=? "
        "AND output_number=? ORDER BY shift_number,id",
        (route_id, day_type, output_number),
    ))


def _manual_link_trips(con, trips, shifts):
    positions = {int(trip["id"]): index for index, trip in enumerate(trips)}
    trip_ids = [int(trip["id"]) for trip in trips]
    marks = ",".join("?" for _ in trip_ids)
    con.execute(f"UPDATE route_trips SET output_shift_id=NULL WHERE id IN ({marks})",
                trip_ids)
    for shift in shifts:
        first = positions[int(shift["trip_from_id"])]
        last = positions[int(shift["trip_to_id"])]
        covered = trip_ids[first:last + 1]
        covered_marks = ",".join("?" for _ in covered)
        con.execute(
            f"UPDATE route_trips SET shift_number=?,output_shift_id=? "
            f"WHERE id IN ({covered_marks})",
            (int(shift["shift_number"]), int(shift["id"]), *covered),
        )


@router.patch("/output-shifts/{shift_id}")
def output_shift_manual_update(shift_id: int, payload: dict = Body(...),
                               user=Depends(current_user)):
    from .route_shifts import replace_shift_boundaries

    require_write(user, "trips")
    reason = str(payload.get("reason") or "").strip()
    if not reason:
        raise HTTPException(400, "Укажите причину ручного изменения")
    con = db.connect()
    try:
        con.execute("BEGIN IMMEDIATE")
        selected = db.one(con.execute("SELECT * FROM output_shifts WHERE id=?",
                                      (shift_id,)))
        if not selected:
            raise HTTPException(404, "Смена не найдена")
        shift_type = _active_type_or_error(con, payload.get("shift_type_id"),
                                           "тип смены")
        try:
            trip_from_id = int(payload.get("trip_from_id"))
            trip_to_id = int(payload.get("trip_to_id"))
        except (TypeError, ValueError):
            raise ValueError("Укажите корректные границы смены")
        trips = [trip for trip in _generation_trips(
            con, selected["route_id"], selected["day_type"]
        ) if int(trip["output_number"]) == int(selected["output_number"])]
        before = _manual_output_rows(con, selected["route_id"],
                                     selected["day_type"],
                                     selected["output_number"])
        plan = replace_shift_boundaries(
            trips, before, shift_id=shift_id, trip_from_id=trip_from_id,
            trip_to_id=trip_to_id, shift_type=shift_type,
        )
        settings = _settings_payload(
            con, selected["route_id"], selected["day_type"]
        )
        _apply_manual_handover(plan, settings["handover_min"])
        edited = next(row for row in plan if int(row["id"]) == shift_id)
        if (int(edited["end_sec"]) - int(edited["start_sec"]) >
                int(shift_type["max_duration_min"]) * 60):
            raise ValueError("Длительность смены превышает максимум выбранного типа")
        _validate_changed_shift_constraints(con, before, plan)
        timestamp = datetime.datetime.now().isoformat(timespec="seconds")
        temporary_base = min(
            [0, *(int(row["shift_number"]) for row in before)]
        ) - len(before) - 1
        if temporary_base - len(before) < -(2 ** 63):
            raise ValueError("Невозможно безопасно перенумеровать смены")
        for index, row in enumerate(before):
            con.execute(
                "UPDATE output_shifts SET shift_number=? WHERE id=?",
                (temporary_base - index, int(row["id"])),
            )
        for row in plan:
            manual = int(row["id"]) == shift_id
            con.execute(
                "UPDATE output_shifts SET shift_number=?,shift_type_id=?,"
                "trip_from_id=?,trip_to_id=?,start_sec=?,end_sec=?,driver_slots=?,"
                "handover_after_min=?,source=?,is_manual_locked=?,manual_reason=?,"
                "updated_at=? WHERE id=?",
                (row["shift_number"], row["shift_type_id"], row["trip_from_id"],
                 row["trip_to_id"], row["start_sec"], row["end_sec"],
                 row["driver_slots"], row.get("handover_after_min", 0),
                 "manual" if manual else row.get("source", "generated"),
                 1 if manual else int(row.get("is_manual_locked", 0)),
                 reason if manual else row.get("manual_reason"), timestamp, row["id"]),
            )
        persisted = _manual_output_rows(con, selected["route_id"],
                                        selected["day_type"],
                                        selected["output_number"])
        _manual_link_trips(con, trips, persisted)
        if validate_output_shift_plan(trips, persisted):
            raise ValueError("Изменённый план не покрывает выпуск без конфликтов")
        db.audit(con, user["username"], "ручное изменение границ смены",
                 "output_shifts", shift_id, old=before, new=persisted)
        con.commit()
        return {"shift": next(row for row in persisted if row["id"] == shift_id),
                "shifts": persisted}
    except sqlite3.OperationalError as exc:
        con.rollback()
        if _sqlite_busy(exc):
            raise HTTPException(409, "База данных занята, повторите операцию")
        raise
    except HTTPException:
        con.rollback()
        raise
    except (KeyError, TypeError, ValueError, sqlite3.IntegrityError) as exc:
        con.rollback()
        raise HTTPException(400, str(exc) or "Не удалось изменить смену")
    finally:
        con.close()


def _manual_regenerate_output(con, route_id, day_type, output_number,
                              settings, timestamp):
    trips = [trip for trip in _generation_trips(con, route_id, day_type)
             if int(trip["output_number"]) == output_number]
    if not trips:
        raise ValueError("В выбранном выпуске нет рейсов")
    shift_type = _require_active_generated_type(_shift_type_for_output(settings, trips))
    plan = build_output_shifts(trips, shift_type=shift_type,
                               handover_min=settings["handover_min"])
    old = _manual_output_rows(con, route_id, day_type, output_number)
    old_ids = [row["id"] for row in old]
    if old_ids:
        marks = ",".join("?" for _ in old_ids)
        con.execute(f"UPDATE roster_assignments SET output_shift_id=NULL "
                    f"WHERE output_shift_id IN ({marks})", old_ids)
    con.execute("UPDATE route_trips SET output_shift_id=NULL WHERE route_id=? "
                "AND day_type=? AND output_number=?", (route_id, day_type, output_number))
    con.execute("DELETE FROM output_shifts WHERE route_id=? AND day_type=? "
                "AND output_number=?", (route_id, day_type, output_number))
    inserted = []
    for shift in plan:
        cursor = con.execute(
            "INSERT INTO output_shifts(route_id,day_type,output_number,shift_number,"
            "shift_type_id,trip_from_id,trip_to_id,start_sec,end_sec,driver_slots,"
            "handover_after_min,source,is_manual_locked,manual_reason,created_at,updated_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,'generated',0,NULL,?,?)",
            (route_id, day_type, output_number, shift["shift_number"],
             shift["shift_type_id"], shift["trip_from_id"], shift["trip_to_id"],
             shift["start_sec"], shift["end_sec"], shift["driver_slots"],
             shift.get("handover_after_min", 0), timestamp, timestamp),
        )
        inserted.append({**shift, "id": cursor.lastrowid})
    _manual_link_trips(con, trips, inserted)
    return old, _manual_output_rows(con, route_id, day_type, output_number)


@router.post("/routes/{route_id}/output-shifts/reset-manual")
def output_shifts_reset_manual(route_id: int, payload: dict = Body(...),
                               user=Depends(current_user)):
    require_write(user, "trips")
    day_type = str(payload.get("day_type") or "").strip()
    if not day_type:
        raise HTTPException(400, "Укажите тип дня")
    shift_id = payload.get("shift_id")
    output_number = payload.get("output_number")
    if shift_id not in (None, "") and output_number not in (None, ""):
        raise HTTPException(400, "Укажите только shift_id или output_number")
    con = db.connect()
    try:
        con.execute("BEGIN IMMEDIATE")
        _route_or_404(con, route_id)
        if shift_id not in (None, ""):
            try:
                shift_id = int(shift_id)
            except (TypeError, ValueError):
                raise ValueError("Некорректный shift_id")
            target = db.one(con.execute("SELECT output_number FROM output_shifts "
                "WHERE id=? AND route_id=? AND day_type=?", (shift_id, route_id, day_type)))
            if not target:
                raise HTTPException(404, "Смена не найдена в выбранной области")
            outputs = [int(target["output_number"])]
        elif output_number not in (None, ""):
            try:
                output_number = int(output_number)
            except (TypeError, ValueError):
                raise ValueError("Некорректный номер выпуска")
            exists = con.execute("SELECT 1 FROM route_trips WHERE route_id=? AND "
                "day_type=? AND output_number=?", (route_id, day_type, output_number)).fetchone()
            if not exists:
                raise HTTPException(404, "Выпуск не найден в выбранной области")
            outputs = [output_number]
        else:
            outputs = [int(row[0]) for row in con.execute("SELECT DISTINCT output_number "
                "FROM route_trips WHERE route_id=? AND day_type=? ORDER BY output_number",
                (route_id, day_type))]
            if not outputs:
                raise HTTPException(404, "Для выбранного дня нет выпусков")
        settings = _settings_payload(con, route_id, day_type)
        timestamp = datetime.datetime.now().isoformat(timespec="seconds")
        before, after = [], []
        for number in outputs:
            old, new = _manual_regenerate_output(con, route_id, day_type, number,
                                                  settings, timestamp)
            before.extend(old); after.extend(new)
        db.audit(con, user["username"], "сброс ручных смен", "output_shifts",
                 route_id, old=before, new=after)
        con.commit()
        return {"route_id": route_id, "day_type": day_type,
                "output_numbers": outputs, "shifts": after}
    except sqlite3.OperationalError as exc:
        con.rollback()
        if _sqlite_busy(exc):
            raise HTTPException(409, "База данных занята, повторите операцию")
        raise
    except HTTPException:
        con.rollback(); raise
    except (KeyError, TypeError, ValueError, sqlite3.IntegrityError) as exc:
        con.rollback()
        raise HTTPException(400, str(exc) or "Не удалось сбросить ручные смены")
    finally:
        con.close()


def _validate_changed_shift_constraints(con, before, proposed):
    before_by_id = {int(row["id"]): row for row in before}
    constrained_fields = (
        "shift_number", "shift_type_id", "trip_from_id", "trip_to_id",
        "start_sec", "end_sec", "driver_slots",
    )
    for shift in proposed:
        previous = before_by_id.get(int(shift["id"]))
        if previous and all(
            previous.get(field) == shift.get(field) for field in constrained_fields
        ):
            continue
        shift_type = _shift_type_by_id(con, shift.get("shift_type_id"))
        if not shift_type:
            raise ValueError("Тип изменённой смены не найден")
        if not shift_type["active"]:
            raise ValueError("Тип изменённой смены должен быть активным")
        try:
            slots = int(shift["driver_slots"])
            expected_slots = int(shift_type["driver_slots"])
            duration = int(shift["end_sec"]) - int(shift["start_sec"])
            maximum = int(shift_type["max_duration_min"]) * 60
        except (KeyError, TypeError, ValueError):
            raise ValueError("Некорректные параметры изменённой смены")
        if slots not in (1, 2) or slots != expected_slots:
            raise ValueError(
                "Количество водительских мест не соответствует типу смены"
            )
        if duration <= 0:
            raise ValueError("Длительность смены должна быть больше нуля")
        if duration > maximum:
            raise ValueError(
                "Длительность изменённой смены превышает максимум её типа"
            )


def _sqlite_busy(exc):
    message = str(exc).lower()
    return "locked" in message or "busy" in message


def _apply_manual_handover(plan, handover_min):
    required_seconds = int(handover_min) * 60
    for previous, current in zip(plan, plan[1:]):
        gap_seconds = int(current["start_sec"]) - int(previous["end_sec"])
        if gap_seconds < required_seconds:
            raise ValueError("Недостаточное время пересмены между сменами")
        previous["handover_after_min"] = gap_seconds // 60
    if plan:
        plan[-1]["handover_after_min"] = 0
