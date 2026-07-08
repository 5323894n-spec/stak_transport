# -*- coding: utf-8 -*-
"""Import helpers for ERM route workbooks."""
import datetime as _dt
import io
import re

from openpyxl import load_workbook


class ErmImportError(ValueError):
    """Raised when an ERM workbook cannot be parsed."""


_ROUTE_RE = re.compile(r"Маршрут\s*№\s*([^\s\"«»]+)\s*[\"«](.*?)[\"»]", re.IGNORECASE)


def _norm(value):
    return str(value or "").strip()


def _cell(row, idx):
    return row[idx] if idx < len(row) else None


def _to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _to_int(value):
    f = _to_float(value)
    return int(f) if f is not None else None


def _time_to_minutes(value):
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        value = value.time()
    if isinstance(value, _dt.time):
        return value.hour * 60 + value.minute + round(value.second / 60)
    if isinstance(value, _dt.timedelta):
        return round(value.total_seconds() / 60)
    if isinstance(value, (int, float)):
        return round(value * 24 * 60) if 0 <= value <= 1 else round(value)
    text = _norm(value)
    parts = text.split(":")
    if len(parts) >= 2:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(float(parts[2])) if len(parts) > 2 else 0
            return hours * 60 + minutes + round(seconds / 60)
        except ValueError:
            return None
    return _to_int(value)


def _time_text(value):
    if value is None:
        return ""
    if isinstance(value, _dt.datetime):
        value = value.time()
    if isinstance(value, _dt.time):
        return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"
    if isinstance(value, _dt.timedelta):
        total = int(value.total_seconds())
        return f"{total // 3600:02d}:{(total % 3600) // 60:02d}:{total % 60:02d}"
    return _norm(value)


def _route_title(workbook):
    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
            for value in row:
                text = _norm(value)
                match = _ROUTE_RE.search(text)
                if match:
                    return text, match.group(1).strip(), match.group(2).strip()
    raise ErmImportError("Не найден заголовок маршрута вида: Маршрут № ...")


def _direction_marker(row):
    text = " ".join(_norm(v).lower() for v in row if v is not None)
    if "прямое направление" in text:
        return "прямое"
    if "обратное направление" in text:
        return "обратное"
    return None


def _sheet_kind(title):
    title_l = title.lower()
    if title_l == "из парка":
        return "из парка"
    if title_l == "в парк":
        return "в парк"
    return "маршрут"


def _default_direction(sheet_title, sections_count):
    title_l = sheet_title.lower()
    if title_l == "из парка":
        return "из парка"
    if title_l == "в парк":
        return "в парк"
    return "прямое" if sections_count == 0 else "обратное"


def _is_header(row):
    return _norm(_cell(row, 1)).lower() == "п.п." and "останов" in _norm(_cell(row, 5)).lower()


def _is_blank(row):
    return not any(_norm(v) for v in row)


def _parse_stop(row, row_number):
    distance_m = _to_float(_cell(row, 9))
    cumulative_m = _to_float(_cell(row, 10))
    distance_km = _to_float(_cell(row, 14))
    cumulative_km = _to_float(_cell(row, 15))
    if distance_km is None and distance_m is not None:
        distance_km = distance_m / 1000
    if cumulative_km is None and cumulative_m is not None:
        cumulative_km = cumulative_m / 1000
    return {
        "row": row_number,
        "seq": _to_int(_cell(row, 1)),
        "stop_id": _to_int(_cell(row, 2)),
        "line_name": _norm(_cell(row, 3)),
        "direction_label": _norm(_cell(row, 4)),
        "stop_name": _norm(_cell(row, 5)),
        "street": _norm(_cell(row, 6)),
        "lat": _to_float(_cell(row, 7)),
        "lon": _to_float(_cell(row, 8)),
        "distance_m": distance_m,
        "cumulative_m": cumulative_m,
        "distance_km": round(distance_km, 6) if distance_km is not None else None,
        "cumulative_km": round(cumulative_km, 6) if cumulative_km is not None else None,
        "travel_time": _time_text(_cell(row, 11)),
        "travel_time_min": _time_to_minutes(_cell(row, 11)),
        "cumulative_time": _time_text(_cell(row, 12)),
        "cumulative_time_min": _time_to_minutes(_cell(row, 12)),
        "mismatch": _to_int(_cell(row, 13)),
    }


def _looks_like_stop(row):
    return _to_int(_cell(row, 1)) is not None and bool(_norm(_cell(row, 5)))


def _parse_sheet(ws):
    sections = []
    current = None
    current_direction = None
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        marker = _direction_marker(row)
        if marker:
            if current and current["stops"]:
                current = None
            current_direction = marker
        if _is_header(row):
            current = {
                "sheet": ws.title,
                "kind": _sheet_kind(ws.title),
                "direction": current_direction or _default_direction(ws.title, len(sections)),
                "header_row": row_number,
                "stops": [],
            }
            sections.append(current)
            continue
        if current is None:
            continue
        if _is_blank(row):
            if current["stops"]:
                current = None
            continue
        if _looks_like_stop(row):
            current["stops"].append(_parse_stop(row, row_number))
    return [s for s in sections if s["stops"]]


def _last_km(stops):
    vals = [s.get("cumulative_km") for s in stops if s.get("cumulative_km") is not None]
    if vals:
        return round(vals[-1], 3)
    return round(sum(s.get("distance_km") or 0 for s in stops), 3)


def _last_minutes(stops):
    vals = [s.get("cumulative_time_min") for s in stops if s.get("cumulative_time_min") is not None]
    return vals[-1] if vals else None


def _stop_names(stops):
    return [s["stop_name"] for s in stops if s.get("stop_name")]


def _pick_route_sections(sections):
    forward = next((s for s in sections if s["direction"] == "прямое"), None)
    backward = next((s for s in sections if s["direction"] == "обратное"), None)
    if not forward and sections:
        forward = sections[0]
    if not backward and len(sections) > 1:
        backward = sections[1]
    return forward, backward


def parse_erm_route_workbook(data):
    """Parse an ERM workbook and return route fields plus detailed sections."""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ErmImportError("Не удалось прочитать файл. Нужен формат .xlsx") from exc
    if "параметры" not in wb.sheetnames:
        raise ErmImportError("В файле не найден лист 'параметры'")

    title, number, name = _route_title(wb)
    sheets = {}
    all_sections = []
    for ws in wb.worksheets:
        sections = _parse_sheet(ws)
        sheets[ws.title] = {"sections": sections, "rows": ws.max_row}
        all_sections.extend(sections)

    route_sections = sheets["параметры"]["sections"]
    forward, backward = _pick_route_sections(route_sections)
    if not forward:
        raise ErmImportError("На листе 'параметры' не найдены остановки маршрута")

    forward_stops = forward["stops"]
    backward_stops = backward["stops"] if backward else []
    forward_names = _stop_names(forward_stops)
    backward_names = _stop_names(backward_stops)
    depot_sections = [s for s in all_sections if s["kind"] in ("из парка", "в парк")]

    details = {
        "title": title,
        "number": number,
        "name": name,
        "sheets": sheets,
        "summary": {
            "sections": len(all_sections),
            "route_stops_forward": len(forward_stops),
            "route_stops_backward": len(backward_stops),
            "depot_sections": len(depot_sections),
            "depot_stops": sum(len(s["stops"]) for s in depot_sections),
        },
    }

    return {
        "number": number,
        "name": name,
        "comm_type": "городское",
        "transport_type": "Регулярные перевозки пассажиров и багажа",
        "start_point": forward_names[0] if forward_names else "",
        "end_point": forward_names[-1] if forward_names else "",
        "stops": ", ".join(forward_names),
        "stops_back": ", ".join(backward_names),
        "length_km": _last_km(forward_stops),
        "length_back_km": _last_km(backward_stops) if backward_stops else None,
        "trip_time_min": _last_minutes(forward_stops),
        "trip_time_back_min": _last_minutes(backward_stops) if backward_stops else None,
        "details": details,
    }
