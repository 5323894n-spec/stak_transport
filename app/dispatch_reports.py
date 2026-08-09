# -*- coding: utf-8 -*-
"""Excel-отчёт диспетчерской сводки за день."""
from openpyxl import Workbook

from .route_document_xlsx import apply_sheet_setup, write_table_header, write_title_band
from . import dispatch_service as ds


def build_dispatch_report(con, date):
    board = ds.build_board(con, date)
    summary = ds.day_summary(con, date)
    wb = Workbook()

    release = wb.active
    release.title = "Выпуск"
    apply_sheet_setup(release)
    write_title_band(release, 1, f"ДИСПЕТЧЕРСКАЯ СВОДКА — ВЫПУСК {date}", end_col=8)
    write_table_header(release, 2, (
        "Маршрут", "Выход", "Смена", "Водитель", "Автобус",
        "План", "Факт", "Откл., мин", "Статус",
    ))
    row = 3
    for item in board["rows"]:
        values = (
            item["route_number"], item["output_number"], item["shift_number"],
            item["driver_fio"], item["garage_number"], item["plan_release"],
            item["actual_release"], item["deviation_min"], item["status"],
        )
        for col, value in enumerate(values, 1):
            release.cell(row, col, value)
        row += 1
    release.cell(row + 1, 1, (
        f"План выходов: {summary['planned']} · выпущено: {summary['released']} · "
        f"на линии: {summary['on_line']} · сходы: {summary['off_line']} · "
        f"срывы: {summary['disrupted']} · регулярность выпуска: {summary['release_regularity']}%"
    ))

    adherence = wb.create_sheet("Регулярность")
    apply_sheet_setup(adherence)
    write_title_band(adherence, 1, f"РЕГУЛЯРНОСТЬ ДВИЖЕНИЯ {date}", end_col=6)
    write_table_header(adherence, 2, (
        "Маршрут", "Выход", "Рейс", "План", "Факт", "Откл., мин",
    ))
    row = 3
    for item in board["rows"]:
        for fact in ds.list_trip_facts(con, date, item["order_line_id"]):
            for col, value in enumerate((
                item["route_number"], item["output_number"], fact["trip_number"],
                fact["plan_dep"], fact["actual_dep"], fact["deviation_min"],
            ), 1):
                adherence.cell(row, col, value)
            row += 1
    adherence.cell(row + 1, 1, f"Регулярность рейсов: {summary['trip_regularity']}%")
    return wb


def dispatch_report_filename(date):
    return f"Диспетчер_{date}.xlsx"
